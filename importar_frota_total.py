import sqlite3
from pathlib import Path
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit(
        "Falta instalar openpyxl. Executa primeiro:\n"
        "python -m pip install openpyxl"
    )

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "stock.db"
EXCEL_PATH = BASE_DIR / "frota 0805_v2.xlsx"
SHEET_NAME = "Vehicles"

def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None

def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None

def clean_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None

def excel_date_to_iso(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return text

def status_to_estado(current_status, status):
    cs = (current_status or "").strip().upper()
    st = str(status or "").strip()

    if "SOLD" in cs or "RETURNED" in cs or st == "4":
        return "Vendida", 0

    if "IMPRO" in cs or st == "3":
        return "Ativa", 1

    if "FREE" in cs or "SHORT" in cs or "MID" in cs or st in ("1", "2"):
        return "Ativa", 1

    return "Ativa", 1

def column_exists(conn, table, column):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r[1] == column for r in rows)

def ensure_columns(conn):
    columns = {
        "estado_frota": "TEXT DEFAULT 'Ativa'",
        "data_venda": "TEXT",
        "valor_venda": "REAL",
        "comprador": "TEXT",
        "motivo_venda": "TEXT",
        "data_baixa": "TEXT",
        "ativa_operacional": "INTEGER DEFAULT 1",
        "current_status_rentway": "TEXT",
        "status_rentway": "TEXT",
        "fleet": "TEXT",
        "rental_station": "TEXT",
        "grupo": "TEXT",
        "categoria": "TEXT",
        "fornecedor": "TEXT",
        "valor_compra": "REAL",
        "valor_compra_com_iva": "REAL",
        "valor_venda_com_iva": "REAL",
        "documento_venda": "TEXT",
        "data_fatura_venda": "TEXT",
    }

    for col, sql_type in columns.items():
        if not column_exists(conn, "viaturas", col):
            conn.execute(f"ALTER TABLE viaturas ADD COLUMN {col} {sql_type}")

def main():
    if not DB_PATH.exists():
        raise SystemExit(f"Não encontrei a base de dados: {DB_PATH}")

    if not EXCEL_PATH.exists():
        raise SystemExit(f"Não encontrei o Excel: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Não encontrei a folha {SHEET_NAME}. Folhas: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    required = ["platenr", "chassinr", "brandid", "modelid", "version", "fuel", "kms", "purchase_date", "last_service", "next_service", "observations", "Year", "CurrentStatus", "status"]
    missing = [h for h in required if h not in col]
    if missing:
        raise SystemExit(f"Faltam colunas no Excel: {missing}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    ensure_columns(conn)

    created = 0
    updated = 0
    skipped = 0
    sold = 0
    active = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        matricula = clean_text(row[col["platenr"]])
        if not matricula:
            skipped += 1
            continue

        matricula = matricula.upper().replace(" ", "")

        current_status = clean_text(row[col["CurrentStatus"]]) if "CurrentStatus" in col else None
        status = clean_text(row[col["status"]]) if "status" in col else None
        estado_frota, ativa_operacional = status_to_estado(current_status, status)

        if estado_frota == "Vendida":
            sold += 1
        else:
            active += 1

        sales_date = excel_date_to_iso(row[col["sales_date"]]) if "sales_date" in col else None
        invoice_date = excel_date_to_iso(row[col["invoice_date"]]) if "invoice_date" in col else None
        data_venda = sales_date or invoice_date

        valor_venda = None
        for candidate in ["sales_value_with_tax", "invoice_value_with_tax", "finantial_sale_value_with_tax", "sales_value"]:
            if candidate in col:
                valor_venda = clean_float(row[col[candidate]])
                if valor_venda not in (None, 0):
                    break

        existing = conn.execute("SELECT id FROM viaturas WHERE matricula = ?", (matricula,)).fetchone()

        values = {
            "matricula": matricula,
            "vin": clean_text(row[col["chassinr"]]),
            "marca": clean_text(row[col["brandid"]]),
            "modelo": clean_text(row[col["modelid"]]),
            "versao": clean_text(row[col["version"]]),
            "motorizacao": clean_text(row[col["version"]]),
            "combustivel": clean_text(row[col["fuel"]]),
            "caixa": None,
            "ano": clean_int(row[col["Year"]]),
            "data_compra": excel_date_to_iso(row[col["purchase_date"]]),
            "km_atual": clean_int(row[col["kms"]]),
            "proxima_revisao_km": None,
            "proxima_revisao_data": excel_date_to_iso(row[col["next_service"]]),
            "ultima_revisao": excel_date_to_iso(row[col["last_service"]]),
            "campanhas_pendentes": 0,
            "risco_tecnico": "Normal",
            "observacoes": clean_text(row[col["observations"]]),
            "ativo": 1,
            "estado_frota": estado_frota,
            "data_venda": data_venda,
            "valor_venda": valor_venda,
            "comprador": clean_text(row[col["Client"]]) if "Client" in col else None,
            "motivo_venda": None,
            "data_baixa": None,
            "ativa_operacional": ativa_operacional,
            "current_status_rentway": current_status,
            "status_rentway": status,
            "fleet": clean_text(row[col["fleet"]]) if "fleet" in col else None,
            "rental_station": clean_text(row[col["rental_station"]]) if "rental_station" in col else None,
            "grupo": clean_text(row[col["groupid"]]) if "groupid" in col else None,
            "categoria": clean_text(row[col["category"]]) if "category" in col else None,
            "fornecedor": clean_text(row[col["supplier_name"]]) if "supplier_name" in col else None,
            "valor_compra": clean_float(row[col["value"]]) if "value" in col else None,
            "valor_compra_com_iva": clean_float(row[col["value_with_tax"]]) if "value_with_tax" in col else None,
            "valor_venda_com_iva": valor_venda,
            "documento_venda": clean_text(row[col["invoice_nr"]]) if "invoice_nr" in col else None,
            "data_fatura_venda": invoice_date,
        }

        if existing:
            conn.execute("""
                UPDATE viaturas SET
                    vin=?, marca=?, modelo=?, versao=?, motorizacao=?, combustivel=?, caixa=?, ano=?,
                    data_compra=?, km_atual=?, proxima_revisao_km=?, proxima_revisao_data=?,
                    ultima_revisao=?, campanhas_pendentes=?, risco_tecnico=?, observacoes=?, ativo=?,
                    estado_frota=?, data_venda=?, valor_venda=?, comprador=?, motivo_venda=?, data_baixa=?,
                    ativa_operacional=?, current_status_rentway=?, status_rentway=?, fleet=?, rental_station=?,
                    grupo=?, categoria=?, fornecedor=?, valor_compra=?, valor_compra_com_iva=?,
                    valor_venda_com_iva=?, documento_venda=?, data_fatura_venda=?
                WHERE matricula=?
            """, (
                values["vin"], values["marca"], values["modelo"], values["versao"], values["motorizacao"],
                values["combustivel"], values["caixa"], values["ano"], values["data_compra"], values["km_atual"],
                values["proxima_revisao_km"], values["proxima_revisao_data"], values["ultima_revisao"],
                values["campanhas_pendentes"], values["risco_tecnico"], values["observacoes"], values["ativo"],
                values["estado_frota"], values["data_venda"], values["valor_venda"], values["comprador"],
                values["motivo_venda"], values["data_baixa"], values["ativa_operacional"],
                values["current_status_rentway"], values["status_rentway"], values["fleet"], values["rental_station"],
                values["grupo"], values["categoria"], values["fornecedor"], values["valor_compra"],
                values["valor_compra_com_iva"], values["valor_venda_com_iva"], values["documento_venda"],
                values["data_fatura_venda"], matricula
            ))
            updated += 1
        else:
            conn.execute("""
                INSERT INTO viaturas (
                    matricula, vin, marca, modelo, versao, motorizacao, combustivel, caixa, ano,
                    data_compra, km_atual, proxima_revisao_km, proxima_revisao_data, ultima_revisao,
                    campanhas_pendentes, risco_tecnico, observacoes, ativo, criado_em,
                    estado_frota, data_venda, valor_venda, comprador, motivo_venda, data_baixa,
                    ativa_operacional, current_status_rentway, status_rentway, fleet, rental_station,
                    grupo, categoria, fornecedor, valor_compra, valor_compra_com_iva,
                    valor_venda_com_iva, documento_venda, data_fatura_venda
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                values["matricula"], values["vin"], values["marca"], values["modelo"], values["versao"],
                values["motorizacao"], values["combustivel"], values["caixa"], values["ano"], values["data_compra"],
                values["km_atual"], values["proxima_revisao_km"], values["proxima_revisao_data"],
                values["ultima_revisao"], values["campanhas_pendentes"], values["risco_tecnico"],
                values["observacoes"], values["ativo"], datetime.now().isoformat(timespec="seconds"),
                values["estado_frota"], values["data_venda"], values["valor_venda"], values["comprador"],
                values["motivo_venda"], values["data_baixa"], values["ativa_operacional"],
                values["current_status_rentway"], values["status_rentway"], values["fleet"], values["rental_station"],
                values["grupo"], values["categoria"], values["fornecedor"], values["valor_compra"],
                values["valor_compra_com_iva"], values["valor_venda_com_iva"], values["documento_venda"],
                values["data_fatura_venda"]
            ))
            created += 1

    conn.commit()
    conn.close()

    print("Importação frota total concluída.")
    print(f"Criadas: {created}")
    print(f"Atualizadas: {updated}")
    print(f"Vendidas/históricas: {sold}")
    print(f"Ativas/operacionais: {active}")
    print(f"Ignoradas sem matrícula: {skipped}")

if __name__ == "__main__":
    main()
