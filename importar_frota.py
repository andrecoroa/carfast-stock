import sqlite3
from pathlib import Path
from datetime import datetime, timedelta

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit(
        "Falta instalar openpyxl. Executa primeiro:\n"
        "python -m pip install openpyxl\n"
    )

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "stock.db"
EXCEL_PATH = BASE_DIR / "frota0805.xlsx"
SHEET_NAME = "Vehicles"


def excel_date_to_iso(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, (int, float)):
        # Excel serial date, Windows 1900 date system
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    return text


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def init_viaturas_table(conn):
    conn.execute("""
    CREATE TABLE IF NOT EXISTS viaturas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        matricula TEXT UNIQUE NOT NULL,
        vin TEXT,
        marca TEXT,
        modelo TEXT,
        versao TEXT,
        motorizacao TEXT,
        combustivel TEXT,
        caixa TEXT,
        ano INTEGER,
        data_compra TEXT,
        km_atual INTEGER,
        proxima_revisao_km INTEGER,
        proxima_revisao_data TEXT,
        ultima_revisao TEXT,
        campanhas_pendentes INTEGER DEFAULT 0,
        risco_tecnico TEXT DEFAULT 'Normal',
        observacoes TEXT,
        ativo INTEGER DEFAULT 1,
        criado_em TEXT NOT NULL
    )
    """)


def main():
    if not EXCEL_PATH.exists():
        raise SystemExit(f"Não encontrei o ficheiro: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Não encontrei a folha '{SHEET_NAME}'. Folhas existentes: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    required = ["platenr", "chassinr", "brandid", "modelid", "version", "fuel", "kms", "purchase_date", "last_service", "next_service", "observations", "Year"]
    missing = [h for h in required if h not in col]
    if missing:
        raise SystemExit(f"Faltam colunas no Excel: {missing}")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    init_viaturas_table(conn)

    created = 0
    updated = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        matricula = clean_text(row[col["platenr"]])
        if not matricula:
            skipped += 1
            continue

        matricula = matricula.upper().replace(" ", "")

        data = {
            "matricula": matricula,
            "vin": clean_text(row[col["chassinr"]]),
            "marca": clean_text(row[col["brandid"]]),
            "modelo": clean_text(row[col["modelid"]]),
            "versao": clean_text(row[col["version"]]),
            "motorizacao": clean_text(row[col["version"]]),  # provisório: a versão tem normalmente motor + versão comercial
            "combustivel": clean_text(row[col["fuel"]]),
            "caixa": None,
            "ano": clean_int(row[col["Year"]]),
            "data_compra": excel_date_to_iso(row[col["purchase_date"]]),
            "km_atual": clean_int(row[col["kms"]]),
            "proxima_revisao_km": None,
            "proxima_revisao_data": excel_date_to_iso(row[col["next_service"]]),
            "ultima_revisao": excel_date_to_iso(row[col["last_service"]]),
            "campanhas_pendentes": 0,
            "risco_tecnico": "Normal",
            "observacoes": clean_text(row[col["observations"]]),
            "ativo": 1,
            "criado_em": datetime.now().isoformat(timespec="seconds"),
        }

        existing = conn.execute("SELECT id FROM viaturas WHERE matricula = ?", (matricula,)).fetchone()

        if existing:
            conn.execute("""
                UPDATE viaturas SET
                    vin = ?,
                    marca = ?,
                    modelo = ?,
                    versao = ?,
                    motorizacao = ?,
                    combustivel = ?,
                    caixa = ?,
                    ano = ?,
                    data_compra = ?,
                    km_atual = ?,
                    proxima_revisao_km = ?,
                    proxima_revisao_data = ?,
                    ultima_revisao = ?,
                    campanhas_pendentes = ?,
                    risco_tecnico = ?,
                    observacoes = ?,
                    ativo = ?
                WHERE matricula = ?
            """, (
                data["vin"],
                data["marca"],
                data["modelo"],
                data["versao"],
                data["motorizacao"],
                data["combustivel"],
                data["caixa"],
                data["ano"],
                data["data_compra"],
                data["km_atual"],
                data["proxima_revisao_km"],
                data["proxima_revisao_data"],
                data["ultima_revisao"],
                data["campanhas_pendentes"],
                data["risco_tecnico"],
                data["observacoes"],
                data["ativo"],
                matricula,
            ))
            updated += 1
        else:
            conn.execute("""
                INSERT INTO viaturas (
                    matricula, vin, marca, modelo, versao, motorizacao, combustivel, caixa, ano,
                    data_compra, km_atual, proxima_revisao_km, proxima_revisao_data, ultima_revisao,
                    campanhas_pendentes, risco_tecnico, observacoes, ativo, criado_em
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                data["matricula"],
                data["vin"],
                data["marca"],
                data["modelo"],
                data["versao"],
                data["motorizacao"],
                data["combustivel"],
                data["caixa"],
                data["ano"],
                data["data_compra"],
                data["km_atual"],
                data["proxima_revisao_km"],
                data["proxima_revisao_data"],
                data["ultima_revisao"],
                data["campanhas_pendentes"],
                data["risco_tecnico"],
                data["observacoes"],
                data["ativo"],
                data["criado_em"],
            ))
            created += 1

    conn.commit()
    conn.close()

    print("Importação concluída.")
    print(f"Criadas: {created}")
    print(f"Atualizadas: {updated}")
    print(f"Ignoradas sem matrícula: {skipped}")


if __name__ == "__main__":
    main()
