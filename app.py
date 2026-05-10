from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, g
import sqlite3
import json
import hashlib
import unicodedata
import calendar
import re
import shutil
import traceback
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    load_workbook = None
    Workbook = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "stock.db"
UPLOAD_DIR = APP_DIR / "uploads"
IMPORT_DIR = UPLOAD_DIR / "imports"
UPLOAD_DIR.mkdir(exist_ok=True)
IMPORT_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = "carfast-v2"

USER_ROLES = ["Admin", "Gestor", "Operador", "Consulta"]
DEFAULT_ADMIN_EMAIL = "admin@carfast.local"
DEFAULT_ADMIN_PASSWORD = "CarFast123!"

ESTADOS = [
    "Aberto",
    "Receção / ServiceBox",
    "Diagnóstico Mecânico",
    "Pendente Validação",
    "FO Criada",
    "Em Execução",
    "Diagnóstico Pós",
    "Concluído pelo Mecânico",
    "Validação Administrativa",
    "Fechado",
]

PROTOCOLO_FASES_INICIAIS = [
    ("abertura", "Abertura Processo", 1, 1, 1),
    ("rececao_servicebox", "Receção / ServiceBox", 2, 1, 1),
    ("diagnostico_mecanico", "Diagnóstico Mecânico", 3, 1, 1),
    ("pendente_validacao", "Pendente Validação", 4, 1, 1),
    ("criacao_fo", "Criação FO", 5, 1, 1),
    ("gestao_tarefas_fo", "Gestão Tarefas FO", 6, 1, 0),
    ("em_execucao", "Em Execução", 7, 1, 0),
    ("diagnostico_pos", "Diagnóstico Pós", 8, 1, 1),
    ("comparacao_diagnosticos", "Comparação Diagnósticos", 9, 0, 0),
    ("validacao_administrativa", "Validação Administrativa", 10, 1, 1),
]

PROTOCOLO_CAMPOS_INICIAIS = [
    ("numero_impro", "Nº Impro", "abertura", "Identificação", "Texto", 1, 0, "Documento operacional de origem."),
    ("matricula", "Matrícula", "abertura", "Identificação", "Texto", 1, 0, "Matrícula da viatura."),
    ("km_abertura", "KM abertura", "abertura", "Identificação", "Número", 1, 0, "Quilometragem no início."),
    ("responsavel", "Responsável", "abertura", "Identificação", "Texto", 0, 0, "Responsável pelo processo."),
    ("descricao_inicial", "Descrição inicial", "abertura", "Identificação", "Observação", 1, 0, "Problema inicial."),
    ("servicebox_verificado", "ServiceBox verificado", "rececao_servicebox", "OEM / ServiceBox", "Sim/Não", 1, 0, ""),
    ("campanhas_verificadas", "Campanhas verificadas", "rececao_servicebox", "OEM / ServiceBox", "Sim/Não", 1, 0, ""),
    ("campanhas_pendentes", "Campanhas pendentes", "rececao_servicebox", "OEM / ServiceBox", "Sim/Não", 0, 0, ""),
    ("plano_manutencao_verificado", "Plano manutenção verificado", "rececao_servicebox", "OEM / ServiceBox", "Lista", 1, 0, ""),
    ("historico_oem_validado", "Histórico OEM validado", "rececao_servicebox", "OEM / ServiceBox", "Sim/Não", 0, 0, ""),
    ("pdf_servicebox", "PDF ServiceBox", "rececao_servicebox", "Evidência", "PDF", 0, 1, ""),
    ("fotos_servicebox", "Fotos ServiceBox", "rececao_servicebox", "Evidência", "Foto", 0, 1, ""),
    ("diagnosticos_oem", "Diagnósticos OEM", "rececao_servicebox", "Evidência", "Documento", 0, 1, ""),
    ("campanhas_anexo", "Documento campanhas", "rececao_servicebox", "Evidência", "Documento", 0, 1, ""),
    ("n_manutencoes_ecu", "Nº manutenções ECU", "diagnostico_mecanico", "ECU", "Texto", 0, 0, ""),
    ("intervencoes_reais", "Intervenções reais", "diagnostico_mecanico", "Diagnóstico", "Texto", 0, 0, ""),
    ("telecarregamento", "Telecarregamento", "diagnostico_mecanico", "ECU", "Texto", 0, 0, ""),
    ("ultima_manutencao_registada", "Última manutenção registada", "diagnostico_mecanico", "Manutenção", "Texto", 0, 0, ""),
    ("manutencao_nao_registada", "Manutenção não registada", "diagnostico_mecanico", "Manutenção", "Sim/Não", 0, 0, ""),
    ("km_manutencao_anterior", "KM manutenção anterior", "diagnostico_mecanico", "Manutenção", "Número", 0, 0, ""),
    ("km_ate_manutencao", "KM até manutenção", "diagnostico_mecanico", "Manutenção", "Número", 0, 0, ""),
    ("diluicao_oleo", "Diluição óleo", "diagnostico_mecanico", "ECU", "Percentagem", 0, 0, ""),
    ("carbono_oleo", "Carbono óleo", "diagnostico_mecanico", "ECU", "Percentagem", 0, 0, ""),
    ("problema_identificado", "Problema identificado", "diagnostico_mecanico", "Diagnóstico", "Observação", 1, 0, ""),
    ("causa_provavel", "Causa provável", "diagnostico_mecanico", "Diagnóstico", "Observação", 0, 0, ""),
    ("intervencao_recomendada", "Intervenção recomendada", "diagnostico_mecanico", "Diagnóstico", "Observação", 1, 0, ""),
    ("oficina_recomendada", "Oficina recomendada", "diagnostico_mecanico", "Execução", "Lista", 0, 0, ""),
    ("pode_circular", "Pode circular", "diagnostico_mecanico", "Auditoria", "Lista", 0, 0, ""),
    ("prioridade", "Prioridade", "diagnostico_mecanico", "Auditoria", "Lista", 0, 0, ""),
    ("inconsistencias", "Inconsistências", "diagnostico_mecanico", "Auditoria", "Observação", 0, 0, ""),
    ("autoscan_pdf", "PDF AutoScan", "diagnostico_mecanico", "Evidência", "PDF", 0, 1, ""),
    ("fotos_ecu", "Fotos ECU", "diagnostico_mecanico", "Evidência", "Foto", 0, 1, ""),
    ("orcamento_validado", "Orçamento validado", "pendente_validacao", "Validação", "Sim/Não", 1, 0, ""),
    ("criticidade_validada", "Criticidade validada", "pendente_validacao", "Validação", "Sim/Não", 1, 0, ""),
    ("autorizacao_execucao", "Autorização execução", "pendente_validacao", "Validação", "Sim/Não", 1, 0, ""),
    ("encaminhamento", "Encaminhamento", "pendente_validacao", "Execução", "Lista", 0, 0, ""),
    ("motivo_decisao", "Motivo decisão", "pendente_validacao", "Validação", "Observação", 0, 0, ""),
    ("numero_fo", "Nº FO", "criacao_fo", "Execução", "Texto", 1, 0, ""),
    ("responsavel_fo", "Responsável FO", "criacao_fo", "Execução", "Texto", 0, 0, ""),
    ("observacoes_fo", "Observações FO", "criacao_fo", "Execução", "Observação", 0, 0, ""),
    ("descricao_tarefa", "Descrição tarefa", "gestao_tarefas_fo", "Execução", "Texto", 1, 0, ""),
    ("tipo_servico", "Tipo serviço", "gestao_tarefas_fo", "Execução", "Lista", 0, 0, ""),
    ("responsavel_tarefa", "Responsável tarefa", "gestao_tarefas_fo", "Execução", "Texto", 0, 0, ""),
    ("tarefa_autorizada", "Tarefa autorizada", "gestao_tarefas_fo", "Execução", "Sim/Não", 1, 0, ""),
    ("estado_tarefa", "Estado tarefa", "gestao_tarefas_fo", "Execução", "Lista", 1, 0, ""),
    ("observacoes_tarefa", "Observações tarefa", "gestao_tarefas_fo", "Execução", "Observação", 0, 0, ""),
    ("execucao_iniciada", "Execução iniciada", "em_execucao", "Execução", "Sim/Não", 1, 0, ""),
    ("evidencia_execucao", "Evidência execução", "em_execucao", "Evidência", "Documento", 0, 1, ""),
    ("observacoes_execucao", "Observações execução", "em_execucao", "Execução", "Observação", 0, 0, ""),
    ("diagnostico_pos_realizado", "Diagnóstico pós realizado", "diagnostico_pos", "Diagnóstico", "Sim/Não", 1, 0, ""),
    ("autoscan_pos_pdf", "PDF AutoScan pós", "diagnostico_pos", "Evidência", "PDF", 1, 1, ""),
    ("diluicao_oleo_pos", "Diluição óleo pós", "diagnostico_pos", "ECU", "Percentagem", 0, 0, ""),
    ("carbono_oleo_pos", "Carbono óleo pós", "diagnostico_pos", "ECU", "Percentagem", 0, 0, ""),
    ("erros_pos", "Erros pós-reparação", "diagnostico_pos", "Diagnóstico", "Observação", 0, 0, ""),
    ("parametros_pos", "Parâmetros pós-reparação", "diagnostico_pos", "Diagnóstico", "Observação", 0, 0, ""),
    ("comparacao_km", "Comparação KM", "comparacao_diagnosticos", "Auditoria", "Número", 0, 0, ""),
    ("comparacao_diluicao", "Comparação diluição óleo", "comparacao_diagnosticos", "Auditoria", "Percentagem", 0, 0, ""),
    ("comparacao_carbono", "Comparação carbono óleo", "comparacao_diagnosticos", "Auditoria", "Percentagem", 0, 0, ""),
    ("comparacao_ecu", "Comparação ECU", "comparacao_diagnosticos", "Auditoria", "Texto", 0, 0, ""),
    ("alerta_reincidencia", "Alerta reincidência", "comparacao_diagnosticos", "Auditoria", "Sim/Não", 0, 0, ""),
    ("conclusao_comparacao", "Conclusão comparação", "comparacao_diagnosticos", "Auditoria", "Observação", 0, 0, ""),
    ("fo_fechada", "FO fechada", "validacao_administrativa", "Validação", "Sim/Não", 1, 0, ""),
    ("diagnostico_pos_anexado", "Diagnóstico pós anexado", "validacao_administrativa", "Validação", "Sim/Não", 1, 0, ""),
    ("ficha_viatura_atualizada", "Ficha viatura atualizada", "validacao_administrativa", "Validação", "Sim/Não", 1, 0, ""),
    ("parametros_atualizados", "Parâmetros atualizados", "validacao_administrativa", "Validação", "Sim/Não", 1, 0, ""),
    ("documentacao_completa", "Documentação completa", "validacao_administrativa", "Validação", "Sim/Não", 1, 0, ""),
    ("validado_por", "Validado por", "validacao_administrativa", "Validação", "Texto", 1, 0, ""),
    ("observacoes_validacao", "Observações validação", "validacao_administrativa", "Validação", "Observação", 0, 0, ""),
]

INCIDENTES_CATEGORIAS_INICIAIS = [
    "Oficina / Marca",
    "Clientes",
    "Outros",
]

INCIDENTES_CLASSIFICACOES_INICIAIS = {
    "Oficina / Marca": [
        "Diagnóstico errado",
        "Reparação incompleta",
        "Avaria reincidente",
        "Campanha não executada",
        "Peça errada",
        "Demora reparação",
        "Garantia recusada",
        "Telecarregamento em falta",
        "Atualização ECU em falta",
        "Falta documentação",
        "Erro técnico",
        "Relatório inconsistente",
        "Orçamento excessivo",
        "Serviço não autorizado",
    ],
    "Clientes": [
        "Dano interior",
        "Dano exterior",
        "Mau uso",
        "Falta combustível",
        "Falta AdBlue",
        "Embraiagem",
        "Pneu",
        "Jante",
        "Chave",
        "Fumador",
        "Limpeza extrema",
        "Sinistro não comunicado",
        "Acessórios em falta",
        "Uso indevido",
        "Dano detetado em manutenção",
    ],
    "Outros": [
        "Erro operacional",
        "Preparação incompleta",
        "Documentação em falta",
        "Limpeza incompleta",
        "Disponibilidade incorreta",
        "Transferência incorreta",
        "Parque",
        "IPO",
        "Transporte",
        "Bloqueio venda",
        "Anomalia diversa",
    ],
}

INCIDENTE_ESTADOS = [
    "Aberto",
    "Em análise",
    "Comunicado",
    "A aguardar resposta",
    "Resolvido",
    "Fechado",
    "Cancelado",
]

INCIDENTE_GRAVIDADES = ["Baixa", "Média", "Alta", "Crítica"]
INCIDENTE_FATURAVEL = ["Sim", "Não", "A avaliar"]
TAREFA_FROTA_ESTADOS = ["Pendente", "Em curso", "Bloqueada", "Concluída", "Validada", "Cancelada"]
VENDA_USADOS_ESTADOS = ["Selecionada", "Em análise", "Aprovada", "Exportada", "Vendida", "Cancelada"]
CENTRO_TAREFAS_CATEGORIAS = ["Tarefas", "Pedidos / Dúvidas", "Incidentes / Reportes", "Sugestões"]
CENTRO_TAREFAS_PRIORIDADES = ["Urgente", "Alta", "Média", "Baixa"]
CENTRO_TAREFAS_ESTADOS = {
    "Tarefas": ["Novo", "Em execução", "Concluído"],
    "Pedidos / Dúvidas": ["Novo", "Respondido", "Fechado"],
    "Incidentes / Reportes": ["Novo", "Em análise", "Resolvido"],
    "Sugestões": ["Nova", "Em análise", "Aceite", "Rejeitada"],
}
CENTRO_TAREFAS_DEPARTAMENTOS = [
    "Direção",
    "Frota",
    "Manutenção",
    "Stock",
    "Operações",
    "Comercial",
    "Financeiro",
    "Aeroporto",
    "Administrativo",
    "IT",
]

DOCUMENT_ROOT_LABEL = "CarFast Documentos Operacionais"

CONHECIMENTO_AREAS = [
    "Comunicação e tarefas",
    "Oficina",
    "Frota",
    "Gestão",
    "Administração e configurações",
    "Como usar a app",
]

CONHECIMENTO_SEEDS = [
    {
        "titulo": "Como abrir um novo processo de oficina",
        "slug": "novo-processo-oficina",
        "area": "Oficina",
        "tema": "Processos",
        "resumo": "Procedimento rápido para abrir uma manutenção a partir de Nº Impro, matrícula, KM e descrição inicial.",
        "conteudo": """Objetivo
Abrir um processo de oficina para acompanhar uma situação técnica da viatura.

Quando usar
- Viatura com impro aberto
- Avaria ou manutenção a acompanhar
- Necessidade de diagnóstico, FO, validação ou follow-up

Passos
1. Confirmar Nº Impro.
2. Confirmar matrícula.
3. Registar KM de abertura sempre que possível.
4. Indicar responsável pelo acompanhamento.
5. Escrever uma descrição inicial clara do problema.
6. Abrir o processo.

Boas práticas
- Não abrir processos duplicados para o mesmo impro sem motivo.
- Usar descrição objetiva: sintoma, contexto, urgência e restrições.
- Depois de abrir, avançar para receção/ServiceBox ou diagnóstico conforme o caso.""",
    },
    {
        "titulo": "Importar frota Rentway",
        "slug": "importar-frota-rentway",
        "area": "Frota",
        "tema": "Importações Rentway",
        "resumo": "Como atualizar a frota integral a partir do relatório Rentway sem manipular o ficheiro.",
        "conteudo": """Regra principal
O ficheiro de frota deve ser importado integralmente, sem manipulação prévia.

Passos
1. Extrair o relatório de frota no Rentway.
2. Guardar o ficheiro original.
3. Abrir Importar Frota Rentway.
4. Indicar responsável.
5. Selecionar o ficheiro Excel.
6. Importar.

O que verificar depois
- Nº de linhas importadas.
- Viaturas criadas e atualizadas.
- Viaturas vendidas/ativas.
- Erros ou linhas ignoradas.

Nota
A importação não deve apagar histórico, processos, incidentes, faturas ou diagnósticos já associados à viatura.""",
    },
    {
        "titulo": "Usar o Centro de Tarefas",
        "slug": "usar-centro-tarefas",
        "area": "Comunicação e tarefas",
        "tema": "Registos internos",
        "resumo": "Como criar tarefas, pedidos, incidentes/reportes e sugestões internas.",
        "conteudo": """Objetivo
Centralizar trabalho interno, pedidos, dúvidas, reportes e sugestões.

Categorias
- Tarefas: algo que exige execução.
- Pedidos / Dúvidas: pedido de apoio, informação ou decisão.
- Incidentes / Reportes: problema ou situação relevante.
- Sugestões: melhoria ou ideia.

Boas práticas
- Escrever assunto curto e claro.
- Escolher o departamento certo.
- Definir prioridade realista.
- Usar matrícula, cliente ou estação quando ajudar o contexto.
- Comentar no registo em vez de dispersar informação.""",
    },
    {
        "titulo": "Ficha da viatura",
        "slug": "ficha-viatura",
        "area": "Frota",
        "tema": "Viaturas",
        "resumo": "Como interpretar a ficha da viatura como entidade permanente e histórico completo.",
        "conteudo": """Princípio
A viatura é uma entidade permanente. Mesmo vendida, mantém histórico.

Usar a ficha para
- Ver identificação e estados.
- Consultar processos, incidentes e tarefas.
- Ver histórico Rentway.
- Consultar faturas, FO, contratos, ARs, sinistros e diagnósticos.

Boas práticas
- Não apagar viaturas.
- Separar estado lifecycle de estado operacional.
- Usar histórico e timeline antes de tirar conclusões sobre uma viatura.""",
    },
    {
        "titulo": "Incidentes Frota",
        "slug": "incidentes-frota",
        "area": "Frota",
        "tema": "Incidentes",
        "resumo": "Como registar incidentes ligados a viaturas e gerar follow-up.",
        "conteudo": """Conceito
Incidente é o que aconteceu. Tarefa é o que vamos fazer.

Quando criar
- Problemas com oficina ou marca.
- Danos ou mau uso de cliente.
- Problemas documentais.
- Preparação incompleta.
- Situações relevantes para auditoria futura.

Boas práticas
- Ligar sempre à matrícula correta.
- Escolher categoria e classificação com cuidado.
- Indicar ação necessária.
- Anexar evidências quando existirem.""",
    },
    {
        "titulo": "Venda de usados",
        "slug": "venda-usados",
        "area": "Frota",
        "tema": "Venda",
        "resumo": "Como selecionar viaturas da frota para venda e exportar listagem para comércio.",
        "conteudo": """Objetivo
Criar uma camada comercial sobre viaturas já existentes na frota.

Fluxo
1. Filtrar frota.
2. Selecionar candidatas.
3. Rever selecionadas.
4. Ajustar preço sugerido, mínimo e comércio.
5. Exportar listagem.

Notas
O valor atual teórico é contabilístico. Não é o preço de venda.""",
    },
    {
        "titulo": "Faturas de fornecedores",
        "slug": "faturas-fornecedores",
        "area": "Gestão",
        "tema": "Custos",
        "resumo": "Como consultar faturas, linhas, serviços detetados e alertas.",
        "conteudo": """Objetivo
Auditar custos por fornecedor, documento, matrícula e tipo de serviço.

Como ler
- Cada documento aparece fechado.
- Expandir só quando for preciso ver linhas.
- Usar badges de serviços para identificar óleo, pneus, travões, discos e outros sinais.
- Verificar alertas de validação.

Boas práticas
- Auditar por documento antes de analisar linha a linha.
- Cruzar com FO, impros e diagnóstico quando houver custo anormal.""",
    },
    {
        "titulo": "Timeline da viatura",
        "slug": "timeline-viatura",
        "area": "Frota",
        "tema": "Auditoria",
        "resumo": "Como usar a timeline mensal para entender contratos, impros, documentos e diagnósticos.",
        "conteudo": """Objetivo
Ver a vida operacional da viatura por mês/dia.

Linhas previstas
- Contrato / impro / livre.
- Documentos operacionais.
- Documentos financeiros.
- Diagnósticos.

Como usar
Passar pelos dias para ver detalhe e abrir documentos associados quando existir link.""",
    },
    {
        "titulo": "Stock de oficina",
        "slug": "stock-oficina",
        "area": "Oficina",
        "tema": "Stock",
        "resumo": "Como usar stock, artigos e movimentos no contexto da oficina.",
        "conteudo": """Objetivo
Controlar peças, consumíveis e movimentos de apoio à oficina.

Usar para
- Criar artigos.
- Consultar stock atual.
- Registar entradas, saídas e ajustes.
- Associar movimentos a viatura quando fizer sentido.

Boas práticas
- Usar descrições consistentes.
- Registar responsável.
- Não misturar stock operacional com histórico financeiro de faturas.""",
    },
]

CONTEXT_HELP_BY_ENDPOINT = {
    "novo_processo": "novo-processo-oficina",
    "processos": "novo-processo-oficina",
    "processo_detail": "novo-processo-oficina",
    "importar_frota_view": "importar-frota-rentway",
    "centro_tarefas": "usar-centro-tarefas",
    "centro_tarefas_novo": "usar-centro-tarefas",
    "centro_tarefa_detail": "usar-centro-tarefas",
    "viaturas": "ficha-viatura",
    "viatura_detail": "ficha-viatura",
    "incidentes_frota": "incidentes-frota",
    "novo_incidente_frota": "incidentes-frota",
    "incidente_frota_detail": "incidentes-frota",
    "venda_usados": "venda-usados",
    "venda_usados_selecionadas": "venda-usados",
    "faturas_fornecedores": "faturas-fornecedores",
    "auditoria_viatura_timeline": "timeline-viatura",
    "stock_home": "stock-oficina",
    "index": "stock-oficina",
    "novo_artigo": "stock-oficina",
    "novo_movimento": "stock-oficina",
    "movimentos": "stock-oficina",
}

LIFECYCLE_COLUMNS = {
    "rentway_unitnr": "TEXT",
    "estado_frota": "TEXT DEFAULT 'Ativa'",
    "data_venda": "TEXT",
    "valor_venda": "REAL",
    "valor_venda_sem_iva": "REAL",
    "iva_venda": "REAL",
    "valor_iva_venda": "REAL",
    "comprador": "TEXT",
    "motivo_venda": "TEXT",
    "data_baixa": "TEXT",
    "ativa_operacional": "INTEGER DEFAULT 1",
    "estado_operacional": "TEXT",
    "current_status_rentway": "TEXT",
    "status_rentway": "TEXT",
    "fleet": "TEXT",
    "rental_station": "TEXT",
    "grupo": "TEXT",
    "categoria": "TEXT",
    "cor": "TEXT",
    "data_matricula": "TEXT",
    "lugares": "INTEGER",
    "km_compra": "INTEGER",
    "ultima_revisao_km": "INTEGER",
    "fornecedor": "TEXT",
    "valor_compra": "REAL",
    "iva_compra": "REAL",
    "valor_iva_compra": "REAL",
    "valor_compra_com_iva": "REAL",
    "valor_venda_com_iva": "REAL",
    "preco_venda_rentway": "REAL",
    "preco_venda_calculado": "REAL",
    "preco_venda_retalho_rentway": "REAL",
    "preco_venda_retalho_calculado": "REAL",
    "documento_rentway": "TEXT",
    "documento_venda": "TEXT",
    "data_fatura_venda": "TEXT",
    "data_retorno": "TEXT",
    "bloqueio_rentway": "TEXT",
    "data_bloqueio": "TEXT",
    "cliente_rentway": "TEXT",
    "programa_venda": "TEXT",
    "data_inspecao": "TEXT",
    "ald_nr_pagamentos": "INTEGER",
    "ald_primeiro_pagamento": "TEXT",
    "ald_valor_residual": "REAL",
    "depreciacao_codigo_fiscal_rentway": "TEXT",
    "depreciacao_codigo_fiscal_calculado": "TEXT",
    "depreciacao_ultima_calculo_rentway": "TEXT",
    "depreciacao_ultima_calculo_calculado": "TEXT",
    "codigo_conta_rentway": "TEXT",
    "codigo_conta_calculado": "TEXT",
    "data_ultimo_movimento": "TEXT",
    "data_bloqueio_peritagem": "TEXT",
    "fornecedor_financeiro": "TEXT",
    "contrato_financeiro": "TEXT",
    "contrato_financeiro_fim": "TEXT",
    "ultimo_servico_feito": "TEXT",
    "garantia_fim": "TEXT",
    "furto_data_declaracao": "TEXT",
    "furto_numero_declaracao": "TEXT",
    "aviso_viatura": "TEXT",
    "documentnr_rentway": "TEXT",
    "cliente_atual_rentway": "TEXT",
    "condutor_atual_rentway": "TEXT",
}

IMPRO_COLUMNS = {
    "marca": "TEXT",
    "modelo": "TEXT",
    "grupo": "TEXT",
    "condutor": "TEXT",
    "estacao_out": "TEXT",
    "data_in": "TEXT",
    "data_out": "TEXT",
    "driven_kms": "INTEGER",
    "oficina": "TEXT",
}

FO_RENTWAY_COLUMNS = {
    "rentway_key": "TEXT UNIQUE NOT NULL",
    "importacao_id": "INTEGER NOT NULL",
    "row_number": "INTEGER NOT NULL",
    "viatura_id": "INTEGER",
    "folha_obra_origem": "TEXT",
    "data_documento": "TEXT",
    "matricula": "TEXT",
    "modelo": "TEXT",
    "fornecedor": "TEXT",
    "numero_fatura": "TEXT",
    "estacao": "TEXT",
    "kms": "REAL",
    "tipo_servico": "TEXT",
    "natureza": "TEXT",
    "quantidade": "REAL",
    "valor_com_iva": "REAL",
    "descricao": "TEXT",
    "numero_documento": "TEXT",
    "estado_importacao": "TEXT",
    "alertas": "TEXT",
    "raw_json": "TEXT NOT NULL",
    "criado_em": "TEXT NOT NULL",
}

SINISTRO_ALLIANZ_COLUMNS = {
    "allianz_key": "TEXT UNIQUE NOT NULL",
    "importacao_id": "INTEGER NOT NULL",
    "row_number": "INTEGER NOT NULL",
    "viatura_id": "INTEGER",
    "adesao": "TEXT",
    "matricula": "TEXT",
    "referencia_sinistro": "TEXT",
    "data_sinistro": "TEXT",
    "data_encerramento": "TEXT",
    "causa": "TEXT",
    "valor_rc": "REAL",
    "valor_dp": "REAL",
    "valor_ids_credor": "REAL",
    "valor_vidros": "REAL",
    "custos_gestao": "REAL",
    "custo_total": "REAL",
    "estado_importacao": "TEXT",
    "alertas": "TEXT",
    "raw_json": "TEXT NOT NULL",
    "criado_em": "TEXT NOT NULL",
}

AR_RENTWAY_COLUMNS = {
    "ar_key": "TEXT UNIQUE NOT NULL",
    "importacao_id": "INTEGER NOT NULL",
    "row_number": "INTEGER NOT NULL",
    "viatura_id": "INTEGER",
    "accident_report_id": "TEXT",
    "document_type": "TEXT",
    "document_no": "TEXT",
    "state": "TEXT",
    "request_date": "TEXT",
    "accident_date": "TEXT",
    "friendly_declaration": "TEXT",
    "matricula": "TEXT",
    "unit_no": "TEXT",
    "insurance_policy": "TEXT",
    "rental_station_out": "TEXT",
    "created_by_rental_station": "TEXT",
    "status": "TEXT",
    "estado_importacao": "TEXT",
    "alertas": "TEXT",
    "raw_json": "TEXT NOT NULL",
    "criado_em": "TEXT NOT NULL",
}

FATURAS_FORNECEDORES_COLUMNS = {
    "linha_key": "TEXT UNIQUE NOT NULL",
    "importacao_id": "INTEGER NOT NULL",
    "row_number": "INTEGER NOT NULL",
    "viatura_id": "INTEGER",
    "fornecedor": "TEXT",
    "documento": "TEXT",
    "data_doc": "TEXT",
    "matricula": "TEXT",
    "kms": "REAL",
    "or_reparacao": "TEXT",
    "chassis": "TEXT",
    "descricao": "TEXT",
    "referencia": "TEXT",
    "pv_unit": "REAL",
    "desc_pct": "REAL",
    "pliq_unit": "REAL",
    "quantidade": "REAL",
    "total_liq": "REAL",
    "ct": "TEXT",
    "fonte_pdf": "TEXT",
    "estado_importacao": "TEXT",
    "alertas": "TEXT",
    "raw_json": "TEXT NOT NULL",
    "criado_em": "TEXT NOT NULL",
}


def now():
    return datetime.now().isoformat(timespec="seconds")


def db():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def column_exists(conn, table, column):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r[1] == column for r in rows)


def table_exists(conn, table):
    row = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table,),
    ).fetchone()
    return row is not None


def ensure_columns(conn, table, columns):
    if not table_exists(conn, table):
        return
    for col, sql_type in columns.items():
        if not column_exists(conn, table, col):
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {sql_type}")


def current_user_id():
    return session.get("user_id")


def log_action(acao, entidade=None, entidade_id=None, detalhe=None, conn=None):
    user_id = current_user_id()
    params = (user_id, acao, entidade, str(entidade_id) if entidade_id is not None else None, detalhe, now())
    if conn is not None:
        conn.execute("""
            INSERT INTO audit_log (user_id, acao, entidade, entidade_id, detalhe, criado_em)
            VALUES (?, ?, ?, ?, ?, ?)
        """, params)
        return
    with db() as c:
        c.execute("""
            INSERT INTO audit_log (user_id, acao, entidade, entidade_id, detalhe, criado_em)
            VALUES (?, ?, ?, ?, ?, ?)
        """, params)


def text_slug(value):
    value = (value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[^a-z0-9]+", "_", value).strip("_")
    return value or "geral"


def sync_departments(conn):
    for nome in CENTRO_TAREFAS_DEPARTAMENTOS:
        conn.execute("""
            INSERT OR IGNORE INTO departments (nome, descricao, created_at)
            VALUES (?, ?, ?)
        """, (nome, f"Departamento {nome}", now()))
        conn.execute("UPDATE departments SET ativo = 1 WHERE nome = ?", (nome,))
    conn.execute("""
        UPDATE users
        SET department_id = (
            SELECT id FROM departments WHERE departments.nome = users.departamento
        )
        WHERE department_id IS NULL
          AND departamento IS NOT NULL
          AND departamento <> ''
          AND EXISTS (SELECT 1 FROM departments WHERE departments.nome = users.departamento)
    """)


def record_task_history(conn, tarefa_id, campo, valor_antigo=None, valor_novo=None):
    if valor_antigo == valor_novo:
        return
    conn.execute("""
        INSERT INTO task_history
        (task_id, user_id, campo_alterado, valor_antigo, valor_novo, changed_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        tarefa_id,
        session.get("user_id"),
        campo,
        "" if valor_antigo is None else str(valor_antigo),
        "" if valor_novo is None else str(valor_novo),
        now(),
    ))


def document_storage_path(area, categoria, entity_id, filename):
    year = datetime.now().strftime("%Y")
    area_slug = text_slug(area)
    categoria_slug = text_slug(categoria)
    entity_slug = text_slug(str(entity_id))
    return f"{DOCUMENT_ROOT_LABEL}/{area_slug}/{year}/{categoria_slug}/{entity_slug}/{filename}"


def register_operational_document(
    conn,
    area,
    entidade,
    entidade_id,
    categoria,
    filename,
    original_name,
    uploaded_by=None,
    file_size=None,
    file_type=None,
    sharepoint_url=None,
):
    storage_path = document_storage_path(area, categoria, entidade_id, filename)
    cur = conn.execute("""
        INSERT INTO documentos_operacionais
        (area, entidade, entidade_id, categoria, file_name, original_name, file_type,
         file_size, storage_provider, storage_path, sharepoint_url, uploaded_by, uploaded_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        area,
        entidade,
        str(entidade_id),
        categoria,
        filename,
        original_name,
        file_type,
        file_size,
        "local",
        storage_path,
        sharepoint_url,
        uploaded_by,
        now(),
    ))
    return cur.lastrowid, storage_path


def create_import_batch(conn, source_system, entity_type, file_name=None, original_name=None, status="Em processamento"):
    cur = conn.execute("""
        INSERT INTO import_batches
        (source_system, entity_type, file_name, original_name, imported_by, imported_at, status)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        source_system,
        entity_type,
        file_name,
        original_name,
        session.get("user_id"),
        now(),
        status,
    ))
    return cur.lastrowid


def finish_import_batch(conn, batch_id, total_rows=0, created_rows=0, updated_rows=0, skipped_rows=0, error_rows=0, status="Concluído", detalhe=None):
    if not batch_id:
        return
    conn.execute("""
        UPDATE import_batches
        SET total_rows = ?,
            created_rows = ?,
            updated_rows = ?,
            skipped_rows = ?,
            error_rows = ?,
            status = ?,
            detalhe = ?
        WHERE id = ?
    """, (total_rows, created_rows, updated_rows, skipped_rows, error_rows, status, detalhe, batch_id))


def register_import_error(conn, batch_id, row_number, entity_type, error_message, raw_data=None):
    conn.execute("""
        INSERT INTO import_errors
        (import_batch_id, row_number, entity_type, error_message, raw_data, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        batch_id,
        row_number,
        entity_type,
        error_message,
        json.dumps(raw_data, ensure_ascii=False, default=str) if isinstance(raw_data, (dict, list)) else raw_data,
        now(),
    ))


def get_conhecimento_by_slug(conn, slug):
    if not slug:
        return None
    return conn.execute("""
        SELECT *
        FROM conhecimento_artigos
        WHERE slug = ? AND estado = 'Publicado'
        LIMIT 1
    """, (slug,)).fetchone()


def context_help_for_endpoint():
    if not session.get("user_id"):
        return None
    slug = CONTEXT_HELP_BY_ENDPOINT.get(request.endpoint)
    if not slug:
        return None
    try:
        with db() as c:
            return get_conhecimento_by_slug(c, slug)
    except sqlite3.Error:
        return None


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if session.get("role") != "Admin":
            flash("Acesso reservado a Admin.", "error")
            return redirect(url_for("home"))
        return view(*args, **kwargs)
    return wrapped


def gestor_required(view):
    @wraps(view)
    @login_required
    def wrapped(*args, **kwargs):
        if session.get("role") not in ("Admin", "Gestor"):
            flash("Acesso reservado a Admin ou Gestor.", "error")
            return redirect(url_for("home"))
        return view(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_user():
    return {
        "current_user": {
            "id": session.get("user_id"),
            "nome": session.get("user_name"),
            "role": session.get("role"),
            "departamento": session.get("departamento"),
        } if session.get("user_id") else None,
        "context_help": context_help_for_endpoint(),
    }


@app.before_request
def enforce_authentication():
    endpoint = request.endpoint or ""
    public_endpoints = {"login", "static"}
    if endpoint in public_endpoints:
        return None
    if not session.get("user_id"):
        return redirect(url_for("login", next=request.full_path if request.query_string else request.path))
    if request.method == "POST" and session.get("role") == "Consulta":
        flash("O perfil Consulta apenas permite leitura.", "error")
        return redirect(request.referrer or url_for("home"))
    if request.method == "POST" and session.get("role") == "Operador":
        blocked = (
            endpoint.startswith("importar_")
            or endpoint in {
                "validacao",
                "atualizar_incidente_frota_estado",
                "atualizar_tarefa_frota_estado",
                "nova_classificacao_incidente",
                "venda_usados_exportar",
            }
        )
        if blocked:
            flash("O perfil Operador não pode validar, importar ou gerir estrutura.", "error")
            return redirect(request.referrer or url_for("home"))
    return None


def clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def clean_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def clean_float(value):
    if value in (None, ""):
        return None
    try:
        if isinstance(value, str):
            value = value.strip().replace(" ", "")
            if "," in value and "." in value:
                value = value.replace(".", "").replace(",", ".")
            else:
                value = value.replace(",", ".")
        return float(value)
    except Exception:
        return None


def excel_date_to_iso(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception:
            return None

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def parse_excel_datetime(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="minutes")
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("#"):
        return None
    text = " ".join(text.split())
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%y %H:%M",
        "%d/%m/%y %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.isoformat(timespec="minutes") if dt.time().isoformat() != "00:00:00" else dt.date().isoformat()
        except ValueError:
            pass
    return excel_date_to_iso(value)


def clean_number_text(value):
    if value in (None, ""):
        return None
    return str(value).replace("\xa0", "").replace(" ", "").strip()


def rentway_status_to_states(current_status, status):
    cs = (current_status or "").strip().upper()
    st = str(status or "").strip()

    if "SOLD" in cs or "RETURNED" in cs or st == "4":
        return "Vendida", "Vendida", 0
    if "IMPRO" in cs or st == "3":
        return "Ativa", "Em impro", 1
    if "FREE" in cs:
        return "Ativa", "Livre", 1
    if "SHORT" in cs or "MID" in cs or "RENT" in cs or st in ("1", "2"):
        return "Ativa", "Em contrato", 1

    return "Ativa", current_status or None, 1


def row_value(row, col, name):
    idx = col.get(name)
    return row[idx] if idx is not None and idx < len(row) else None


def normalize_header(name):
    text = unicodedata.normalize("NFKD", str(name or "").lower())
    return "".join(ch for ch in text if ch.isalnum() and not unicodedata.combining(ch))


def build_column_lookup(headers):
    lookup = {}
    for idx, name in enumerate(headers):
        lookup[name] = idx
        lookup[normalize_header(name)] = idx
    return lookup


def first_row_value(row, col, candidates):
    for name in candidates:
        value = row_value(row, col, name)
        if value not in (None, ""):
            return value
        value = row_value(row, col, normalize_header(name))
        if value not in (None, ""):
            return value
    return None


def read_spreadsheet_rows(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".xls":
        if xlrd is None:
            raise RuntimeError("Falta instalar xlrd para ler ficheiros .xls.")
        wb = xlrd.open_workbook(path)
        sheet = wb.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            values = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        if not cell.value:
                            values.append(None)
                        else:
                            values.append(datetime(*xlrd.xldate_as_tuple(cell.value, wb.datemode)))
                    except Exception:
                        values.append(None)
                else:
                    values.append(cell.value)
            rows.append(values)
        return sheet.name, rows

    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return ws.title, [list(row) for row in ws.iter_rows(values_only=True)]


def build_vehicle_values(row, col):
    matricula = clean_text(row_value(row, col, "platenr"))
    if not matricula:
        return None

    matricula = matricula.upper().replace(" ", "")
    current_status = clean_text(row_value(row, col, "CurrentStatus"))
    status = clean_text(row_value(row, col, "status"))
    estado_frota, estado_operacional, ativa_operacional = rentway_status_to_states(current_status, status)
    sales_date = excel_date_to_iso(row_value(row, col, "sales_date"))
    invoice_date = excel_date_to_iso(row_value(row, col, "invoice_date"))

    valor_venda = clean_float(row_value(row, col, "sales_value_with_tax"))
    if valor_venda in (None, 0):
        valor_venda = clean_float(row_value(row, col, "sales_value"))

    return {
        "rentway_unitnr": clean_text(row_value(row, col, "unitnr")),
        "matricula": matricula,
        "vin": clean_text(row_value(row, col, "chassinr")),
        "marca": clean_text(row_value(row, col, "brandid")),
        "modelo": clean_text(row_value(row, col, "modelid")),
        "versao": clean_text(row_value(row, col, "version")),
        "motorizacao": clean_text(row_value(row, col, "version")),
        "combustivel": clean_text(row_value(row, col, "fuel")),
        "caixa": None,
        "ano": clean_int(row_value(row, col, "Year")),
        "cor": clean_text(row_value(row, col, "colour")),
        "data_matricula": excel_date_to_iso(row_value(row, col, "plate_date")),
        "lugares": clean_int(row_value(row, col, "seats")),
        "data_compra": excel_date_to_iso(row_value(row, col, "purchase_date")),
        "km_atual": clean_int(row_value(row, col, "kms")),
        "km_compra": clean_int(row_value(row, col, "kms_buy")),
        "proxima_revisao_km": clean_int(row_value(row, col, "next_service")),
        "proxima_revisao_data": None,
        "ultima_revisao": excel_date_to_iso(row_value(row, col, "last_service_done")),
        "ultima_revisao_km": clean_int(row_value(row, col, "last_service")),
        "ultimo_servico_feito": excel_date_to_iso(row_value(row, col, "last_service_done")),
        "campanhas_pendentes": 0,
        "risco_tecnico": "Normal",
        "observacoes": clean_text(row_value(row, col, "observations")),
        "ativo": 1,
        "estado_frota": estado_frota,
        "data_venda": sales_date or invoice_date,
        "valor_venda": valor_venda,
        "valor_venda_sem_iva": clean_float(row_value(row, col, "sales_value")),
        "iva_venda": clean_float(row_value(row, col, "sales_tax")),
        "valor_iva_venda": clean_float(row_value(row, col, "sales_tax_value")),
        "comprador": clean_text(row_value(row, col, "Client")) or clean_text(row_value(row, col, "client_name")),
        "motivo_venda": None,
        "data_baixa": None,
        "ativa_operacional": ativa_operacional,
        "estado_operacional": estado_operacional,
        "current_status_rentway": current_status,
        "status_rentway": status,
        "fleet": clean_text(row_value(row, col, "fleet")),
        "rental_station": clean_text(row_value(row, col, "rental_station")),
        "grupo": clean_text(row_value(row, col, "groupid")),
        "categoria": clean_text(row_value(row, col, "category")),
        "fornecedor": clean_text(row_value(row, col, "supplier_name")),
        "valor_compra": clean_float(row_value(row, col, "value")),
        "iva_compra": clean_float(row_value(row, col, "tax")),
        "valor_iva_compra": clean_float(row_value(row, col, "tax_value")),
        "valor_compra_com_iva": clean_float(row_value(row, col, "value_with_tax")),
        "valor_venda_com_iva": valor_venda,
        "preco_venda_rentway": clean_float(row_value(row, col, "sales_price")),
        "preco_venda_calculado": None,
        "preco_venda_retalho_rentway": clean_float(row_value(row, col, "retail_sales_price")),
        "preco_venda_retalho_calculado": None,
        "documento_rentway": clean_text(row_value(row, col, "document_nr")),
        "documento_venda": clean_text(row_value(row, col, "invoice_nr")),
        "data_fatura_venda": invoice_date,
        "data_retorno": excel_date_to_iso(row_value(row, col, "return_date")),
        "bloqueio_rentway": clean_text(row_value(row, col, "block")),
        "data_bloqueio": excel_date_to_iso(row_value(row, col, "date_block")),
        "cliente_rentway": clean_text(row_value(row, col, "client_name")),
        "programa_venda": clean_text(row_value(row, col, "sales_program")),
        "data_inspecao": excel_date_to_iso(row_value(row, col, "inspection_date")),
        "ald_nr_pagamentos": clean_int(row_value(row, col, "ald_nr_payments")),
        "ald_primeiro_pagamento": excel_date_to_iso(row_value(row, col, "ald_dt_first_payment")),
        "ald_valor_residual": clean_float(row_value(row, col, "ald_residual_value")),
        "depreciacao_codigo_fiscal_rentway": clean_text(row_value(row, col, "depreciation_fiscal_code")),
        "depreciacao_codigo_fiscal_calculado": None,
        "depreciacao_ultima_calculo_rentway": excel_date_to_iso(row_value(row, col, "depreciation_last_calculation")),
        "depreciacao_ultima_calculo_calculado": None,
        "codigo_conta_rentway": clean_text(row_value(row, col, "account_code")),
        "codigo_conta_calculado": None,
        "data_ultimo_movimento": excel_date_to_iso(row_value(row, col, "date_last_movment")),
        "data_bloqueio_peritagem": excel_date_to_iso(row_value(row, col, "date_block_expertie")),
        "fornecedor_financeiro": clean_text(row_value(row, col, "finantial_supplier")),
        "contrato_financeiro": clean_text(row_value(row, col, "finantial_contract")),
        "contrato_financeiro_fim": excel_date_to_iso(row_value(row, col, "finantial_contract_end_date")),
        "garantia_fim": excel_date_to_iso(row_value(row, col, "warrantyenddate")),
        "furto_data_declaracao": excel_date_to_iso(row_value(row, col, "stolen_declaration_date")),
        "furto_numero_declaracao": clean_text(row_value(row, col, "stolen_declaration_number")),
        "aviso_viatura": clean_text(row_value(row, col, "vehicle_warning")),
        "documentnr_rentway": clean_text(row_value(row, col, "DocumentNr")),
        "cliente_atual_rentway": clean_text(row_value(row, col, "Client")),
        "condutor_atual_rentway": clean_text(row_value(row, col, "Driver")),
    }


def init_db():
    with db() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT NOT NULL,
            ultimo_login TEXT
        )
        """)
        ensure_columns(c, "users", {
            "departamento": "TEXT",
            "department_id": "INTEGER",
            "updated_at": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT UNIQUE NOT NULL,
            descricao TEXT,
            ativo INTEGER DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            acao TEXT NOT NULL,
            entidade TEXT,
            entidade_id TEXT,
            detalhe TEXT,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS task_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            user_id INTEGER,
            campo_alterado TEXT NOT NULL,
            valor_antigo TEXT,
            valor_novo TEXT,
            changed_at TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS task_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            document_id INTEGER,
            file_name TEXT NOT NULL,
            file_type TEXT,
            file_size INTEGER,
            sharepoint_url TEXT,
            uploaded_by INTEGER,
            uploaded_at TEXT NOT NULL,
            category TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS documentos_operacionais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            area TEXT NOT NULL,
            entidade TEXT NOT NULL,
            entidade_id TEXT NOT NULL,
            categoria TEXT,
            file_name TEXT NOT NULL,
            original_name TEXT NOT NULL,
            file_type TEXT,
            file_size INTEGER,
            storage_provider TEXT DEFAULT 'local',
            storage_path TEXT,
            sharepoint_url TEXT,
            uploaded_by TEXT,
            uploaded_at TEXT NOT NULL,
            archived INTEGER DEFAULT 0
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_system TEXT NOT NULL,
            entity_type TEXT,
            file_name TEXT,
            original_name TEXT,
            imported_by INTEGER,
            imported_at TEXT NOT NULL,
            total_rows INTEGER DEFAULT 0,
            created_rows INTEGER DEFAULT 0,
            updated_rows INTEGER DEFAULT 0,
            skipped_rows INTEGER DEFAULT 0,
            error_rows INTEGER DEFAULT 0,
            status TEXT DEFAULT 'Concluído',
            detalhe TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS import_errors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_batch_id INTEGER,
            row_number INTEGER,
            entity_type TEXT,
            error_message TEXT NOT NULL,
            raw_data TEXT,
            created_at TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS conhecimento_artigos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            slug TEXT UNIQUE NOT NULL,
            area TEXT NOT NULL,
            tema TEXT,
            resumo TEXT,
            conteudo TEXT NOT NULL,
            estado TEXT DEFAULT 'Publicado',
            visivel_contexto INTEGER DEFAULT 1,
            criado_por INTEGER,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT
        )
        """)

        sync_departments(c)

        for artigo in CONHECIMENTO_SEEDS:
            c.execute("""
                INSERT OR IGNORE INTO conhecimento_artigos
                (titulo, slug, area, tema, resumo, conteudo, estado, visivel_contexto, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, 'Publicado', 1, ?)
            """, (
                artigo["titulo"],
                artigo["slug"],
                artigo["area"],
                artigo["tema"],
                artigo["resumo"],
                artigo["conteudo"],
                now(),
            ))

        user_count = c.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
        if user_count == 0:
            c.execute("""
                INSERT INTO users (nome, email, password_hash, role, ativo, criado_em)
                VALUES (?, ?, ?, 'Admin', 1, ?)
            """, (
                "Administrador",
                DEFAULT_ADMIN_EMAIL,
                generate_password_hash(DEFAULT_ADMIN_PASSWORD),
                now(),
            ))

        c.execute("""
        CREATE TABLE IF NOT EXISTS centro_tarefas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria TEXT NOT NULL,
            assunto TEXT NOT NULL,
            descricao TEXT NOT NULL,
            departamento TEXT NOT NULL,
            prioridade TEXT NOT NULL DEFAULT 'Média',
            estado TEXT NOT NULL DEFAULT 'Novo',
            matricula TEXT,
            estacao TEXT,
            cliente TEXT,
            responsavel TEXT,
            data_limite TEXT,
            criado_por_id INTEGER,
            criado_por_nome TEXT,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT,
            fechado_em TEXT
        )
        """)
        ensure_columns(c, "centro_tarefas", {
            "source_system": "TEXT DEFAULT 'CarFast'",
            "external_reference": "TEXT",
            "archived": "INTEGER DEFAULT 0",
            "deleted": "INTEGER DEFAULT 0",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS centro_tarefas_comentarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tarefa_id INTEGER NOT NULL,
            user_id INTEGER,
            user_nome TEXT,
            comentario TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS centro_tarefas_anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tarefa_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            carregado_por TEXT,
            carregado_em TEXT NOT NULL
        )
        """)
        ensure_columns(c, "centro_tarefas_anexos", {
            "document_id": "INTEGER",
            "file_type": "TEXT",
            "file_size": "INTEGER",
            "sharepoint_url": "TEXT",
            "category": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS viaturas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT UNIQUE NOT NULL,
            vin TEXT,
            marca TEXT,
            modelo TEXT,
            versao TEXT,
            motorizacao TEXT,
            combustivel TEXT,
            caixa TEXT,
            ano INTEGER,
            data_compra TEXT,
            km_atual INTEGER,
            proxima_revisao_km INTEGER,
            proxima_revisao_data TEXT,
            ultima_revisao TEXT,
            campanhas_pendentes INTEGER DEFAULT 0,
            risco_tecnico TEXT DEFAULT 'Normal',
            observacoes TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT NOT NULL
        )
        """)

        ensure_columns(c, "viaturas", LIFECYCLE_COLUMNS)
        ensure_columns(c, "viaturas", {
            "rentway_id": "TEXT",
            "external_reference": "TEXT",
            "source_system": "TEXT DEFAULT 'Rentway'",
            "last_imported_at": "TEXT",
            "import_batch_id": "INTEGER",
            "data_hash": "TEXT",
            "sync_status": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS artigos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nome TEXT NOT NULL,
            categoria TEXT,
            unidade TEXT,
            localizacao TEXT,
            fornecedor TEXT,
            preco_unitario REAL DEFAULT 0,
            stock_minimo INTEGER DEFAULT 0,
            stock_atual INTEGER DEFAULT 0,
            criado_em TEXT NOT NULL
        )
        """)
        ensure_columns(c, "artigos", {
            "codigo": "TEXT",
            "nome": "TEXT",
            "categoria": "TEXT",
            "unidade": "TEXT",
            "localizacao": "TEXT",
            "fornecedor": "TEXT",
            "preco_unitario": "REAL DEFAULT 0",
            "stock_minimo": "INTEGER DEFAULT 0",
            "stock_atual": "INTEGER DEFAULT 0",
            "criado_em": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS movimentos_stock (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            artigo_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            motivo TEXT,
            matricula TEXT,
            responsavel TEXT,
            observacoes TEXT,
            criado_em TEXT NOT NULL,
            FOREIGN KEY(artigo_id) REFERENCES artigos(id)
        )
        """)
        ensure_columns(c, "movimentos_stock", {
            "artigo_id": "INTEGER",
            "tipo": "TEXT",
            "quantidade": "INTEGER DEFAULT 0",
            "motivo": "TEXT",
            "matricula": "TEXT",
            "responsavel": "TEXT",
            "observacoes": "TEXT",
            "criado_em": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS importacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            filename TEXT,
            original_name TEXT,
            sheet_name TEXT,
            total_linhas INTEGER DEFAULT 0,
            criadas INTEGER DEFAULT 0,
            atualizadas INTEGER DEFAULT 0,
            vendidas INTEGER DEFAULT 0,
            ativas INTEGER DEFAULT 0,
            ignoradas INTEGER DEFAULT 0,
            responsavel TEXT,
            observacoes TEXT,
            colunas_json TEXT,
            criado_em TEXT NOT NULL
        )
        """)

        ensure_columns(c, "importacoes", {
            "sheet_name": "TEXT",
            "total_linhas": "INTEGER DEFAULT 0",
            "colunas_json": "TEXT",
            "source_system": "TEXT DEFAULT 'Rentway'",
            "import_batch_id": "INTEGER",
            "status": "TEXT DEFAULT 'Concluído'",
            "error_rows": "INTEGER DEFAULT 0",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS importacoes_linhas_raw (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            matricula TEXT,
            raw_json TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS rentway_impros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rentway_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            matricula TEXT,
            numero_impro TEXT,
            marca TEXT,
            modelo TEXT,
            grupo TEXT,
            condutor TEXT,
            data_abertura TEXT,
            data_fecho TEXT,
            data_in TEXT,
            data_out TEXT,
            estado TEXT,
            estacao TEXT,
            estacao_out TEXT,
            km INTEGER,
            driven_kms INTEGER,
            oficina TEXT,
            motivo TEXT,
            descricao TEXT,
            custo REAL,
            dias_imobilizado INTEGER,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        ensure_columns(c, "rentway_impros", IMPRO_COLUMNS)

        c.execute("""
        CREATE TABLE IF NOT EXISTS rentway_folhas_obra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rentway_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            folha_obra_origem TEXT,
            data_documento TEXT,
            matricula TEXT,
            modelo TEXT,
            fornecedor TEXT,
            numero_fatura TEXT,
            estacao TEXT,
            kms REAL,
            tipo_servico TEXT,
            natureza TEXT,
            quantidade REAL,
            valor_com_iva REAL,
            descricao TEXT,
            numero_documento TEXT,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS sinistros_allianz (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            allianz_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            adesao TEXT,
            matricula TEXT,
            referencia_sinistro TEXT,
            data_sinistro TEXT,
            data_encerramento TEXT,
            causa TEXT,
            valor_rc REAL,
            valor_dp REAL,
            valor_ids_credor REAL,
            valor_vidros REAL,
            custos_gestao REAL,
            custo_total REAL,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS rentway_accident_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ar_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            accident_report_id TEXT,
            document_type TEXT,
            document_no TEXT,
            state TEXT,
            request_date TEXT,
            accident_date TEXT,
            friendly_declaration TEXT,
            matricula TEXT,
            unit_no TEXT,
            insurance_policy TEXT,
            rental_station_out TEXT,
            created_by_rental_station TEXT,
            status TEXT,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS faturas_fornecedores_linhas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            linha_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            fornecedor TEXT,
            documento TEXT,
            data_doc TEXT,
            matricula TEXT,
            kms REAL,
            or_reparacao TEXT,
            chassis TEXT,
            descricao TEXT,
            referencia TEXT,
            pv_unit REAL,
            desc_pct REAL,
            pliq_unit REAL,
            quantidade REAL,
            total_liq REAL,
            ct TEXT,
            fonte_pdf TEXT,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS contratos_financeiro (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contrato_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            contrato_nr TEXT,
            data_out TEXT,
            data_in TEXT,
            data_fim TEXT,
            dias REAL,
            valor_dia REAL,
            valor_fatura REAL,
            valor_fatura_caixa REAL,
            valor_nota_debito REAL,
            valor_nota_credito REAL,
            valor_total REAL,
            cliente_id TEXT,
            tipo_cliente TEXT,
            origem TEXT,
            grupo_reservado TEXT,
            grupo_entregue TEXT,
            matricula_final TEXT,
            vendedor TEXT,
            nome_vendedor TEXT,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS contratos_viaturas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            movimento_key TEXT UNIQUE NOT NULL,
            importacao_id INTEGER NOT NULL,
            row_number INTEGER NOT NULL,
            viatura_id INTEGER,
            contrato_nr TEXT,
            cliente_nome TEXT,
            data_out TEXT,
            data_in TEXT,
            matricula TEXT,
            estado_importacao TEXT,
            alertas TEXT,
            raw_json TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS processos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_impro TEXT UNIQUE NOT NULL,
            viatura_id INTEGER NOT NULL,
            matricula TEXT NOT NULL,
            km_abertura INTEGER,
            responsavel TEXT,
            descricao_inicial TEXT,
            estado TEXT NOT NULL DEFAULT 'Aberto',
            data_abertura TEXT NOT NULL,
            data_fecho TEXT,
            observacoes TEXT
        )
        """)
        ensure_columns(c, "processos", {
            "origem": "TEXT DEFAULT 'CarFast'",
            "arquivo": "INTEGER DEFAULT 0",
            "visivel_gestao": "INTEGER DEFAULT 1",
        })
        c.execute("""
            UPDATE processos
            SET origem = COALESCE(origem, 'Importação Diagnósticos PDF'),
                arquivo = 1,
                visivel_gestao = 0
            WHERE numero_impro LIKE 'PDF-DIAG-%'
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS rececoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER UNIQUE NOT NULL,
            servicebox_verificado INTEGER DEFAULT 0,
            campanhas_verificadas INTEGER DEFAULT 0,
            campanhas_pendentes INTEGER DEFAULT 0,
            plano_manutencao_verificado TEXT DEFAULT 'Não',
            historico_oem_verificado INTEGER DEFAULT 0,
            observacoes TEXT,
            atualizado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS diagnosticos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            viatura_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            data_diagnostico TEXT NOT NULL,
            km_diagnostico INTEGER,
            n_manutencoes_ecu TEXT,
            intervencoes_reais TEXT,
            telecarregamento TEXT,
            ultima_manutencao_registada TEXT,
            manutencao_nao_registada TEXT,
            km_manutencao_anterior INTEGER,
            km_ate_manutencao INTEGER,
            diluicao_oleo REAL,
            carbono_oleo REAL,
            problema_identificado TEXT,
            causa_provavel TEXT,
            intervencao_recomendada TEXT,
            oficina_recomendada TEXT,
            pode_circular TEXT,
            prioridade TEXT,
            inconsistencias TEXT,
            observacoes TEXT,
            criado_em TEXT NOT NULL
        )
        """)
        ensure_columns(c, "diagnosticos", {
            "import_key": "TEXT",
            "origem_importacao": "TEXT",
            "origem_linha": "INTEGER",
            "estado_importacao": "TEXT",
            "alertas": "TEXT",
            "fonte_pdf": "TEXT",
            "fonte_path": "TEXT",
            "vin": "TEXT",
            "numero_relatorio": "TEXT",
            "texto_extraido": "TEXT",
        })
        c.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_diagnosticos_import_key
            ON diagnosticos(import_key)
            WHERE import_key IS NOT NULL
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS folhas_obra (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER UNIQUE NOT NULL,
            numero_fo TEXT,
            estado TEXT NOT NULL DEFAULT 'Aberta',
            responsavel TEXT,
            observacoes TEXT,
            criada_em TEXT NOT NULL,
            fechada_em TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS tarefas_fo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fo_id INTEGER NOT NULL,
            descricao TEXT NOT NULL,
            tipo_servico TEXT,
            responsavel TEXT,
            autorizada INTEGER DEFAULT 1,
            estado TEXT DEFAULT 'Pendente',
            observacoes TEXT,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS validacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER UNIQUE NOT NULL,
            fo_fechada INTEGER DEFAULT 0,
            diagnostico_pos_anexado INTEGER DEFAULT 0,
            ficha_viatura_atualizada INTEGER DEFAULT 0,
            parametros_atualizados INTEGER DEFAULT 0,
            documentacao_completa INTEGER DEFAULT 0,
            observacoes TEXT,
            validado_por TEXT,
            atualizado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            area TEXT NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS processos_manutencao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER,
            numero_impro TEXT,
            viatura_id INTEGER,
            matricula TEXT,
            km_abertura INTEGER,
            responsavel TEXT,
            descricao_inicial TEXT,
            estado_processo TEXT DEFAULT 'Aberto',
            data_abertura TEXT,
            data_fecho TEXT,
            protocolo_regra_id INTEGER,
            criado_em TEXT NOT NULL
        )
        """)
        ensure_columns(c, "processos_manutencao", {
            "processo_id": "INTEGER",
            "numero_impro": "TEXT",
            "viatura_id": "INTEGER",
            "matricula": "TEXT",
            "km_abertura": "INTEGER",
            "responsavel": "TEXT",
            "descricao_inicial": "TEXT",
            "estado_processo": "TEXT DEFAULT 'Aberto'",
            "data_abertura": "TEXT",
            "data_fecho": "TEXT",
            "protocolo_regra_id": "INTEGER",
            "criado_em": "TEXT",
        })

        c.execute("""
        CREATE TABLE IF NOT EXISTS protocolo_fases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_fase TEXT UNIQUE NOT NULL,
            nome_fase TEXT NOT NULL,
            ordem INTEGER DEFAULT 0,
            obrigatoria INTEGER DEFAULT 1,
            bloqueia_avanco INTEGER DEFAULT 0,
            ativo INTEGER DEFAULT 1
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS protocolo_campos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_campo TEXT UNIQUE NOT NULL,
            nome_campo TEXT NOT NULL,
            fase_id INTEGER,
            categoria TEXT,
            tipo_campo TEXT,
            obrigatorio_default INTEGER DEFAULT 0,
            aceita_anexo INTEGER DEFAULT 0,
            descricao TEXT,
            ativo INTEGER DEFAULT 1,
            ordem INTEGER DEFAULT 0
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS protocolo_regras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_regra TEXT NOT NULL,
            marca TEXT,
            modelo TEXT,
            versao TEXT,
            motorizacao TEXT,
            combustivel TEXT,
            categoria_viatura TEXT,
            grupo TEXT,
            ano_min INTEGER,
            ano_max INTEGER,
            km_min INTEGER,
            km_max INTEGER,
            tipo_intervencao TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS protocolo_regra_campos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            regra_id INTEGER NOT NULL,
            campo_id INTEGER NOT NULL,
            obrigatorio INTEGER DEFAULT 0,
            exige_anexo INTEGER DEFAULT 0,
            ordem INTEGER DEFAULT 0,
            instrucoes TEXT,
            UNIQUE(regra_id, campo_id)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS processo_campos_valores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            campo_id INTEGER NOT NULL,
            valor_texto TEXT,
            valor_numero REAL,
            valor_data TEXT,
            valor_booleano INTEGER,
            observacoes TEXT,
            preenchido_por TEXT,
            preenchido_em TEXT,
            UNIQUE(processo_id, campo_id)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS processo_campos_anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            campo_id INTEGER,
            fase_id INTEGER,
            tipo_anexo TEXT,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            carregado_por TEXT,
            carregado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS validacoes_processo (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            fase_id INTEGER,
            validado INTEGER DEFAULT 0,
            validado_por TEXT,
            validado_em TEXT,
            observacoes TEXT,
            UNIQUE(processo_id, fase_id)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS comparacoes_diagnostico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            processo_id INTEGER NOT NULL,
            viatura_id INTEGER,
            campo_id INTEGER,
            valor_anterior TEXT,
            valor_atual TEXT,
            diferenca TEXT,
            resultado TEXT,
            alerta INTEGER DEFAULT 0
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS incidentes_categorias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            ativo INTEGER DEFAULT 1
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS incidentes_classificacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            categoria_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            descricao TEXT,
            ativo INTEGER DEFAULT 1,
            prioridade_default TEXT DEFAULT 'Média',
            gera_tarefa INTEGER DEFAULT 1,
            faturavel_default TEXT DEFAULT 'A avaliar',
            sla_horas_default INTEGER DEFAULT 24,
            equipa_responsavel_default TEXT,
            exige_anexo INTEGER DEFAULT 0,
            instrucoes TEXT,
            criado_em TEXT NOT NULL,
            UNIQUE(categoria_id, nome)
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS incidentes_frota (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            viatura_id INTEGER NOT NULL,
            matricula TEXT NOT NULL,
            categoria_id INTEGER NOT NULL,
            classificacao_id INTEGER NOT NULL,
            processo_id INTEGER,
            fo_id INTEGER,
            contrato_id TEXT,
            cliente TEXT,
            oficina_marca TEXT,
            descricao TEXT NOT NULL,
            origem TEXT,
            gravidade TEXT DEFAULT 'Média',
            estado TEXT DEFAULT 'Aberto',
            faturavel TEXT DEFAULT 'A avaliar',
            valor_estimado REAL,
            responsavel_followup TEXT,
            acao_necessaria TEXT,
            data_limite TEXT,
            criado_por TEXT,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT,
            fechado_em TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS incidentes_anexos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            incidente_id INTEGER NOT NULL,
            tipo_anexo TEXT,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            carregado_por TEXT,
            carregado_em TEXT NOT NULL
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS tarefas_frota (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            origem_tipo TEXT,
            origem_id TEXT,
            viatura_id INTEGER,
            matricula TEXT,
            titulo TEXT NOT NULL,
            descricao TEXT,
            responsavel TEXT,
            prioridade TEXT DEFAULT 'Média',
            estado TEXT DEFAULT 'Pendente',
            data_limite TEXT,
            criada_em TEXT NOT NULL,
            concluida_em TEXT,
            validada_por TEXT,
            validada_em TEXT
        )
        """)

        c.execute("""
        CREATE TABLE IF NOT EXISTS viaturas_venda (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            viatura_id INTEGER NOT NULL,
            matricula TEXT NOT NULL,
            selecionada_venda INTEGER DEFAULT 1,
            estado_venda TEXT DEFAULT 'Selecionada',
            preco_sugerido REAL,
            preco_minimo REAL,
            preco_comercio REAL,
            observacoes TEXT,
            fornecedor_financeiro TEXT,
            proxima_revisao TEXT,
            valor_aquisicao REAL,
            data_aquisicao TEXT,
            meses_depreciacao INTEGER,
            depreciacao_mensal REAL,
            depreciacao_acumulada REAL,
            valor_atual_teorico REAL,
            exportada INTEGER DEFAULT 0,
            exportada_em TEXT,
            criado_em TEXT NOT NULL,
            atualizado_em TEXT
        )
        """)
        c.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_viaturas_venda_ativa
            ON viaturas_venda(viatura_id)
            WHERE selecionada_venda = 1 AND estado_venda NOT IN ('Vendida', 'Cancelada')
        """)

        ensure_columns(c, "folhas_obra", {
            "estado_fo": "TEXT",
        })
        ensure_columns(c, "tarefas_fo", {
            "concluida_em": "TEXT",
        })

        for nome in INCIDENTES_CATEGORIAS_INICIAIS:
            c.execute("""
                INSERT INTO incidentes_categorias (nome, ativo)
                VALUES (?, 1)
                ON CONFLICT(nome) DO UPDATE SET ativo = 1
            """, (nome,))

        categorias_incidentes = {
            row["nome"]: row["id"]
            for row in c.execute("SELECT id, nome FROM incidentes_categorias").fetchall()
        }
        for categoria_nome, classificacoes in INCIDENTES_CLASSIFICACOES_INICIAIS.items():
            categoria_id = categorias_incidentes.get(categoria_nome)
            if not categoria_id:
                continue
            for nome in classificacoes:
                c.execute("""
                    INSERT INTO incidentes_classificacoes
                    (categoria_id, nome, descricao, ativo, prioridade_default,
                     gera_tarefa, faturavel_default, sla_horas_default,
                     equipa_responsavel_default, exige_anexo, instrucoes, criado_em)
                    VALUES (?, ?, NULL, 1, 'Média', 1, 'A avaliar', 24, NULL, 0, NULL, ?)
                    ON CONFLICT(categoria_id, nome) DO UPDATE SET ativo = 1
                """, (categoria_id, nome, now()))

        for codigo, nome, ordem, obrigatoria, bloqueia in PROTOCOLO_FASES_INICIAIS:
            c.execute("""
                INSERT INTO protocolo_fases
                (codigo_fase, nome_fase, ordem, obrigatoria, bloqueia_avanco, ativo)
                VALUES (?, ?, ?, ?, ?, 1)
                ON CONFLICT(codigo_fase) DO UPDATE SET
                    nome_fase = excluded.nome_fase,
                    ordem = excluded.ordem,
                    obrigatoria = excluded.obrigatoria,
                    bloqueia_avanco = excluded.bloqueia_avanco,
                    ativo = 1
            """, (codigo, nome, ordem, obrigatoria, bloqueia))

        fases = {
            row["codigo_fase"]: row["id"]
            for row in c.execute("SELECT id, codigo_fase FROM protocolo_fases").fetchall()
        }

        for ordem, item in enumerate(PROTOCOLO_CAMPOS_INICIAIS, start=1):
            codigo, nome, codigo_fase, categoria, tipo, obrigatorio, aceita_anexo, descricao = item
            c.execute("""
                INSERT INTO protocolo_campos
                (codigo_campo, nome_campo, fase_id, categoria, tipo_campo,
                 obrigatorio_default, aceita_anexo, descricao, ativo, ordem)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                ON CONFLICT(codigo_campo) DO UPDATE SET
                    nome_campo = excluded.nome_campo,
                    fase_id = excluded.fase_id,
                    categoria = excluded.categoria,
                    tipo_campo = excluded.tipo_campo,
                    obrigatorio_default = excluded.obrigatorio_default,
                    aceita_anexo = excluded.aceita_anexo,
                    descricao = excluded.descricao,
                    ativo = 1,
                    ordem = excluded.ordem
            """, (
                codigo,
                nome,
                fases.get(codigo_fase),
                categoria,
                tipo,
                obrigatorio,
                aceita_anexo,
                descricao,
                ordem,
            ))

        c.execute("""
            UPDATE protocolo_regras
            SET nome_regra = 'Protocolo Base de Manutenção',
                marca = NULL,
                modelo = NULL,
                versao = NULL,
                motorizacao = NULL,
                combustivel = NULL,
                categoria_viatura = NULL,
                grupo = NULL,
                ano_min = NULL,
                ano_max = NULL,
                km_min = NULL,
                km_max = NULL,
                tipo_intervencao = NULL,
                ativo = 1
            WHERE nome_regra = 'Protocolo Stellantis 1.5 BlueHDi'
        """)
        regra = c.execute(
            "SELECT id FROM protocolo_regras WHERE nome_regra = ?",
            ("Protocolo Base de Manutenção",)
        ).fetchone()
        if regra:
            regra_id = regra["id"]
        else:
            cur = c.execute("""
                INSERT INTO protocolo_regras
                (nome_regra, marca, motorizacao, combustivel, tipo_intervencao, ativo, criado_em)
                VALUES (?, ?, ?, ?, ?, 1, ?)
            """, (
                "Protocolo Base de Manutenção",
                None,
                None,
                None,
                None,
                now(),
            ))
            regra_id = cur.lastrowid

        campos_obrigatorios_base = [
            "servicebox_verificado",
            "campanhas_verificadas",
            "plano_manutencao_verificado",
            "autoscan_pdf",
            "n_manutencoes_ecu",
            "telecarregamento",
            "diluicao_oleo",
            "carbono_oleo",
            "problema_identificado",
            "intervencao_recomendada",
            "diagnostico_pos_realizado",
            "autoscan_pos_pdf",
            "fo_fechada",
            "documentacao_completa",
        ]
        for ordem, codigo_campo in enumerate(campos_obrigatorios_base, start=1):
            campo = c.execute(
                "SELECT id, aceita_anexo FROM protocolo_campos WHERE codigo_campo = ?",
                (codigo_campo,)
            ).fetchone()
            if campo:
                c.execute("""
                    INSERT INTO protocolo_regra_campos
                    (regra_id, campo_id, obrigatorio, exige_anexo, ordem, instrucoes)
                    VALUES (?, ?, 1, ?, ?, ?)
                    ON CONFLICT(regra_id, campo_id) DO UPDATE SET
                        obrigatorio = 1,
                        exige_anexo = excluded.exige_anexo,
                        ordem = excluded.ordem
                """, (
                    regra_id,
                    campo["id"],
                    1 if codigo_campo in {"autoscan_pdf", "autoscan_pos_pdf"} else 0,
                    ordem,
                    "Campo obrigatório no protocolo base de manutenção.",
                ))


def ensure_viatura(matricula, km=None):
    matricula = matricula.strip().upper()

    with db() as c:
        v = c.execute(
            "SELECT * FROM viaturas WHERE matricula = ?",
            (matricula,)
        ).fetchone()

        if v:
            if km:
                c.execute(
                    "UPDATE viaturas SET km_atual = ? WHERE id = ?",
                    (km, v["id"])
                )
            return v["id"]

        cur = c.execute(
            "INSERT INTO viaturas (matricula, km_atual, criado_em) VALUES (?, ?, ?)",
            (matricula, km, now())
        )
        return cur.lastrowid


def processo(processo_id):
    with db() as c:
        return c.execute("""
            SELECT p.*, v.vin, v.marca, v.modelo, v.versao, v.motorizacao,
                   v.combustivel, v.ano, v.grupo, v.categoria, v.km_atual
            FROM processos p
            JOIN viaturas v ON v.id = p.viatura_id
            WHERE p.id = ?
        """, (processo_id,)).fetchone()


def protocol_text(value):
    value = "" if value is None else str(value)
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return value.strip().lower()


def protocol_matches(rule_value, vehicle_value):
    rule_value = protocol_text(rule_value)
    vehicle_value = protocol_text(vehicle_value)
    if not rule_value:
        return True
    if not vehicle_value:
        return False
    parts = [p.strip() for p in re.split(r"[/,;|]", rule_value) if p.strip()]
    return any(part in vehicle_value or vehicle_value in part for part in parts)


def protocolo_regra_aplicavel(conn, viatura, tipo_intervencao=None):
    rules = conn.execute("""
        SELECT *
        FROM protocolo_regras
        WHERE ativo = 1
        ORDER BY
            CASE
              WHEN marca IS NULL AND modelo IS NULL AND versao IS NULL
               AND motorizacao IS NULL AND combustivel IS NULL
               AND categoria_viatura IS NULL AND grupo IS NULL
               AND ano_min IS NULL AND ano_max IS NULL
               AND km_min IS NULL AND km_max IS NULL
               AND tipo_intervencao IS NULL
              THEN 1 ELSE 0
            END,
            id
    """).fetchall()
    best = None
    best_score = -1
    km = viatura["km_abertura"] or viatura["km_atual"] or 0
    ano = viatura["ano"] or 0

    for rule in rules:
        checks = [
            ("marca", "marca", 3),
            ("modelo", "modelo", 3),
            ("versao", "versao", 2),
            ("motorizacao", "motorizacao", 4),
            ("combustivel", "combustivel", 2),
            ("categoria_viatura", "categoria", 1),
            ("grupo", "grupo", 1),
        ]
        score = 0
        failed = False
        for rule_col, vehicle_col, weight in checks:
            if rule[rule_col]:
                if protocol_matches(rule[rule_col], viatura[vehicle_col]):
                    score += weight
                else:
                    failed = True
                    break
        if failed:
            continue
        if rule["ano_min"] and (not ano or ano < rule["ano_min"]):
            continue
        if rule["ano_max"] and (not ano or ano > rule["ano_max"]):
            continue
        if rule["km_min"] and km < rule["km_min"]:
            continue
        if rule["km_max"] and km > rule["km_max"]:
            continue
        if tipo_intervencao and rule["tipo_intervencao"]:
            if protocol_matches(rule["tipo_intervencao"], tipo_intervencao):
                score += 1
            else:
                continue
        if score > best_score:
            best = rule
            best_score = score

    return best


def protocolo_campos_da_regra(conn, regra_id=None):
    if regra_id:
        return conn.execute("""
            SELECT pc.*, pf.codigo_fase, pf.nome_fase, pf.ordem AS fase_ordem,
                   prc.obrigatorio AS obrigatorio_regra,
                   prc.exige_anexo AS exige_anexo,
                   prc.instrucoes AS instrucoes_regra,
                   prc.ordem AS regra_ordem
            FROM protocolo_regra_campos prc
            JOIN protocolo_campos pc ON pc.id = prc.campo_id
            LEFT JOIN protocolo_fases pf ON pf.id = pc.fase_id
            WHERE prc.regra_id = ? AND pc.ativo = 1
            ORDER BY pf.ordem, prc.ordem, pc.ordem
        """, (regra_id,)).fetchall()

    return conn.execute("""
        SELECT pc.*, pf.codigo_fase, pf.nome_fase, pf.ordem AS fase_ordem,
               pc.obrigatorio_default AS obrigatorio_regra,
               pc.aceita_anexo AS exige_anexo,
               NULL AS instrucoes_regra,
               pc.ordem AS regra_ordem
        FROM protocolo_campos pc
        LEFT JOIN protocolo_fases pf ON pf.id = pc.fase_id
        WHERE pc.ativo = 1
        ORDER BY pf.ordem, pc.ordem
    """).fetchall()


def agrupar_campos_por_fase(rows):
    fases = []
    idx = {}
    for row in rows:
        key = row["codigo_fase"] or "sem_fase"
        if key not in idx:
            idx[key] = {
                "codigo": key,
                "nome": row["nome_fase"] or "Sem fase",
                "campos": [],
            }
            fases.append(idx[key])
        idx[key]["campos"].append(row)
    return fases


def ensure_motor_processo(conn, p, regra_id):
    row = conn.execute(
        "SELECT id FROM processos_manutencao WHERE processo_id = ?",
        (p["id"],)
    ).fetchone()
    if row:
        conn.execute(
            "UPDATE processos_manutencao SET protocolo_regra_id = ? WHERE id = ?",
            (regra_id, row["id"])
        )
        return row["id"]
    cur = conn.execute("""
        INSERT INTO processos_manutencao
        (processo_id, numero_impro, viatura_id, matricula, km_abertura,
         responsavel, descricao_inicial, estado_processo, data_abertura,
         protocolo_regra_id, criado_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        p["id"],
        p["numero_impro"],
        p["viatura_id"],
        p["matricula"],
        p["km_abertura"],
        p["responsavel"],
        p["descricao_inicial"],
        p["estado"],
        p["data_abertura"],
        regra_id,
        now(),
    ))
    return cur.lastrowid


def valor_para_campo(campo, raw_value, presente=False):
    tipo = campo["tipo_campo"]
    if tipo == "Sim/Não":
        return None, None, None, 1 if presente else 0
    if tipo in {"Número", "Percentagem"}:
        return None, clean_float(raw_value), None, None
    if tipo == "Data":
        return None, None, raw_value or None, None
    return raw_value or None, None, None, None


def valor_visivel(campo, valor):
    if not valor:
        return ""
    tipo = campo["tipo_campo"]
    if tipo == "Sim/Não":
        return "Sim" if valor["valor_booleano"] else "Não"
    if tipo in {"Número", "Percentagem"}:
        return valor["valor_numero"] if valor["valor_numero"] is not None else ""
    if tipo == "Data":
        return valor["valor_data"] or ""
    return valor["valor_texto"] or ""


def save_incidente_anexos(conn, incidente_id, files, carregado_por=None):
    for f in files:
        if not f or not f.filename:
            continue
        safe = secure_filename(f.filename)
        stored_name = f"incidente_{incidente_id}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe}"
        path = UPLOAD_DIR / stored_name
        f.save(path)
        conn.execute("""
            INSERT INTO incidentes_anexos
            (incidente_id, tipo_anexo, filename, original_name, carregado_por, carregado_em)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            incidente_id,
            Path(f.filename).suffix.lower().lstrip(".") or "documento",
            stored_name,
            f.filename,
            carregado_por,
            now(),
        ))


def criar_tarefa_frota_para_incidente(conn, incidente, classificacao):
    if not classificacao or not classificacao["gera_tarefa"]:
        return None

    data_limite = (
        datetime.now() + timedelta(hours=classificacao["sla_horas_default"] or 24)
    ).isoformat(timespec="minutes")
    instrucoes = classificacao["instrucoes"] or ""
    descricao = incidente["descricao"] or ""
    if instrucoes:
        descricao = f"{descricao}\n\nInstruções: {instrucoes}"

    cur = conn.execute("""
        INSERT INTO tarefas_frota
        (origem_tipo, origem_id, viatura_id, matricula, titulo, descricao,
         responsavel, prioridade, estado, data_limite, criada_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Pendente', ?, ?)
    """, (
        "incidente_frota",
        str(incidente["id"]),
        incidente["viatura_id"],
        incidente["matricula"],
        f"Follow-up incidente {incidente['matricula']} — {classificacao['nome']}",
        descricao,
        incidente["responsavel_followup"] or classificacao["equipa_responsavel_default"],
        classificacao["prioridade_default"] or incidente["gravidade"] or "Média",
        incidente["data_limite"] or data_limite,
        now(),
    ))
    return cur.lastrowid


def parse_iso_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    if "T" in text:
        text = text.split("T", 1)[0]
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def meses_entre_datas(data_inicio, data_fim=None):
    if not data_inicio:
        return 0
    data_fim = data_fim or datetime.now().date()
    meses = (data_fim.year - data_inicio.year) * 12 + (data_fim.month - data_inicio.month)
    if data_fim.day < data_inicio.day:
        meses -= 1
    return max(0, meses)


def calcular_depreciacao_venda(valor_aquisicao, data_aquisicao):
    valor = clean_float(valor_aquisicao) or 0
    data = parse_iso_date(data_aquisicao)
    meses = min(meses_entre_datas(data), 96) if data else 0
    depreciacao_mensal = round(valor / 96, 2) if valor else 0
    depreciacao_acumulada = round(depreciacao_mensal * meses, 2)
    valor_atual = round(max(0, valor - depreciacao_acumulada), 2)
    return {
        "valor_aquisicao": round(valor, 2) if valor else None,
        "data_aquisicao": data.isoformat() if data else None,
        "meses_depreciacao": meses,
        "depreciacao_mensal": depreciacao_mensal,
        "depreciacao_acumulada": depreciacao_acumulada,
        "valor_atual_teorico": valor_atual,
    }


def valor_aquisicao_viatura(row):
    return (
        row["valor_compra_com_iva"]
        if "valor_compra_com_iva" in row.keys() and row["valor_compra_com_iva"] is not None
        else row["valor_compra"]
    )


def dados_venda_da_viatura(row):
    dep = calcular_depreciacao_venda(valor_aquisicao_viatura(row), row["data_compra"])
    return {
        **dep,
        "fornecedor_financeiro": row["fornecedor_financeiro"] if "fornecedor_financeiro" in row.keys() else None,
        "proxima_revisao": row["proxima_revisao_data"] if "proxima_revisao_data" in row.keys() else None,
    }


def centro_estados_para_categoria(categoria):
    return CENTRO_TAREFAS_ESTADOS.get(categoria, ["Novo"])


def can_view_centro_tarefa(row):
    role = session.get("role")
    if role in ("Admin", "Gestor"):
        if role == "Admin":
            return True
        user_dep = session.get("departamento")
        return not user_dep or row["departamento"] == user_dep or row["criado_por_id"] == session.get("user_id")
    return row["criado_por_id"] == session.get("user_id")


def save_centro_tarefa_anexos(conn, tarefa_id, files):
    for f in files:
        if not f or not f.filename:
            continue
        safe = secure_filename(f.filename)
        stored_name = f"centro_tarefa_{tarefa_id}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe}"
        target = UPLOAD_DIR / stored_name
        f.save(target)
        file_size = target.stat().st_size if target.exists() else None
        file_type = f.mimetype or None
        document_id, storage_path = register_operational_document(
            conn,
            area="Centro de Tarefas",
            entidade="centro_tarefas",
            entidade_id=tarefa_id,
            categoria="Anexo",
            filename=stored_name,
            original_name=f.filename,
            uploaded_by=session.get("user_name"),
            file_size=file_size,
            file_type=file_type,
        )
        conn.execute("""
            INSERT INTO centro_tarefas_anexos
            (tarefa_id, filename, original_name, carregado_por, carregado_em,
             document_id, file_type, file_size, category)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            tarefa_id,
            stored_name,
            f.filename,
            session.get("user_name"),
            now(),
            document_id,
            file_type,
            file_size,
            "Anexo",
        ))
        conn.execute("""
            INSERT INTO task_attachments
            (task_id, document_id, file_name, file_type, file_size, sharepoint_url,
             uploaded_by, uploaded_at, category)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            tarefa_id,
            document_id,
            stored_name,
            file_type,
            file_size,
            None,
            session.get("user_id"),
            now(),
            "Anexo",
        ))
        record_task_history(conn, tarefa_id, "anexo", None, storage_path)


def uploads(processo_id, area, files):
    with db() as c:
        for f in files:
            if not f or not f.filename:
                continue

            original = f.filename
            safe = secure_filename(original)
            name = f"{processo_id}_{area}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe}"

            f.save(UPLOAD_DIR / name)

            c.execute("""
                INSERT INTO anexos
                (processo_id, area, filename, original_name, criado_em)
                VALUES (?, ?, ?, ?, ?)
            """, (processo_id, area, name, original, now()))


def upsert_vehicle_from_import(conn, values):
    existing = conn.execute(
        "SELECT id FROM viaturas WHERE matricula = ?",
        (values["matricula"],)
    ).fetchone()

    fields = [
        "rentway_unitnr", "vin", "marca", "modelo", "versao", "motorizacao",
        "combustivel", "caixa", "ano", "cor", "data_matricula", "lugares",
        "data_compra", "km_atual", "km_compra", "proxima_revisao_km",
        "proxima_revisao_data", "ultima_revisao", "ultima_revisao_km", "ultimo_servico_feito",
        "campanhas_pendentes", "risco_tecnico", "observacoes", "ativo",
        "estado_frota", "data_venda", "valor_venda", "valor_venda_sem_iva",
        "iva_venda", "valor_iva_venda", "comprador", "motivo_venda",
        "data_baixa", "ativa_operacional", "estado_operacional",
        "current_status_rentway", "status_rentway", "fleet", "rental_station",
        "grupo", "categoria", "fornecedor", "valor_compra", "iva_compra",
        "valor_iva_compra", "valor_compra_com_iva", "valor_venda_com_iva",
        "preco_venda_rentway", "preco_venda_retalho_rentway",
        "documento_rentway", "documento_venda", "data_fatura_venda",
        "data_retorno", "bloqueio_rentway", "data_bloqueio", "cliente_rentway",
        "programa_venda", "data_inspecao", "ald_nr_pagamentos",
        "ald_primeiro_pagamento", "ald_valor_residual",
        "depreciacao_codigo_fiscal_rentway",
        "depreciacao_ultima_calculo_rentway", "codigo_conta_rentway",
        "data_ultimo_movimento", "data_bloqueio_peritagem",
        "fornecedor_financeiro", "contrato_financeiro", "contrato_financeiro_fim",
        "garantia_fim", "furto_data_declaracao", "furto_numero_declaracao",
        "aviso_viatura", "documentnr_rentway", "cliente_atual_rentway",
        "condutor_atual_rentway",
    ]

    if existing:
        assignments = ", ".join(f"{field} = ?" for field in fields)
        conn.execute(
            f"UPDATE viaturas SET {assignments} WHERE matricula = ?",
            [values[field] for field in fields] + [values["matricula"]]
        )
        return "updated"

    insert_fields = ["matricula", *fields, "criado_em"]
    placeholders = ", ".join("?" for _ in insert_fields)
    conn.execute(
        f"INSERT INTO viaturas ({', '.join(insert_fields)}) VALUES ({placeholders})",
        [values.get(field) for field in insert_fields[:-1]] + [now()]
    )
    return "created"


def import_frota_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Vehicles"] if "Vehicles" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = build_column_lookup(headers)

    stats = {
        "criadas": 0,
        "atualizadas": 0,
        "vendidas": 0,
        "ativas": 0,
        "ignoradas": 0,
        "total_linhas": 0,
    }

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Frota Total",
            Path(path).name,
            original_name,
            ws.title,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            stats["total_linhas"] += 1
            raw = {
                headers[idx] or f"coluna_{idx + 1}": value
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            values = build_vehicle_values(row, col)
            matricula = values["matricula"] if values else clean_text(raw.get("platenr"))

            c.execute("""
                INSERT INTO importacoes_linhas_raw
                (importacao_id, row_number, matricula, raw_json)
                VALUES (?, ?, ?, ?)
            """, (
                importacao_id,
                row_number,
                matricula,
                json.dumps(raw, ensure_ascii=False, default=str)
            ))

            if not values:
                stats["ignoradas"] += 1
                continue

            result = upsert_vehicle_from_import(c, values)
            if result == "created":
                stats["criadas"] += 1
            else:
                stats["atualizadas"] += 1

            if values["estado_frota"] == "Vendida":
                stats["vendidas"] += 1
            else:
                stats["ativas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                vendidas = ?,
                ativas = ?,
                ignoradas = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["vendidas"],
            stats["ativas"],
            stats["ignoradas"],
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def build_impro_values(row, col, raw, filename, row_number):
    matricula = clean_text(first_row_value(row, col, [
        "platenr", "matricula", "licenseplate", "plate", "registration"
    ]))
    if matricula:
        matricula = matricula.upper().replace(" ", "")

    numero_impro = clean_text(first_row_value(row, col, [
        "numero_impro", "n_impro", "impro", "impronr", "impro_nr", "orderid",
        "workorder", "work_order", "processo", "nrprocesso", "document_nr"
    ]))

    raw_json = json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True)
    raw_hash = hashlib.sha1(raw_json.encode("utf-8")).hexdigest()
    rentway_key = f"impro:{numero_impro}" if numero_impro else f"rowhash:{raw_hash}"

    return {
        "rentway_key": rentway_key,
        "matricula": matricula,
        "numero_impro": numero_impro,
        "marca": clean_text(first_row_value(row, col, ["marca", "brand", "brandid"])),
        "modelo": clean_text(first_row_value(row, col, ["modelo", "model", "modelid"])),
        "grupo": clean_text(first_row_value(row, col, ["grupo", "group", "groupid"])),
        "condutor": clean_text(first_row_value(row, col, ["condutor", "driver"])),
        "data_abertura": excel_date_to_iso(first_row_value(row, col, [
            "data_abertura", "opening_date", "open_date", "start_date", "date_start",
            "date_in", "entry_date", "created_date", "date"
        ])),
        "data_fecho": excel_date_to_iso(first_row_value(row, col, [
            "data_fecho", "closing_date", "close_date", "end_date", "date_end",
            "date_out", "exit_date"
        ])),
        "data_in": excel_date_to_iso(first_row_value(row, col, ["data_in", "date_in", "entry_date"])),
        "data_out": parse_excel_datetime(first_row_value(row, col, ["data_out", "date_out", "exit_date"])),
        "estado": clean_text(first_row_value(row, col, [
            "estado", "status", "state", "currentstatus", "current_status"
        ])),
        "estacao": clean_text(first_row_value(row, col, [
            "rental_station", "station", "estacao", "location", "branch"
        ])),
        "estacao_out": clean_text(first_row_value(row, col, ["estacao_out", "station_out", "checkout_station"])),
        "km": clean_int(first_row_value(row, col, [
            "kms", "km", "kilometers", "quilometros", "mileage"
        ])),
        "driven_kms": clean_int(clean_number_text(first_row_value(row, col, ["driven_kms", "drivenkms"]))),
        "oficina": clean_text(first_row_value(row, col, ["oficina", "workshop", "garage"])),
        "motivo": clean_text(first_row_value(row, col, [
            "motivo", "reason", "reason_code", "fault", "avaria", "tipo"
        ])),
        "descricao": clean_text(first_row_value(row, col, [
            "descricao", "description", "observations", "obs", "notes", "comment",
            "problem", "complaint"
        ])),
        "custo": clean_float(first_row_value(row, col, [
            "custo", "cost", "amount", "value", "total", "expenses"
        ])),
        "dias_imobilizado": clean_int(first_row_value(row, col, [
            "dias_imobilizado", "days", "downtime_days", "immobilized_days"
        ])),
        "raw_json": raw_json,
    }


def upsert_rentway_impro(conn, importacao_id, row_number, values):
    viatura_id = None
    if values["matricula"]:
        vehicle = conn.execute(
            "SELECT id FROM viaturas WHERE matricula = ?",
            (values["matricula"],)
        ).fetchone()
        viatura_id = vehicle["id"] if vehicle else None

    conn.execute("""
        INSERT INTO rentway_impros
        (rentway_key, importacao_id, row_number, viatura_id, matricula, numero_impro,
         marca, modelo, grupo, condutor, data_abertura, data_fecho, data_in, data_out,
         estado, estacao, estacao_out, km, driven_kms, oficina, motivo, descricao, custo,
         dias_imobilizado, raw_json, criado_em)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(rentway_key) DO UPDATE SET
            importacao_id = excluded.importacao_id,
            row_number = excluded.row_number,
            viatura_id = excluded.viatura_id,
            matricula = excluded.matricula,
            numero_impro = excluded.numero_impro,
            marca = excluded.marca,
            modelo = excluded.modelo,
            grupo = excluded.grupo,
            condutor = excluded.condutor,
            data_abertura = excluded.data_abertura,
            data_fecho = excluded.data_fecho,
            data_in = excluded.data_in,
            data_out = excluded.data_out,
            estado = excluded.estado,
            estacao = excluded.estacao,
            estacao_out = excluded.estacao_out,
            km = excluded.km,
            driven_kms = excluded.driven_kms,
            oficina = excluded.oficina,
            motivo = excluded.motivo,
            descricao = excluded.descricao,
            custo = excluded.custo,
            dias_imobilizado = excluded.dias_imobilizado,
            raw_json = excluded.raw_json
    """, (
        values["rentway_key"],
        importacao_id,
        row_number,
        viatura_id,
        values["matricula"],
        values["numero_impro"],
        values.get("marca"),
        values.get("modelo"),
        values.get("grupo"),
        values.get("condutor"),
        values["data_abertura"],
        values["data_fecho"],
        values.get("data_in"),
        values.get("data_out"),
        values["estado"],
        values["estacao"],
        values.get("estacao_out"),
        values["km"],
        values.get("driven_kms"),
        values.get("oficina"),
        values["motivo"],
        values["descricao"],
        values["custo"],
        values["dias_imobilizado"],
        values["raw_json"],
        now()
    ))


def import_impros_excel(path, original_name, responsavel=None):
    sheet_name, rows = read_spreadsheet_rows(path)
    headers = [str(c).strip() if c not in (None, "") else "" for c in rows[0]] if rows else []
    col = build_column_lookup(headers)
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0}

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Histórico Impros",
            Path(path).name,
            original_name,
            sheet_name,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        is_closed_impros = any(
            clean_text(row[0]) == "Impros fechados"
            for row in rows[:5]
            if row
        )

        if is_closed_impros:
            current_data_in = None
            current_record = None

            def flush_current():
                if not current_record:
                    return
                existed = c.execute(
                    "SELECT id FROM rentway_impros WHERE rentway_key = ?",
                    (current_record["rentway_key"],)
                ).fetchone()
                upsert_rentway_impro(c, importacao_id, current_record["row_number"], current_record)
                if existed:
                    stats["atualizadas"] += 1
                else:
                    stats["criadas"] += 1

            for idx, row in enumerate(rows, start=1):
                stats["total_linhas"] += 1
                label = clean_text(row[0] if len(row) > 0 else None)
                raw = {
                    f"coluna_{col_idx + 1}": value
                    for col_idx, value in enumerate(row)
                    if value not in (None, "")
                }

                if label == "Data in :":
                    flush_current()
                    current_record = None
                    current_data_in = excel_date_to_iso(row[2] if len(row) > 2 else None)
                    continue

                if label == "Oficina" and current_record:
                    current_record["oficina"] = clean_text(row[2] if len(row) > 2 else None)
                    current_record["raw_json"] = json.dumps({**json.loads(current_record["raw_json"]), **raw}, ensure_ascii=False, default=str, sort_keys=True)
                    continue

                if label == "Driven kms :" and current_record:
                    current_record["driven_kms"] = clean_int(clean_number_text(row[2] if len(row) > 2 else None))
                    current_record["km"] = current_record["driven_kms"]
                    current_record["raw_json"] = json.dumps({**json.loads(current_record["raw_json"]), **raw}, ensure_ascii=False, default=str, sort_keys=True)
                    continue

                numero = clean_text(row[0] if len(row) > 0 else None)
                matricula = clean_text(row[2] if len(row) > 2 else None)
                if not numero or not numero.replace(".", "", 1).isdigit() or not matricula:
                    stats["ignoradas"] += 1
                    continue

                flush_current()
                numero_impro = str(int(float(numero)))
                matricula = matricula.upper().replace(" ", "")
                raw_json = json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True)
                current_record = {
                    "row_number": idx,
                    "rentway_key": f"impro:{numero_impro}",
                    "matricula": matricula,
                    "numero_impro": numero_impro,
                    "marca": clean_text(row[3] if len(row) > 3 else None),
                    "modelo": clean_text(row[4] if len(row) > 4 else None),
                    "grupo": clean_text(row[7] if len(row) > 7 else None),
                    "condutor": clean_text(row[5] if len(row) > 5 else None),
                    "data_abertura": None,
                    "data_fecho": current_data_in,
                    "data_in": current_data_in,
                    "data_out": parse_excel_datetime(row[9] if len(row) > 9 else None),
                    "estado": "Fechado",
                    "estacao": None,
                    "estacao_out": clean_text(row[8] if len(row) > 8 else None),
                    "km": None,
                    "driven_kms": None,
                    "oficina": None,
                    "motivo": None,
                    "descricao": None,
                    "custo": None,
                    "dias_imobilizado": None,
                    "raw_json": raw_json,
                }

            flush_current()
        else:
            for row_number, row in enumerate(rows[1:], start=2):
                stats["total_linhas"] += 1
                raw = {
                    headers[idx] or f"coluna_{idx + 1}": value
                    for idx, value in enumerate(row)
                    if idx < len(headers)
                }
                values = build_impro_values(row, col, raw, Path(path).name, row_number)
                if not values["matricula"] and not values["numero_impro"]:
                    stats["ignoradas"] += 1
                    continue

                existed = c.execute(
                    "SELECT id FROM rentway_impros WHERE rentway_key = ?",
                    (values["rentway_key"],)
                ).fetchone()
                upsert_rentway_impro(c, importacao_id, row_number, values)
                if existed:
                    stats["atualizadas"] += 1
                else:
                    stats["criadas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def safe_cell(value):
    if value is None:
        return None
    try:
        if value != value:
            return None
    except Exception:
        pass
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def build_fo_key(row):
    parts = [
        row.get("matricula"),
        row.get("data_documento"),
        row.get("fornecedor"),
        row.get("numero_fatura"),
        row.get("descricao"),
        row.get("valor_com_iva"),
    ]
    raw = "|".join(str(safe_cell(part) or "") for part in parts)
    return "fo:" + hashlib.sha1(raw.encode("utf-8")).hexdigest()


def upsert_rentway_fo(conn, importacao_id, row):
    matricula = row.get("matricula")
    viatura_id = None
    if matricula:
        vehicle = conn.execute(
            "SELECT id FROM viaturas WHERE matricula = ?",
            (matricula,)
        ).fetchone()
        viatura_id = vehicle["id"] if vehicle else None

    raw_json = json.dumps(
        {key: safe_cell(value) for key, value in row.items()},
        ensure_ascii=False,
        default=str,
        sort_keys=True
    )
    rentway_key = build_fo_key(row)

    fields = [
        "rentway_key", "importacao_id", "row_number", "viatura_id",
        "folha_obra_origem", "data_documento", "matricula", "modelo",
        "fornecedor", "numero_fatura", "estacao", "kms", "tipo_servico",
        "natureza", "quantidade", "valor_com_iva", "descricao",
        "numero_documento", "estado_importacao", "alertas", "raw_json",
        "criado_em",
    ]
    values = {
        "rentway_key": rentway_key,
        "importacao_id": importacao_id,
        "row_number": int(row.get("linha_origem") or 0),
        "viatura_id": viatura_id,
        "folha_obra_origem": safe_cell(row.get("folha_obra_origem")),
        "data_documento": safe_cell(row.get("data_documento")),
        "matricula": matricula,
        "modelo": safe_cell(row.get("modelo")),
        "fornecedor": safe_cell(row.get("fornecedor")),
        "numero_fatura": safe_cell(row.get("numero_fatura")),
        "estacao": safe_cell(row.get("estacao")),
        "kms": safe_cell(row.get("kms")),
        "tipo_servico": safe_cell(row.get("tipo_servico")),
        "natureza": safe_cell(row.get("natureza")),
        "quantidade": safe_cell(row.get("quantidade")),
        "valor_com_iva": safe_cell(row.get("valor_com_iva")),
        "descricao": safe_cell(row.get("descricao")),
        "numero_documento": safe_cell(row.get("numero_documento")),
        "estado_importacao": safe_cell(row.get("estado_importacao")),
        "alertas": safe_cell(row.get("alertas")),
        "raw_json": raw_json,
        "criado_em": now(),
    }

    existed = conn.execute(
        "SELECT id FROM rentway_folhas_obra WHERE rentway_key = ?",
        (rentway_key,)
    ).fetchone()

    placeholders = ", ".join("?" for _ in fields)
    update_fields = [field for field in fields if field not in ("rentway_key", "criado_em")]
    updates = ", ".join(f"{field} = excluded.{field}" for field in update_fields)
    conn.execute(
        f"""
        INSERT INTO rentway_folhas_obra ({', '.join(fields)})
        VALUES ({placeholders})
        ON CONFLICT(rentway_key) DO UPDATE SET {updates}
        """,
        [values[field] for field in fields]
    )
    return "updated" if existed else "created"


def import_folhas_obra_excel(path, original_name, responsavel=None):
    from preparar_importacao_fo import (
        clean_data,
        detect_duplicates,
        normalize_columns,
        read_file,
        set_import_status,
        validate_data,
    )

    df = read_file(path)
    df = normalize_columns(df)
    df = clean_data(df)
    df, alerts_df = validate_data(df)
    df, alerts_df = detect_duplicates(df, alerts_df)
    df = set_import_status(df)

    importable = df[df["estado_importacao"].isin(["OK", "VALIDAR"])].copy()
    stats = {
        "total_linhas": len(df),
        "criadas": 0,
        "atualizadas": 0,
        "ignoradas": len(df) - len(importable),
        "alertas": len(alerts_df),
    }

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em,
             total_linhas, ignoradas)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            "Histórico FO",
            Path(path).name,
            original_name,
            "Folha1",
            responsavel,
            json.dumps(list(df.columns), ensure_ascii=False),
            now(),
            stats["total_linhas"],
            stats["ignoradas"],
        ))
        importacao_id = cur.lastrowid

        for _, row in importable.iterrows():
            result = upsert_rentway_fo(c, importacao_id, row.to_dict())
            if result == "created":
                stats["criadas"] += 1
            else:
                stats["atualizadas"] += 1

        c.execute("""
            UPDATE importacoes SET
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"Alertas no ficheiro: {stats['alertas']}",
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def parse_money(value):
    if value in (None, ""):
        return 0.0
    text = clean_text(value)
    if not text or text in ("-", "-   "):
        return 0.0
    text = text.replace("€", "").replace("\xa0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def import_sinistros_allianz_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [clean_text(c.value) for c in ws[1]]
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0}
    agregados = {}

    def add_unique(current, value, sep=" | "):
        value = clean_text(value)
        if not value:
            return current
        parts = [p.strip() for p in str(current or "").split(sep) if p.strip()]
        if value not in parts:
            parts.append(value)
        return sep.join(parts)

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stats["total_linhas"] += 1
        raw = {
            headers[idx] or f"coluna_{idx + 1}": value
            for idx, value in enumerate(row)
            if idx < len(headers)
        }
        matricula = (clean_text(row[1] if len(row) > 1 else None) or "").upper().replace(" ", "")
        referencia = clean_text(row[2] if len(row) > 2 else None)
        data_sinistro = excel_date_to_iso(row[3] if len(row) > 3 else None)
        if not matricula and not referencia:
            stats["ignoradas"] += 1
            continue

        if matricula and data_sinistro:
            allianz_key = f"allianz:{matricula}:{data_sinistro}"
        else:
            allianz_key = "allianz:" + hashlib.sha1(
                json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True).encode("utf-8")
            ).hexdigest()

        alertas = []
        if not matricula:
            alertas.append("Matricula em falta")
        if not referencia:
            alertas.append("Referencia de sinistro em falta")
        if not data_sinistro:
            alertas.append("Data de sinistro em falta")

        values = agregados.setdefault(allianz_key, {
            "allianz_key": allianz_key,
            "importacao_id": None,
            "row_number": row_number,
            "viatura_id": None,
            "adesao": clean_text(row[0] if len(row) > 0 else None),
            "matricula": matricula,
            "referencia_sinistro": None,
            "data_sinistro": data_sinistro,
            "data_encerramento": excel_date_to_iso(row[4] if len(row) > 4 else None),
            "causa": None,
            "valor_rc": 0.0,
            "valor_dp": 0.0,
            "valor_ids_credor": 0.0,
            "valor_vidros": 0.0,
            "custos_gestao": 0.0,
            "custo_total": 0.0,
            "estado_importacao": "OK",
            "alertas": None,
            "raw_json": "",
            "criado_em": now(),
        })
        values["referencia_sinistro"] = add_unique(values["referencia_sinistro"], referencia)
        values["causa"] = add_unique(values["causa"], clean_text(row[5] if len(row) > 5 else None))
        values["alertas"] = add_unique(values["alertas"], "; ".join(alertas), sep="; ")
        values["estado_importacao"] = "VALIDAR" if values["alertas"] else "OK"
        values["valor_rc"] += parse_money(row[6] if len(row) > 6 else None)
        values["valor_dp"] += parse_money(row[7] if len(row) > 7 else None)
        values["valor_ids_credor"] += parse_money(row[8] if len(row) > 8 else None)
        values["valor_vidros"] += parse_money(row[9] if len(row) > 9 else None)
        values["custos_gestao"] += parse_money(row[10] if len(row) > 10 else None)
        values["custo_total"] += parse_money(row[11] if len(row) > 11 else None)
        values["raw_json"] = add_unique(
            values["raw_json"],
            json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True),
            sep="\n"
        )

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Sinistros Allianz",
            Path(path).name,
            original_name,
            ws.title,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        for values in agregados.values():
            viatura = c.execute(
                "SELECT id FROM viaturas WHERE matricula = ?",
                (values["matricula"],)
            ).fetchone() if values["matricula"] else None
            values["importacao_id"] = importacao_id
            values["viatura_id"] = viatura["id"] if viatura else None
            fields = list(values.keys())
            placeholders = ", ".join("?" for _ in fields)
            updates = ", ".join(
                f"{field} = excluded.{field}"
                for field in fields
                if field not in ("allianz_key", "criado_em")
            )
            existed = c.execute(
                "SELECT id FROM sinistros_allianz WHERE allianz_key = ?",
                (values["allianz_key"],)
            ).fetchone()
            c.execute(
                f"""
                INSERT INTO sinistros_allianz ({', '.join(fields)})
                VALUES ({placeholders})
                ON CONFLICT(allianz_key) DO UPDATE SET {updates}
                """,
                [values[field] for field in fields]
            )
            if existed:
                stats["atualizadas"] += 1
            else:
                stats["criadas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def import_rentway_accident_reports_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row = None
    headers = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [normalize_header(value) for value in row]
        if "accidentreportid" in normalized and "platenumber" in normalized:
            header_row = idx
            headers = [clean_text(value) for value in row]
            break
    if not header_row:
        raise RuntimeError("Não encontrei a linha de cabeçalho com accidentReportID e plateNumber.")

    col = build_column_lookup(headers)
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0, "alertas": 0}

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Histórico AR Rentway",
            Path(path).name,
            original_name,
            ws.title,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            stats["total_linhas"] += 1
            accident_report_id = clean_text(first_row_value(row, col, ["accidentReportID"]))
            matricula = (clean_text(first_row_value(row, col, ["plateNumber"])) or "").upper().replace(" ", "")
            document_no = clean_text(first_row_value(row, col, ["documentNo"]))
            accident_date = parse_excel_datetime(first_row_value(row, col, ["accidentDate"]))

            if not accident_report_id and not matricula and not document_no:
                stats["ignoradas"] += 1
                continue

            alertas = []
            if not accident_report_id:
                alertas.append("AR sem accidentReportID")
            if not matricula:
                alertas.append("Matrícula em falta")
            if not accident_date:
                alertas.append("Data de acidente/reparação em falta")

            viatura = c.execute(
                "SELECT id FROM viaturas WHERE matricula = ?",
                (matricula,)
            ).fetchone() if matricula else None

            raw = {
                headers[idx] or f"coluna_{idx + 1}": value
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            if accident_report_id:
                ar_key = f"rentway-ar:{accident_report_id}"
            else:
                ar_key = "rentway-ar:" + hashlib.sha1(
                    json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True).encode("utf-8")
                ).hexdigest()

            values = {
                "ar_key": ar_key,
                "importacao_id": importacao_id,
                "row_number": row_number,
                "viatura_id": viatura["id"] if viatura else None,
                "accident_report_id": accident_report_id,
                "document_type": clean_text(first_row_value(row, col, ["documentType"])),
                "document_no": document_no,
                "state": clean_text(first_row_value(row, col, ["state"])),
                "request_date": parse_excel_datetime(first_row_value(row, col, ["requestDate"])),
                "accident_date": accident_date,
                "friendly_declaration": clean_text(first_row_value(row, col, ["friendlyDeclaration"])),
                "matricula": matricula,
                "unit_no": clean_text(first_row_value(row, col, ["unitNo"])),
                "insurance_policy": clean_text(first_row_value(row, col, ["insurancePolicy"])),
                "rental_station_out": clean_text(first_row_value(row, col, ["rentalStationOut"])),
                "created_by_rental_station": clean_text(first_row_value(row, col, ["createdByRentalSation", "createdByRentalStation"])),
                "status": clean_text(first_row_value(row, col, ["Status"])),
                "estado_importacao": "VALIDAR" if alertas else "OK",
                "alertas": "; ".join(alertas),
                "raw_json": json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True),
                "criado_em": now(),
            }
            fields = list(values.keys())
            placeholders = ", ".join("?" for _ in fields)
            updates = ", ".join(
                f"{field} = excluded.{field}"
                for field in fields
                if field not in ("ar_key", "criado_em")
            )
            existed = c.execute(
                "SELECT id FROM rentway_accident_reports WHERE ar_key = ?",
                (ar_key,)
            ).fetchone()
            c.execute(
                f"""
                INSERT INTO rentway_accident_reports ({', '.join(fields)})
                VALUES ({placeholders})
                ON CONFLICT(ar_key) DO UPDATE SET {updates}
                """,
                [values[field] for field in fields]
            )
            if existed:
                stats["atualizadas"] += 1
            else:
                stats["criadas"] += 1
            if alertas:
                stats["alertas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"{stats['alertas']} registos a validar",
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def parse_supplier_number(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("€", "").replace("\xa0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_service_text(value):
    text = str(value or "").lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def position_from_service_text(text):
    if any(term in text for term in ("frente", "dianteir", "front")):
        return "frente"
    if any(term in text for term in ("tras", "traseir", "posterior", "rear")):
        return "tras"
    return None


def detect_supplier_services(lines):
    found = []

    def add(label):
        if label not in found:
            found.append(label)

    for row in lines:
        text = normalize_service_text(
            " ".join([
                row["descricao"] or "",
                row["referencia"] or "",
                row["ct"] or "",
            ])
        )
        if any(term in text for term in ("oleo", "oil", "lubrificante")):
            add("Oleo")
        position = position_from_service_text(text)
        if "pneu" in text:
            add(f"Pneus {position}" if position else "Pneus")
        if any(term in text for term in ("calco", "calcos", "pastilha", "pastilhas")):
            add(f"Calcos {position}" if position else "Calcos")
        if "disco" in text:
            add(f"Discos {position}" if position else "Discos")

    return found


def import_faturas_fornecedores_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = "Importacao" if "Importacao" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows_iter)]
    col = build_column_lookup(headers)
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0, "alertas": 0}

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Faturas Fornecedores",
            Path(path).name,
            original_name,
            sheet_name,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        for row_number, row in enumerate(rows_iter, start=2):
            stats["total_linhas"] += 1
            fornecedor = clean_text(first_row_value(row, col, ["fornecedor"]))
            documento = clean_text(first_row_value(row, col, ["documento"]))
            matricula = (clean_text(first_row_value(row, col, ["matricula"])) or "").upper().replace(" ", "")
            descricao = clean_text(first_row_value(row, col, ["descricao"]))
            referencia = clean_text(first_row_value(row, col, ["referencia"]))
            total_liq = parse_supplier_number(first_row_value(row, col, ["total_liq"]))
            fonte_pdf = clean_text(first_row_value(row, col, ["fonte_pdf"]))

            if not any((fornecedor, documento, matricula, descricao, referencia, total_liq, fonte_pdf)):
                stats["ignoradas"] += 1
                continue

            alertas = []
            if not fornecedor:
                alertas.append("Fornecedor em falta")
            if not documento:
                alertas.append("Documento em falta")
            if not matricula:
                alertas.append("Matrícula em falta")
            if not descricao:
                alertas.append("Descrição em falta")
            if total_liq is None:
                alertas.append("Valor total líquido em falta/inválido")

            viatura = c.execute(
                "SELECT id FROM viaturas WHERE matricula = ?",
                (matricula,)
            ).fetchone() if matricula else None

            raw = {
                headers[idx] or f"coluna_{idx + 1}": value
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            key_payload = {
                "fornecedor": fornecedor,
                "documento": documento,
                "data_doc": parse_excel_datetime(first_row_value(row, col, ["data_doc"])),
                "matricula": matricula,
                "or_reparacao": clean_text(first_row_value(row, col, ["or_reparacao"])),
                "descricao": descricao,
                "referencia": referencia,
                "total_liq": total_liq,
                "ct": clean_text(first_row_value(row, col, ["ct"])),
                "fonte_pdf": fonte_pdf,
            }
            linha_key = "fornecedor-fatura:" + hashlib.sha1(
                json.dumps(key_payload, ensure_ascii=False, default=str, sort_keys=True).encode("utf-8")
            ).hexdigest()

            values = {
                "linha_key": linha_key,
                "importacao_id": importacao_id,
                "row_number": row_number,
                "viatura_id": viatura["id"] if viatura else None,
                "fornecedor": fornecedor,
                "documento": documento,
                "data_doc": key_payload["data_doc"],
                "matricula": matricula,
                "kms": parse_supplier_number(first_row_value(row, col, ["kms"])),
                "or_reparacao": key_payload["or_reparacao"],
                "chassis": clean_text(first_row_value(row, col, ["chassis"])),
                "descricao": descricao,
                "referencia": referencia,
                "pv_unit": parse_supplier_number(first_row_value(row, col, ["pv_unit"])),
                "desc_pct": parse_supplier_number(first_row_value(row, col, ["desc_pct"])),
                "pliq_unit": parse_supplier_number(first_row_value(row, col, ["pliq_unit"])),
                "quantidade": parse_supplier_number(first_row_value(row, col, ["tmp_qt", "quantidade"])),
                "total_liq": total_liq,
                "ct": key_payload["ct"],
                "fonte_pdf": fonte_pdf,
                "estado_importacao": "VALIDAR" if alertas else "OK",
                "alertas": "; ".join(alertas),
                "raw_json": json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True),
                "criado_em": now(),
            }
            fields = list(values.keys())
            placeholders = ", ".join("?" for _ in fields)
            updates = ", ".join(
                f"{field} = excluded.{field}"
                for field in fields
                if field not in ("linha_key", "criado_em")
            )
            existed = c.execute(
                "SELECT id FROM faturas_fornecedores_linhas WHERE linha_key = ?",
                (linha_key,)
            ).fetchone()
            c.execute(
                f"""
                INSERT INTO faturas_fornecedores_linhas ({', '.join(fields)})
                VALUES ({placeholders})
                ON CONFLICT(linha_key) DO UPDATE SET {updates}
                """,
                [values[field] for field in fields]
            )
            if existed:
                stats["atualizadas"] += 1
            else:
                stats["criadas"] += 1
            if alertas:
                stats["alertas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"{stats['alertas']} linhas a validar",
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def import_contratos_financeiro_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    headers = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [normalize_header(value) for value in row]
        if "contrato" in normalized and "dataout" in normalized and "matricula" in normalized:
            header_row = idx
            headers = [clean_text(value) for value in row]
            break
    if not header_row:
        raise RuntimeError("Não encontrei cabeçalho com Contrato, Data out e Matrícula.")

    col = build_column_lookup(headers)
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0, "alertas": 0}

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Contratos Financeiro",
            Path(path).name,
            original_name,
            ws.title,
            responsavel,
            json.dumps(headers, ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid

        for row_number, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            stats["total_linhas"] += 1
            contrato_nr = clean_text(first_row_value(row, col, ["Contrato"]))
            matricula_final = (clean_text(first_row_value(row, col, ["Matrícula", "Matricula"])) or "").upper().replace(" ", "")
            if not contrato_nr and not matricula_final:
                stats["ignoradas"] += 1
                continue

            alertas = []
            if not contrato_nr:
                alertas.append("Contrato em falta")
            if not matricula_final:
                alertas.append("Matrícula final em falta")

            viatura = c.execute(
                "SELECT id FROM viaturas WHERE matricula = ?",
                (matricula_final,)
            ).fetchone() if matricula_final else None

            raw = {
                headers[idx] or f"coluna_{idx + 1}": value
                for idx, value in enumerate(row)
                if idx < len(headers)
            }
            values = {
                "contrato_key": f"contrato-financeiro:{contrato_nr}",
                "importacao_id": importacao_id,
                "row_number": row_number,
                "viatura_id": viatura["id"] if viatura else None,
                "contrato_nr": contrato_nr,
                "data_out": parse_excel_datetime(first_row_value(row, col, ["Data out"])),
                "data_in": parse_excel_datetime(first_row_value(row, col, ["Data in"])),
                "data_fim": parse_excel_datetime(first_row_value(row, col, ["Data de fim"])),
                "dias": parse_supplier_number(first_row_value(row, col, ["Dias"])),
                "valor_dia": parse_supplier_number(first_row_value(row, col, ["Valor dia"])),
                "valor_fatura": parse_supplier_number(first_row_value(row, col, ["Fatura"])),
                "valor_fatura_caixa": parse_supplier_number(first_row_value(row, col, ["Fatura (caixa)"])),
                "valor_nota_debito": parse_supplier_number(first_row_value(row, col, ["Nota de débito", "Nota de debito"])),
                "valor_nota_credito": parse_supplier_number(first_row_value(row, col, ["Nota de crédito", "Nota de credito"])),
                "valor_total": parse_supplier_number(first_row_value(row, col, ["Total"])),
                "cliente_id": clean_text(first_row_value(row, col, ["ID do cliente"])),
                "tipo_cliente": clean_text(first_row_value(row, col, ["Tipo de cliente"])),
                "origem": clean_text(first_row_value(row, col, ["Origem"])),
                "grupo_reservado": clean_text(first_row_value(row, col, ["Grupo reservado"])),
                "grupo_entregue": clean_text(first_row_value(row, col, ["Grupo entregue"])),
                "matricula_final": matricula_final,
                "vendedor": clean_text(first_row_value(row, col, ["Vendedor"])),
                "nome_vendedor": clean_text(first_row_value(row, col, ["Nome do vendedor"])),
                "estado_importacao": "VALIDAR" if alertas else "OK",
                "alertas": "; ".join(alertas),
                "raw_json": json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True),
                "criado_em": now(),
            }
            fields = list(values.keys())
            placeholders = ", ".join("?" for _ in fields)
            updates = ", ".join(
                f"{field} = excluded.{field}"
                for field in fields
                if field not in ("contrato_key", "criado_em")
            )
            existed = c.execute(
                "SELECT id FROM contratos_financeiro WHERE contrato_key = ?",
                (values["contrato_key"],)
            ).fetchone()
            c.execute(
                f"""
                INSERT INTO contratos_financeiro ({', '.join(fields)})
                VALUES ({placeholders})
                ON CONFLICT(contrato_key) DO UPDATE SET {updates}
                """,
                [values[field] for field in fields]
            )
            if existed:
                stats["atualizadas"] += 1
            else:
                stats["criadas"] += 1
            if alertas:
                stats["alertas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"{stats['alertas']} contratos a validar",
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def import_contratos_viaturas_excel(path, original_name, responsavel=None):
    if load_workbook is None:
        raise RuntimeError("Falta instalar openpyxl para ler ficheiros Excel.")

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0, "alertas": 0}

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, responsavel, colunas_json, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Contratos Viaturas",
            Path(path).name,
            original_name,
            ws.title,
            responsavel,
            json.dumps(["Contrato nº", "Data out", "Data in", "Matrícula", "Cliente"], ensure_ascii=False),
            now()
        ))
        importacao_id = cur.lastrowid
        cliente_atual = None

        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values_row = [clean_text(value) for value in row]
            if len(values_row) > 3 and values_row[2] == "Cliente :":
                cliente_atual = values_row[3]
                continue

            contrato_nr = clean_text(row[0] if len(row) > 0 else None)
            if not contrato_nr or not contrato_nr.isdigit():
                continue

            stats["total_linhas"] += 1
            matricula = (clean_text(row[4] if len(row) > 4 else None) or "").upper().replace(" ", "")
            data_out = parse_excel_datetime(row[2] if len(row) > 2 else None)
            data_in = parse_excel_datetime(row[3] if len(row) > 3 else None)
            alertas = []
            if not matricula:
                alertas.append("Matrícula em falta")
            if not data_out:
                alertas.append("Data out em falta/inválida")
            if row[3] not in (None, "") and not data_in:
                alertas.append("Data in inválida")

            viatura = c.execute(
                "SELECT id FROM viaturas WHERE matricula = ?",
                (matricula,)
            ).fetchone() if matricula else None

            raw = {
                "Contrato nº": contrato_nr,
                "Data out": row[2] if len(row) > 2 else None,
                "Data in": row[3] if len(row) > 3 else None,
                "Matrícula": matricula,
                "Cliente": cliente_atual,
            }
            key_payload = {
                "contrato": contrato_nr,
                "matricula": matricula,
                "data_out": data_out,
                "data_in": data_in,
                "cliente": cliente_atual,
            }
            movimento_key = "contrato-viatura:" + hashlib.sha1(
                json.dumps(key_payload, ensure_ascii=False, default=str, sort_keys=True).encode("utf-8")
            ).hexdigest()
            values = {
                "movimento_key": movimento_key,
                "importacao_id": importacao_id,
                "row_number": row_number,
                "viatura_id": viatura["id"] if viatura else None,
                "contrato_nr": contrato_nr,
                "cliente_nome": cliente_atual,
                "data_out": data_out,
                "data_in": data_in,
                "matricula": matricula,
                "estado_importacao": "VALIDAR" if alertas else "OK",
                "alertas": "; ".join(alertas),
                "raw_json": json.dumps(raw, ensure_ascii=False, default=str, sort_keys=True),
                "criado_em": now(),
            }
            fields = list(values.keys())
            placeholders = ", ".join("?" for _ in fields)
            updates = ", ".join(
                f"{field} = excluded.{field}"
                for field in fields
                if field not in ("movimento_key", "criado_em")
            )
            existed = c.execute(
                "SELECT id FROM contratos_viaturas WHERE movimento_key = ?",
                (movimento_key,)
            ).fetchone()
            c.execute(
                f"""
                INSERT INTO contratos_viaturas ({', '.join(fields)})
                VALUES ({placeholders})
                ON CONFLICT(movimento_key) DO UPDATE SET {updates}
                """,
                [values[field] for field in fields]
            )
            if existed:
                stats["atualizadas"] += 1
            else:
                stats["criadas"] += 1
            if alertas:
                stats["alertas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"{stats['alertas']} movimentos a validar",
            importacao_id
        ))

    stats["importacao_id"] = importacao_id
    return stats


def normalize_matricula(value):
    text = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    if len(text) != 6:
        return text or None
    if text[:2].isalpha() and text[2:4].isdigit() and text[4:].isalpha():
        return f"{text[:2]}-{text[2:4]}-{text[4:]}"
    if text[:2].isdigit() and text[2:4].isalpha() and text[4:].isdigit():
        return f"{text[:2]}-{text[2:4]}-{text[4:]}"
    if text[:4].isdigit() and text[4:].isalpha():
        return f"{text[:2]}-{text[2:4]}-{text[4:]}"
    if text[:2].isalpha() and text[2:].isdigit():
        return f"{text[:2]}-{text[2:4]}-{text[4:]}"
    return text


def extract_pdf_text(path):
    if PdfReader is None:
        raise RuntimeError("Falta instalar PyPDF2 para ler PDFs.")
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def first_regex(text, patterns):
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return clean_text(match.group(1))
    return None


def classify_diagnostic_pdf(filename, text):
    name = filename.lower()
    if "relatorio-de-diagnostico" in name or "relatório-de-diagnóstico" in name:
        return "diagnostico_autel"
    if "lubrificacao" in name or "lubrificação" in name:
        return "lubrificacao_motor"
    if "manutencao" in name or "manutenção" in name:
        return "informacoes_manutencao"
    if "dtc" in text.lower() or "relatório de diagnóstico" in text.lower():
        return "diagnostico_pdf"
    return "documento_tecnico_pdf"


def find_vehicle_by_pdf_data(conn, matricula, vin):
    if vin:
        row = conn.execute("SELECT * FROM viaturas WHERE vin = ?", (vin,)).fetchone()
        if row:
            return row
    if matricula:
        compact = matricula.replace("-", "")
        row = conn.execute("""
            SELECT *
            FROM viaturas
            WHERE matricula = ? OR REPLACE(matricula, '-', '') = ?
        """, (matricula, compact)).fetchone()
        if row:
            return row
    return None


def build_pdf_diagnostic_payload(path, text):
    filename = Path(path).name
    tipo = classify_diagnostic_pdf(filename, text)
    vin = first_regex(text, [r"\bVIN:\s*([A-Z0-9]{17})", r"\b([A-HJ-NPR-Z0-9]{17})\b"])
    matricula = normalize_matricula(first_regex(text, [
        r"Matr[íi]cula:\s*([A-Z0-9\- ]{6,10})",
        r"\b([A-Z]{2}[- ]?\d{2}[- ]?[A-Z]{2})\b",
        r"\b(\d{2}[- ]?[A-Z]{2}[- ]?\d{2})\b",
        r"\b(\d{2}[- ]?\d{2}[- ]?[A-Z]{2})\b",
    ]))
    data = parse_excel_datetime(first_regex(text, [
        r"Tempo de teste:\s*([0-9]{4}-[0-9]{2}-[0-9]{2}\s+[0-9]{2}:[0-9]{2}:[0-9]{2})",
        r"([0-9]{4}-[0-9]{2}-[0-9]{2})",
        r"([0-9]{2}/[0-9]{2}/[0-9]{4})",
    ]))
    if not data:
        file_date = first_regex(filename, [r"^([0-9]{4}-[0-9]{2}-[0-9]{2})"])
        data = parse_excel_datetime(file_date)
    km = clean_int(first_regex(text, [
        r"Quilometragem:\s*([0-9\.,]+)\s*km",
        r"\bKms?:\s*([0-9\.,]+)",
        r"\bKM:\s*([0-9\.,]+)",
    ]))
    numero_relatorio = first_regex(text, [
        r"N[úu]mero do relat[óo]rio:\s*([A-Z0-9\-]+)",
        r"N[úu]mero do pedido de repara[çc][ãa]o:\s*([A-Z0-9\-]+)",
    ])
    dtc_count = first_regex(text, [r"DTC\s*\(\s*([0-9]+)\s*\)"])
    sistemas_count = first_regex(text, [r"Sistemas analisados\s*\(\s*([0-9]+)\s*\)"])

    if tipo == "diagnostico_autel":
        problema = "Relatorio Autel"
        if sistemas_count or dtc_count:
            problema += f": {sistemas_count or '?'} sistemas analisados, {dtc_count or '?'} DTC"
    elif tipo == "informacoes_manutencao":
        problema = "Informacoes de manutencao importadas de PDF"
    elif tipo == "lubrificacao_motor":
        problema = "Informacoes de lubrificacao do motor importadas de PDF"
    else:
        problema = "Documento tecnico importado de PDF"

    return {
        "filename": filename,
        "tipo": tipo,
        "vin": vin,
        "matricula": matricula,
        "data_diagnostico": data,
        "km_diagnostico": km,
        "numero_relatorio": numero_relatorio,
        "problema_identificado": problema,
        "observacoes": text[:5000],
    }


def import_diagnosticos_pdf_paths(paths, original_source, responsavel=None, copy_files=True):
    stats = {"total_linhas": 0, "criadas": 0, "atualizadas": 0, "ignoradas": 0, "alertas": 0}
    pdf_paths = [Path(p) for p in paths if str(p).lower().endswith(".pdf")]
    if not pdf_paths:
        raise RuntimeError("Não encontrei PDFs para importar.")

    with db() as c:
        cur = c.execute("""
            INSERT INTO importacoes
            (tipo, filename, original_name, sheet_name, total_linhas, responsavel, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            "Diagnosticos PDF",
            Path(str(original_source)).name,
            str(original_source),
            "PDF",
            len(pdf_paths),
            responsavel,
            now(),
        ))
        importacao_id = cur.lastrowid

        for row_number, pdf_path in enumerate(pdf_paths, start=1):
            stats["total_linhas"] += 1
            try:
                text = extract_pdf_text(pdf_path)
            except Exception as exc:
                stats["ignoradas"] += 1
                stats["alertas"] += 1
                continue

            payload = build_pdf_diagnostic_payload(pdf_path, text)
            vehicle = find_vehicle_by_pdf_data(c, payload["matricula"], payload["vin"])
            if vehicle and not payload["matricula"]:
                payload["matricula"] = vehicle["matricula"]

            alertas_linha = []
            if not text.strip():
                alertas_linha.append("PDF sem texto extraivel")
            if not payload["matricula"] and not payload["vin"]:
                alertas_linha.append("sem matricula e sem VIN")
            if not payload["data_diagnostico"]:
                alertas_linha.append("data nao identificada")
            if payload["vin"] and not vehicle:
                alertas_linha.append("VIN nao encontrado na frota")

            if not payload["matricula"] and not vehicle:
                stats["ignoradas"] += 1
                stats["alertas"] += 1
                continue

            matricula = payload["matricula"] or vehicle["matricula"]
            data_diagnostico = payload["data_diagnostico"] or now()
            if vehicle:
                viatura_id = vehicle["id"]
            else:
                existing_vehicle = c.execute("SELECT id FROM viaturas WHERE matricula = ?", (matricula,)).fetchone()
                if existing_vehicle:
                    viatura_id = existing_vehicle["id"]
                else:
                    viatura_id = c.execute(
                        "INSERT INTO viaturas (matricula, vin, km_atual, criado_em) VALUES (?, ?, ?, ?)",
                        (matricula, payload["vin"], payload["km_diagnostico"], now())
                    ).lastrowid
            if copy_files:
                stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{secure_filename(pdf_path.name)}"
                stored_path = IMPORT_DIR / stored_name
                shutil.copy2(pdf_path, stored_path)
                fonte_path = str(stored_path)
                fonte_pdf = stored_name
            else:
                fonte_path = str(pdf_path)
                fonte_pdf = pdf_path.name

            text_hash = hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()
            import_key = "diag-pdf:" + hashlib.sha1(json.dumps({
                "vin": payload["vin"],
                "matricula": matricula,
                "data": data_diagnostico,
                "tipo": payload["tipo"],
                "filename": pdf_path.name,
                "hash": text_hash,
            }, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()

            processo_numero = f"PDF-DIAG-{import_key[-12:].upper()}"
            processo = c.execute("SELECT id FROM processos WHERE numero_impro = ?", (processo_numero,)).fetchone()
            if processo:
                processo_id = processo["id"]
            else:
                processo_id = c.execute("""
                    INSERT INTO processos
                    (numero_impro, viatura_id, matricula, km_abertura, responsavel, descricao_inicial,
                     estado, data_abertura, observacoes, origem, arquivo, visivel_gestao)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    processo_numero,
                    viatura_id,
                    matricula,
                    payload["km_diagnostico"],
                    responsavel or "Importacao PDF",
                    payload["problema_identificado"],
                    "Fechado",
                    data_diagnostico,
                    f"Processo criado automaticamente a partir do PDF {pdf_path.name}.",
                    "Importação Diagnósticos PDF",
                    1,
                    0,
                )).lastrowid

            values = {
                "processo_id": processo_id,
                "viatura_id": viatura_id,
                "tipo": payload["tipo"],
                "data_diagnostico": data_diagnostico,
                "km_diagnostico": payload["km_diagnostico"],
                "problema_identificado": payload["problema_identificado"],
                "observacoes": payload["observacoes"],
                "criado_em": now(),
                "import_key": import_key,
                "origem_importacao": str(original_source),
                "origem_linha": row_number,
                "estado_importacao": "VALIDAR" if alertas_linha else "OK",
                "alertas": "; ".join(alertas_linha),
                "fonte_pdf": fonte_pdf,
                "fonte_path": fonte_path,
                "vin": payload["vin"],
                "numero_relatorio": payload["numero_relatorio"],
                "texto_extraido": text[:20000],
            }
            existing = c.execute("SELECT id FROM diagnosticos WHERE import_key = ?", (import_key,)).fetchone()
            if existing:
                assignments = ", ".join(f"{field} = ?" for field in values)
                c.execute(f"UPDATE diagnosticos SET {assignments} WHERE id = ?", list(values.values()) + [existing["id"]])
                stats["atualizadas"] += 1
            else:
                fields = list(values.keys())
                placeholders = ", ".join("?" for _ in fields)
                c.execute(
                    f"INSERT INTO diagnosticos ({', '.join(fields)}) VALUES ({placeholders})",
                    [values[field] for field in fields]
                )
                stats["criadas"] += 1
            if alertas_linha:
                stats["alertas"] += 1

        c.execute("""
            UPDATE importacoes SET
                total_linhas = ?,
                criadas = ?,
                atualizadas = ?,
                ignoradas = ?,
                observacoes = ?
            WHERE id = ?
        """, (
            stats["total_linhas"],
            stats["criadas"],
            stats["atualizadas"],
            stats["ignoradas"],
            f"{stats['alertas']} PDFs a validar",
            importacao_id,
        ))

    stats["importacao_id"] = importacao_id
    return stats


@app.route("/")
def home():
    return render_template("home.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        with db() as c:
            user = c.execute("SELECT * FROM users WHERE email = ?", (email,)).fetchone()
            if not user or not user["ativo"] or not check_password_hash(user["password_hash"], password):
                flash("Email ou password inválidos.", "error")
                return redirect(url_for("login"))

            session.clear()
            session["user_id"] = user["id"]
            session["user_name"] = user["nome"]
            session["role"] = user["role"]
            session["departamento"] = user["departamento"] if "departamento" in user.keys() else None
            c.execute("UPDATE users SET ultimo_login = ? WHERE id = ?", (now(), user["id"]))
            log_action("login", "users", user["id"], f"Login de {user['email']}", conn=c)

        next_url = request.args.get("next") or url_for("home")
        return redirect(next_url)

    return render_template(
        "auth/login.html",
        default_email=DEFAULT_ADMIN_EMAIL,
        default_password=DEFAULT_ADMIN_PASSWORD,
    )


@app.route("/logout")
def logout():
    user_id = session.get("user_id")
    if user_id:
        log_action("logout", "users", user_id, f"Logout de {session.get('user_name')}")
    session.clear()
    return redirect(url_for("login"))


@app.route("/users")
@admin_required
def users():
    with db() as c:
        rows = c.execute("""
            SELECT id, nome, email, role, ativo, departamento, criado_em, ultimo_login
            FROM users
            ORDER BY nome
        """).fetchall()
    return render_template("auth/users.html", users=rows)


@app.route("/users/novo", methods=["GET", "POST"])
@admin_required
def novo_user():
    if request.method == "POST":
        role = request.form.get("role")
        if role not in USER_ROLES:
            flash("Perfil inválido.", "error")
            return redirect(url_for("novo_user"))
        password = request.form.get("password", "")
        if len(password) < 6:
            flash("A password deve ter pelo menos 6 caracteres.", "error")
            return redirect(url_for("novo_user"))
        try:
            with db() as c:
                cur = c.execute("""
                    INSERT INTO users
                    (nome, email, password_hash, role, ativo, departamento, criado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    request.form.get("nome", "").strip(),
                    request.form.get("email", "").strip().lower(),
                    generate_password_hash(password),
                    role,
                    1 if request.form.get("ativo") else 0,
                    request.form.get("departamento"),
                    now(),
                ))
                log_action("criação utilizador", "users", cur.lastrowid, request.form.get("email"), conn=c)
            flash("Utilizador criado.", "success")
            return redirect(url_for("users"))
        except sqlite3.IntegrityError:
            flash("Já existe um utilizador com esse email.", "error")

    return render_template("auth/novo_user.html", roles=USER_ROLES, departamentos=CENTRO_TAREFAS_DEPARTAMENTOS)


@app.route("/centro-tarefas")
def centro_tarefas():
    categoria = request.args.get("categoria", "")
    estado = request.args.get("estado", "")
    prioridade = request.args.get("prioridade", "")
    departamento = request.args.get("departamento", "")
    q = request.args.get("q", "").strip()
    page = max(clean_int(request.args.get("page")) or 1, 1)
    per_page = 25

    where = ["1=1"]
    params = []
    if categoria:
        where.append("categoria = ?")
        params.append(categoria)
    if estado:
        where.append("estado = ?")
        params.append(estado)
    if prioridade:
        where.append("prioridade = ?")
        params.append(prioridade)
    if departamento:
        where.append("departamento = ?")
        params.append(departamento)
    if q:
        where.append("(assunto LIKE ? OR descricao LIKE ? OR matricula LIKE ? OR cliente LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])

    role = session.get("role")
    if role == "Gestor" and session.get("departamento"):
        where.append("(departamento = ? OR criado_por_id = ?)")
        params.extend([session.get("departamento"), session.get("user_id")])
    elif role not in ("Admin", "Gestor"):
        where.append("criado_por_id = ?")
        params.append(session.get("user_id"))

    with db() as c:
        total = c.execute(f"""
            SELECT COUNT(*) c
            FROM centro_tarefas
            WHERE {' AND '.join(where)}
        """, params).fetchone()["c"]
        rows = c.execute(f"""
            SELECT *
            FROM centro_tarefas
            WHERE {' AND '.join(where)}
            ORDER BY
              CASE prioridade WHEN 'Urgente' THEN 1 WHEN 'Alta' THEN 2 WHEN 'Média' THEN 3 ELSE 4 END,
              id DESC
            LIMIT ? OFFSET ?
        """, params + [per_page, (page - 1) * per_page]).fetchall()
        resumo = {
            cat: c.execute("""
                SELECT COUNT(*) c
                FROM centro_tarefas
                WHERE categoria = ?
            """, (cat,)).fetchone()["c"]
            for cat in CENTRO_TAREFAS_CATEGORIAS
        }
    return render_template(
        "centro_tarefas/index.html",
        tarefas=rows,
        resumo=resumo,
        categorias=CENTRO_TAREFAS_CATEGORIAS,
        prioridades=CENTRO_TAREFAS_PRIORIDADES,
        departamentos=CENTRO_TAREFAS_DEPARTAMENTOS,
        estados_por_categoria=CENTRO_TAREFAS_ESTADOS,
        filtros={
            "categoria": categoria,
            "estado": estado,
            "prioridade": prioridade,
            "departamento": departamento,
            "q": q,
        },
        pagination={
            "page": page,
            "per_page": per_page,
            "total": total,
            "pages": max(1, (total + per_page - 1) // per_page),
        },
    )


@app.route("/centro-tarefas/novo", methods=["GET", "POST"])
def centro_tarefas_novo():
    if request.method == "POST":
        categoria = request.form.get("categoria")
        if categoria not in CENTRO_TAREFAS_CATEGORIAS:
            flash("Categoria inválida.", "error")
            return redirect(url_for("centro_tarefas_novo"))
        assunto = request.form.get("assunto", "").strip()
        descricao = request.form.get("descricao", "").strip()
        departamento = request.form.get("departamento", "").strip()
        prioridade = request.form.get("prioridade") or "Média"
        if not assunto or not descricao or not departamento:
            flash("Categoria, assunto, descrição e departamento são obrigatórios.", "error")
            return redirect(url_for("centro_tarefas_novo"))
        estado = "Nova" if categoria == "Sugestões" else "Novo"
        with db() as c:
            cur = c.execute("""
                INSERT INTO centro_tarefas
                (categoria, assunto, descricao, departamento, prioridade, estado,
                 matricula, estacao, cliente, responsavel, data_limite,
                 criado_por_id, criado_por_nome, criado_em, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                categoria,
                assunto,
                descricao,
                departamento,
                prioridade,
                estado,
                request.form.get("matricula", "").strip().upper(),
                request.form.get("estacao"),
                request.form.get("cliente"),
                request.form.get("responsavel"),
                request.form.get("data_limite") or None,
                session.get("user_id"),
                session.get("user_name"),
                now(),
                now(),
            ))
            tarefa_id = cur.lastrowid
            record_task_history(c, tarefa_id, "criação", None, estado)
            record_task_history(c, tarefa_id, "prioridade", None, prioridade)
            record_task_history(c, tarefa_id, "departamento", None, departamento)
            save_centro_tarefa_anexos(c, tarefa_id, request.files.getlist("anexos"))
            log_action("criação centro tarefas", "centro_tarefas", tarefa_id, assunto, conn=c)
        flash("Registo criado.", "success")
        return redirect(url_for("centro_tarefa_detail", tarefa_id=tarefa_id))

    return render_template(
        "centro_tarefas/novo.html",
        categorias=CENTRO_TAREFAS_CATEGORIAS,
        prioridades=CENTRO_TAREFAS_PRIORIDADES,
        departamentos=CENTRO_TAREFAS_DEPARTAMENTOS,
    )


@app.route("/centro-tarefas/<int:tarefa_id>")
def centro_tarefa_detail(tarefa_id):
    with db() as c:
        tarefa = c.execute("SELECT * FROM centro_tarefas WHERE id = ?", (tarefa_id,)).fetchone()
        if not tarefa or not can_view_centro_tarefa(tarefa):
            flash("Registo não encontrado ou sem acesso.", "error")
            return redirect(url_for("centro_tarefas"))
        comentarios = c.execute("""
            SELECT * FROM centro_tarefas_comentarios
            WHERE tarefa_id = ?
            ORDER BY id
        """, (tarefa_id,)).fetchall()
        anexos = c.execute("""
            SELECT * FROM centro_tarefas_anexos
            WHERE tarefa_id = ?
            ORDER BY id DESC
        """, (tarefa_id,)).fetchall()
        historico = c.execute("""
            SELECT th.*, u.nome AS user_nome
            FROM task_history th
            LEFT JOIN users u ON u.id = th.user_id
            WHERE th.task_id = ?
            ORDER BY th.id DESC
        """, (tarefa_id,)).fetchall()
    can_update = session.get("role") in ("Admin", "Gestor") or tarefa["criado_por_id"] == session.get("user_id")
    return render_template(
        "centro_tarefas/detail.html",
        tarefa=tarefa,
        comentarios=comentarios,
        anexos=anexos,
        historico=historico,
        estados=centro_estados_para_categoria(tarefa["categoria"]),
        can_update=can_update,
    )


@app.route("/centro-tarefas/<int:tarefa_id>/atualizar", methods=["POST"])
def centro_tarefa_atualizar(tarefa_id):
    with db() as c:
        tarefa = c.execute("SELECT * FROM centro_tarefas WHERE id = ?", (tarefa_id,)).fetchone()
        if not tarefa or not can_view_centro_tarefa(tarefa):
            flash("Sem acesso ao registo.", "error")
            return redirect(url_for("centro_tarefas"))
        if session.get("role") not in ("Admin", "Gestor") and tarefa["criado_por_id"] != session.get("user_id"):
            flash("Sem permissão para atualizar este registo.", "error")
            return redirect(url_for("centro_tarefa_detail", tarefa_id=tarefa_id))
        estado = request.form.get("estado") or tarefa["estado"]
        fechado_em = now() if estado in {"Concluído", "Fechado", "Resolvido", "Aceite", "Rejeitada"} else None
        c.execute("""
            UPDATE centro_tarefas
            SET estado = ?,
                responsavel = ?,
                data_limite = ?,
                atualizado_em = ?,
                fechado_em = COALESCE(?, fechado_em)
            WHERE id = ?
        """, (
            estado,
            request.form.get("responsavel"),
            request.form.get("data_limite") or None,
            now(),
            fechado_em,
            tarefa_id,
        ))
        record_task_history(c, tarefa_id, "estado", tarefa["estado"], estado)
        record_task_history(c, tarefa_id, "responsavel", tarefa["responsavel"], request.form.get("responsavel"))
        record_task_history(c, tarefa_id, "data_limite", tarefa["data_limite"], request.form.get("data_limite") or None)
        comentario = request.form.get("comentario", "").strip()
        if comentario:
            c.execute("""
                INSERT INTO centro_tarefas_comentarios
                (tarefa_id, user_id, user_nome, comentario, criado_em)
                VALUES (?, ?, ?, ?, ?)
            """, (tarefa_id, session.get("user_id"), session.get("user_name"), comentario, now()))
            record_task_history(c, tarefa_id, "comentário", None, comentario[:240])
        save_centro_tarefa_anexos(c, tarefa_id, request.files.getlist("anexos"))
        log_action("atualização centro tarefas", "centro_tarefas", tarefa_id, estado, conn=c)
    flash("Registo atualizado.", "success")
    return redirect(url_for("centro_tarefa_detail", tarefa_id=tarefa_id))


@app.route("/centro-tarefas/exportar")
@gestor_required
def centro_tarefas_exportar():
    if Workbook is None:
        flash("Falta instalar openpyxl para exportar Excel.", "error")
        return redirect(url_for("centro_tarefas"))
    with db() as c:
        rows = c.execute("""
            SELECT *
            FROM centro_tarefas
            ORDER BY criado_em DESC
        """).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = "Centro de Tarefas"
    headers = ["ID", "Categoria", "Assunto", "Departamento", "Estado", "Prioridade", "Responsável", "Criado por", "Data", "Matrícula", "Estação", "Cliente"]
    ws.append(headers)
    for row in rows:
        ws.append([
            row["id"], row["categoria"], row["assunto"], row["departamento"], row["estado"],
            row["prioridade"], row["responsavel"], row["criado_por_nome"], row["criado_em"],
            row["matricula"], row["estacao"], row["cliente"],
        ])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"centro_tarefas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/conhecimento")
def conhecimento_index():
    area = request.args.get("area", "").strip()
    q = request.args.get("q", "").strip()
    where = ["estado = 'Publicado'"]
    params = []
    if area:
        where.append("area = ?")
        params.append(area)
    if q:
        where.append("(titulo LIKE ? OR tema LIKE ? OR resumo LIKE ? OR conteudo LIKE ?)")
        params.extend([f"%{q}%"] * 4)
    with db() as c:
        artigos = c.execute(f"""
            SELECT *
            FROM conhecimento_artigos
            WHERE {" AND ".join(where)}
            ORDER BY area, tema, titulo
        """, params).fetchall()
        areas = c.execute("""
            SELECT area, COUNT(*) total
            FROM conhecimento_artigos
            WHERE estado = 'Publicado'
            GROUP BY area
            ORDER BY area
        """).fetchall()
    return render_template(
        "conhecimento/index.html",
        artigos=artigos,
        areas=areas,
        filtros={"area": area, "q": q},
    )


@app.route("/conhecimento/novo", methods=["GET", "POST"])
@gestor_required
def conhecimento_novo():
    if request.method == "POST":
        titulo = request.form.get("titulo", "").strip()
        slug = text_slug(request.form.get("slug") or titulo)
        area = request.form.get("area", "").strip()
        conteudo = request.form.get("conteudo", "").strip()
        if not titulo or not area or not conteudo:
            flash("Título, área e conteúdo são obrigatórios.", "error")
            return redirect(url_for("conhecimento_novo"))
        with db() as c:
            try:
                cur = c.execute("""
                    INSERT INTO conhecimento_artigos
                    (titulo, slug, area, tema, resumo, conteudo, estado, visivel_contexto,
                     criado_por, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    titulo,
                    slug,
                    area,
                    request.form.get("tema", "").strip(),
                    request.form.get("resumo", "").strip(),
                    conteudo,
                    request.form.get("estado") or "Publicado",
                    1 if request.form.get("visivel_contexto") else 0,
                    session.get("user_id"),
                    now(),
                    now(),
                ))
                log_action("criação artigo conhecimento", "conhecimento_artigos", cur.lastrowid, titulo, conn=c)
            except sqlite3.IntegrityError:
                flash("Já existe um artigo com esse slug.", "error")
                return redirect(url_for("conhecimento_novo"))
        flash("Artigo criado.", "success")
        return redirect(url_for("conhecimento_detail", slug=slug))
    return render_template("conhecimento/form.html", artigo=None, areas=CONHECIMENTO_AREAS)


@app.route("/conhecimento/<slug>")
def conhecimento_detail(slug):
    with db() as c:
        artigo = get_conhecimento_by_slug(c, slug)
    if not artigo:
        flash("Artigo não encontrado.", "error")
        return redirect(url_for("conhecimento_index"))
    return render_template("conhecimento/detail.html", artigo=artigo)


@app.route("/conhecimento/<int:artigo_id>/editar", methods=["GET", "POST"])
@gestor_required
def conhecimento_editar(artigo_id):
    with db() as c:
        artigo = c.execute("SELECT * FROM conhecimento_artigos WHERE id = ?", (artigo_id,)).fetchone()
        if not artigo:
            flash("Artigo não encontrado.", "error")
            return redirect(url_for("conhecimento_index"))
        if request.method == "POST":
            titulo = request.form.get("titulo", "").strip()
            slug = text_slug(request.form.get("slug") or titulo)
            area = request.form.get("area", "").strip()
            conteudo = request.form.get("conteudo", "").strip()
            if not titulo or not area or not conteudo:
                flash("Título, área e conteúdo são obrigatórios.", "error")
                return redirect(url_for("conhecimento_editar", artigo_id=artigo_id))
            try:
                c.execute("""
                    UPDATE conhecimento_artigos
                    SET titulo = ?,
                        slug = ?,
                        area = ?,
                        tema = ?,
                        resumo = ?,
                        conteudo = ?,
                        estado = ?,
                        visivel_contexto = ?,
                        atualizado_em = ?
                    WHERE id = ?
                """, (
                    titulo,
                    slug,
                    area,
                    request.form.get("tema", "").strip(),
                    request.form.get("resumo", "").strip(),
                    conteudo,
                    request.form.get("estado") or "Publicado",
                    1 if request.form.get("visivel_contexto") else 0,
                    now(),
                    artigo_id,
                ))
                log_action("edição artigo conhecimento", "conhecimento_artigos", artigo_id, titulo, conn=c)
            except sqlite3.IntegrityError:
                flash("Já existe um artigo com esse slug.", "error")
                return redirect(url_for("conhecimento_editar", artigo_id=artigo_id))
            flash("Artigo atualizado.", "success")
            return redirect(url_for("conhecimento_detail", slug=slug))
    return render_template("conhecimento/form.html", artigo=artigo, areas=CONHECIMENTO_AREAS)


@app.route("/stock-home")
def stock_home():
    return render_template("stock/home.html")


@app.route("/manutencao-home")
def manutencao_home():
    return render_template("oficina/home.html")


@app.route("/frota-home")
def frota_home():
    return render_template("frota/home.html")


@app.route("/gestao-home")
def gestao_home():
    return render_template("gestao/home.html")


@app.route("/admin-home")
@admin_required
def admin_home():
    return render_template("admin/home.html")


@app.route("/auditoria-home")
def auditoria_home():
    with db() as c:
        alertas = {
            "contratos_sem_viaturas": c.execute("""
                SELECT COUNT(*) total
                FROM contratos_financeiro cf
                LEFT JOIN contratos_viaturas cv ON cv.contrato_nr = cf.contrato_nr
                WHERE cv.id IS NULL
            """).fetchone()["total"],
            "contratos_multiviatura": c.execute("""
                SELECT COUNT(*) total
                FROM (
                    SELECT contrato_nr
                    FROM contratos_viaturas
                    GROUP BY contrato_nr
                    HAVING COUNT(DISTINCT matricula) > 1
                )
            """).fetchone()["total"],
            "contratos_total_negativo": c.execute("""
                SELECT COUNT(*) total FROM contratos_financeiro WHERE valor_total < 0
            """).fetchone()["total"],
            "contratos_total_zero": c.execute("""
                SELECT COUNT(*) total FROM contratos_financeiro WHERE valor_total = 0
            """).fetchone()["total"],
            "faturas_validar": c.execute("""
                SELECT COUNT(*) total FROM faturas_fornecedores_linhas WHERE estado_importacao = 'VALIDAR'
            """).fetchone()["total"],
            "sinistros_validar": c.execute("""
                SELECT COUNT(*) total FROM sinistros_allianz WHERE estado_importacao = 'VALIDAR'
            """).fetchone()["total"],
            "fo_validar": c.execute("""
                SELECT COUNT(*) total FROM rentway_folhas_obra WHERE estado_importacao <> 'OK'
            """).fetchone()["total"],
            "ars_validar": c.execute("""
                SELECT COUNT(*) total FROM rentway_accident_reports WHERE estado_importacao <> 'OK'
            """).fetchone()["total"],
            "impros_validar": 0,
        }

        prioridades = [
            {
                "nivel": "Critico",
                "titulo": "Contratos financeiros sem movimentos de viatura",
                "valor": alertas["contratos_sem_viaturas"],
                "descricao": "Contrato existe financeiramente, mas nao tem historico operacional de viaturas associado.",
                "url": url_for("contratos_rentway"),
            },
            {
                "nivel": "Validar",
                "titulo": "Contratos com mais do que uma viatura",
                "valor": alertas["contratos_multiviatura"],
                "descricao": "Pode ser substituicao normal, mas deve aparecer na analise financeira do contrato.",
                "url": url_for("contratos_rentway"),
            },
            {
                "nivel": "Financeiro",
                "titulo": "Contratos com total negativo ou zero",
                "valor": (alertas["contratos_total_negativo"] or 0) + (alertas["contratos_total_zero"] or 0),
                "descricao": "Possiveis acertos, anulacoes, contratos sem faturacao ou dados incompletos.",
                "url": url_for("contratos_rentway"),
            },
            {
                "nivel": "Importacao",
                "titulo": "Linhas importadas a validar",
                "valor": (
                    (alertas["faturas_validar"] or 0)
                    + (alertas["sinistros_validar"] or 0)
                    + (alertas["fo_validar"] or 0)
                    + (alertas["ars_validar"] or 0)
                    + (alertas["impros_validar"] or 0)
                ),
                "descricao": "Dados com campos em falta, formatos estranhos ou risco de leitura incorreta.",
                "url": url_for("faturas_fornecedores"),
            },
        ]

        contratos_sem_viaturas = c.execute("""
            SELECT cf.contrato_nr, cf.data_out, cf.data_in, cf.matricula_final, cf.valor_total, cf.origem
            FROM contratos_financeiro cf
            LEFT JOIN contratos_viaturas cv ON cv.contrato_nr = cf.contrato_nr
            WHERE cv.id IS NULL
            ORDER BY cf.data_out DESC, cf.id DESC
            LIMIT 15
        """).fetchall()

        contratos_multiviatura = c.execute("""
            SELECT
                cv.contrato_nr,
                MIN(cv.data_out) data_out,
                MAX(cv.data_in) data_in,
                GROUP_CONCAT(DISTINCT cv.matricula) matriculas,
                COUNT(DISTINCT cv.matricula) viaturas,
                MAX(cv.cliente_nome) cliente
            FROM contratos_viaturas cv
            GROUP BY cv.contrato_nr
            HAVING COUNT(DISTINCT cv.matricula) > 1
            ORDER BY MAX(cv.data_out) DESC
            LIMIT 15
        """).fetchall()

        contratos_valor_anormal = c.execute("""
            SELECT contrato_nr, data_out, data_in, matricula_final, valor_total, origem
            FROM contratos_financeiro
            WHERE valor_total <= 0
            ORDER BY data_out DESC, id DESC
            LIMIT 15
        """).fetchall()

        faturas_validar = c.execute("""
            SELECT
                fornecedor,
                COALESCE(NULLIF(documento, ''), fonte_pdf, 'Sem documento') documento,
                MAX(data_doc) data_doc,
                GROUP_CONCAT(DISTINCT matricula) matriculas,
                COUNT(*) linhas,
                ROUND(COALESCE(SUM(total_liq), 0), 2) valor_total
            FROM faturas_fornecedores_linhas
            WHERE estado_importacao = 'VALIDAR'
            GROUP BY fornecedor, COALESCE(NULLIF(documento, ''), fonte_pdf, 'Sem documento')
            ORDER BY MAX(data_doc) DESC
            LIMIT 15
        """).fetchall()

        sinistros_validar = c.execute("""
            SELECT matricula, referencia_sinistro, data_sinistro, causa, custo_total, alertas
            FROM sinistros_allianz
            WHERE estado_importacao = 'VALIDAR'
            ORDER BY data_sinistro DESC, id DESC
            LIMIT 15
        """).fetchall()

        viaturas_custo = c.execute("""
            SELECT
                fl.matricula,
                MAX(v.marca) marca,
                MAX(v.modelo) modelo,
                COUNT(DISTINCT COALESCE(fl.fornecedor, '') || '|' || COALESCE(NULLIF(fl.documento, ''), fl.fonte_pdf, '')) documentos,
                COUNT(*) linhas,
                ROUND(COALESCE(SUM(fl.total_liq), 0), 2) valor_total
            FROM faturas_fornecedores_linhas fl
            LEFT JOIN viaturas v ON v.matricula = fl.matricula
            WHERE fl.matricula IS NOT NULL AND fl.matricula <> ''
            GROUP BY fl.matricula
            HAVING COALESCE(SUM(fl.total_liq), 0) > 0
            ORDER BY SUM(fl.total_liq) DESC
            LIMIT 15
        """).fetchall()

        viaturas_reincidencia = c.execute("""
            WITH fo AS (
                SELECT matricula, COUNT(DISTINCT folha_obra_origem) fo_count
                FROM rentway_folhas_obra
                WHERE matricula IS NOT NULL AND matricula <> ''
                GROUP BY matricula
            ),
            impro AS (
                SELECT matricula, COUNT(*) impro_count
                FROM rentway_impros
                WHERE matricula IS NOT NULL AND matricula <> ''
                GROUP BY matricula
            ),
            ar AS (
                SELECT matricula, COUNT(*) ar_count
                FROM rentway_accident_reports
                WHERE matricula IS NOT NULL AND matricula <> ''
                GROUP BY matricula
            ),
            sin AS (
                SELECT matricula, COUNT(*) sin_count
                FROM sinistros_allianz
                WHERE matricula IS NOT NULL AND matricula <> ''
                GROUP BY matricula
            )
            SELECT
                v.matricula,
                v.marca,
                v.modelo,
                COALESCE(fo.fo_count, 0) fo_count,
                COALESCE(impro.impro_count, 0) impro_count,
                COALESCE(ar.ar_count, 0) ar_count,
                COALESCE(sin.sin_count, 0) sin_count,
                (
                    COALESCE(fo.fo_count, 0)
                    + COALESCE(impro.impro_count, 0)
                    + COALESCE(ar.ar_count, 0)
                    + COALESCE(sin.sin_count, 0)
                ) total_eventos
            FROM viaturas v
            LEFT JOIN fo ON fo.matricula = v.matricula
            LEFT JOIN impro ON impro.matricula = v.matricula
            LEFT JOIN ar ON ar.matricula = v.matricula
            LEFT JOIN sin ON sin.matricula = v.matricula
            WHERE (
                COALESCE(fo.fo_count, 0)
                + COALESCE(impro.impro_count, 0)
                + COALESCE(ar.ar_count, 0)
                + COALESCE(sin.sin_count, 0)
            ) >= 3
            ORDER BY total_eventos DESC, v.matricula
            LIMIT 20
        """).fetchall()

        top_servicos = c.execute("""
            SELECT
                CASE
                    WHEN UPPER(descricao) LIKE '%PNEU%' THEN 'Pneus'
                    WHEN UPPER(descricao) LIKE '%CALC%' OR UPPER(descricao) LIKE '%PASTILH%' THEN 'Calcos'
                    WHEN UPPER(descricao) LIKE '%DISCO%' THEN 'Discos'
                    WHEN UPPER(descricao) LIKE '%OLEO%' OR UPPER(descricao) LIKE '%ÓLEO%' THEN 'Oleo'
                    ELSE 'Outros'
                END servico,
                COUNT(*) linhas,
                COUNT(DISTINCT matricula) viaturas,
                ROUND(COALESCE(SUM(total_liq), 0), 2) valor_total
            FROM faturas_fornecedores_linhas
            GROUP BY servico
            ORDER BY valor_total DESC
        """).fetchall()

    return render_template(
        "auditoria/home.html",
        alertas=alertas,
        prioridades=prioridades,
        contratos_sem_viaturas=contratos_sem_viaturas,
        contratos_multiviatura=contratos_multiviatura,
        contratos_valor_anormal=contratos_valor_anormal,
        faturas_validar=faturas_validar,
        sinistros_validar=sinistros_validar,
        viaturas_custo=viaturas_custo,
        viaturas_reincidencia=viaturas_reincidencia,
        top_servicos=top_servicos,
    )


@app.route("/auditoria/viatura-timeline")
def auditoria_viatura_timeline():
    matricula = request.args.get("matricula", "").strip().upper().replace(" ", "")
    eventos = []
    periodos_estado = []
    viatura = None

    with db() as c:
        viaturas_sugeridas = c.execute("""
            WITH eventos AS (
                SELECT matricula, COUNT(*) total FROM rentway_folhas_obra WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
                UNION ALL
                SELECT matricula, COUNT(*) total FROM faturas_fornecedores_linhas WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
                UNION ALL
                SELECT matricula, COUNT(*) total FROM rentway_impros WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
                UNION ALL
                SELECT matricula, COUNT(*) total FROM sinistros_allianz WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
                UNION ALL
                SELECT matricula, COUNT(*) total FROM rentway_accident_reports WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
                UNION ALL
                SELECT matricula, COUNT(*) total FROM contratos_viaturas WHERE matricula IS NOT NULL AND matricula <> '' GROUP BY matricula
            )
            SELECT e.matricula, MAX(v.marca) marca, MAX(v.modelo) modelo, SUM(e.total) total_eventos
            FROM eventos e
            LEFT JOIN viaturas v ON v.matricula = e.matricula
            GROUP BY e.matricula
            ORDER BY SUM(e.total) DESC, e.matricula
            LIMIT 20
        """).fetchall()

        if matricula:
            viatura = c.execute("""
                SELECT *
                FROM viaturas
                WHERE matricula = ?
            """, (matricula,)).fetchone()

            for row in c.execute("""
                SELECT *
                FROM contratos_viaturas
                WHERE matricula = ?
                ORDER BY data_out DESC, id DESC
            """, (matricula,)).fetchall():
                if row["data_out"]:
                    periodos_estado.append({
                        "inicio": row["data_out"],
                        "fim": row["data_in"] or row["data_out"],
                        "tipo": "contrato",
                        "titulo": f"Contrato {row['contrato_nr']}",
                        "detalhe": row["cliente_nome"] or "",
                        "url": url_for("contratos_rentway", q=row["contrato_nr"]),
                    })
                eventos.append({
                    "data": row["data_out"],
                    "tipo": "Contrato",
                    "titulo": f"Contrato {row['contrato_nr']}",
                    "descricao": row["cliente_nome"] or "",
                    "detalhe": f"Saida {row['data_out'] or ''} | Entrada {row['data_in'] or 'em aberto'}",
                    "valor": None,
                    "url": url_for("contratos_rentway", q=row["contrato_nr"]),
                })
                if row["data_in"]:
                    eventos.append({
                        "data": row["data_in"],
                        "tipo": "Fim contrato",
                        "titulo": f"Contrato {row['contrato_nr']}",
                        "descricao": row["cliente_nome"] or "",
                        "detalhe": "Entrada/fecho do movimento operacional",
                        "valor": None,
                        "url": url_for("contratos_rentway", q=row["contrato_nr"]),
                    })

            for row in c.execute("""
                SELECT *
                FROM rentway_impros
                WHERE matricula = ?
                ORDER BY COALESCE(data_fecho, data_in, data_abertura, data_out) DESC, id DESC
            """, (matricula,)).fetchall():
                inicio_impro = row["data_out"] or row["data_abertura"] or row["data_in"] or row["data_fecho"]
                fim_impro = row["data_fecho"] or row["data_in"] or inicio_impro
                if inicio_impro:
                    periodos_estado.append({
                        "inicio": inicio_impro,
                        "fim": fim_impro,
                        "tipo": "impro",
                        "titulo": row["numero_impro"] or "Impro Rentway",
                        "detalhe": row["motivo"] or row["descricao"] or row["oficina"] or "",
                        "url": url_for("impros_rentway", matricula=matricula),
                    })
                eventos.append({
                    "data": row["data_fecho"] or row["data_in"] or row["data_abertura"] or row["data_out"],
                    "tipo": "Impro",
                    "titulo": row["numero_impro"] or "Impro Rentway",
                    "descricao": row["motivo"] or row["descricao"] or row["oficina"] or "",
                    "detalhe": f"Oficina {row['oficina'] or ''} | Dias {row['dias_imobilizado'] or ''}",
                    "valor": row["custo"],
                    "url": url_for("impros_rentway", matricula=matricula),
                })

            for row in c.execute("""
                SELECT *
                FROM rentway_folhas_obra
                WHERE matricula = ?
                ORDER BY data_documento DESC, id DESC
            """, (matricula,)).fetchall():
                eventos.append({
                    "data": row["data_documento"],
                    "tipo": "FO Rentway",
                    "titulo": row["folha_obra_origem"] or row["numero_documento"] or "Folha de obra",
                    "descricao": row["descricao"] or row["natureza"] or "",
                    "detalhe": f"{row['fornecedor'] or ''} | {row['natureza'] or ''}",
                    "valor": row["valor_com_iva"],
                    "url": url_for("folhas_obra_rentway", q=row["folha_obra_origem"] or matricula),
                })

            for row in c.execute("""
                SELECT *
                FROM faturas_fornecedores_linhas
                WHERE matricula = ?
                ORDER BY data_doc DESC, id DESC
            """, (matricula,)).fetchall():
                eventos.append({
                    "data": row["data_doc"],
                    "tipo": "Fatura fornecedor",
                    "titulo": row["documento"] or row["fonte_pdf"] or "Fatura",
                    "descricao": row["descricao"] or "",
                    "detalhe": f"{row['fornecedor'] or ''} | OR {row['or_reparacao'] or ''} | {row['ct'] or ''}",
                    "valor": row["total_liq"],
                    "url": url_for("faturas_fornecedores", q=row["documento"] or matricula),
                })

            for row in c.execute("""
                SELECT *
                FROM sinistros_allianz
                WHERE matricula = ?
                ORDER BY data_sinistro DESC, id DESC
            """, (matricula,)).fetchall():
                eventos.append({
                    "data": row["data_sinistro"],
                    "tipo": "Sinistro Allianz",
                    "titulo": row["referencia_sinistro"] or "Sinistro",
                    "descricao": row["causa"] or row["alertas"] or "",
                    "detalhe": f"Encerramento {row['data_encerramento'] or ''}",
                    "valor": row["custo_total"],
                    "url": url_for("sinistros_allianz", matricula=matricula),
                })

            for row in c.execute("""
                SELECT *
                FROM rentway_accident_reports
                WHERE matricula = ?
                ORDER BY COALESCE(accident_date, request_date) DESC, id DESC
            """, (matricula,)).fetchall():
                eventos.append({
                    "data": row["accident_date"] or row["request_date"],
                    "tipo": "AR Rentway",
                    "titulo": row["accident_report_id"] or row["document_no"] or "Accident report",
                    "descricao": row["status"] or row["state"] or "",
                    "detalhe": f"{row['document_type'] or ''} | Contrato {row['document_no'] or ''}",
                    "valor": None,
                    "url": url_for("ars_rentway", matricula=matricula),
                })

            for row in c.execute("""
                SELECT *
                FROM processos
                WHERE matricula = ?
                ORDER BY data_abertura DESC, id DESC
            """, (matricula,)).fetchall():
                eventos.append({
                    "data": row["data_abertura"],
                    "tipo": "Processo interno",
                    "titulo": row["numero_impro"] or f"Processo {row['id']}",
                    "descricao": row["descricao_inicial"] or "",
                    "detalhe": row["estado"] or "",
                    "valor": None,
                    "url": url_for("processo_detail", processo_id=row["id"]),
                })

            if viatura:
                for row in c.execute("""
                    SELECT d.*, p.numero_impro, p.estado AS estado_processo
                    FROM diagnosticos d
                    LEFT JOIN processos p ON p.id = d.processo_id
                    WHERE d.viatura_id = ?
                    ORDER BY d.data_diagnostico DESC, d.id DESC
                """, (viatura["id"],)).fetchall():
                    eventos.append({
                        "data": row["data_diagnostico"],
                        "tipo": "Diagnostico",
                        "titulo": row["tipo"] or row["numero_impro"] or f"Diagnostico {row['id']}",
                        "descricao": row["problema_identificado"] or row["causa_provavel"] or "",
                        "detalhe": f"{row['prioridade'] or ''} | {row['estado_processo'] or ''}",
                        "valor": None,
                        "url": url_for("processo_detail", processo_id=row["processo_id"]),
                    })

    eventos.sort(key=lambda item: item["data"] or "", reverse=True)

    meses_pt = {
        1: "Janeiro",
        2: "Fevereiro",
        3: "Marco",
        4: "Abril",
        5: "Maio",
        6: "Junho",
        7: "Julho",
        8: "Agosto",
        9: "Setembro",
        10: "Outubro",
        11: "Novembro",
        12: "Dezembro",
    }

    timeline_meses = []
    meses_idx = {}
    for evento in eventos:
        data_evento = parse_excel_datetime(evento["data"])
        mes_key = "sem-data"
        mes_label = "Sem data"
        mes_ordem = "0000-00"
        if data_evento:
            data_base = data_evento[:10]
            try:
                dt = datetime.strptime(data_base, "%Y-%m-%d")
                mes_key = dt.strftime("%Y-%m")
                mes_label = f"{meses_pt[dt.month]} {dt.year}"
                mes_ordem = mes_key
            except ValueError:
                mes_key = data_base[:7] if len(data_base) >= 7 else data_base
                mes_label = mes_key
                mes_ordem = mes_key
        if mes_key not in meses_idx:
            meses_idx[mes_key] = {
                "key": mes_key,
                "label": mes_label,
                "ordem": mes_ordem,
                "eventos": [],
                "valor": 0,
                "riscos": 0,
            }
            timeline_meses.append(meses_idx[mes_key])
        meses_idx[mes_key]["eventos"].append(evento)
        meses_idx[mes_key]["valor"] += evento["valor"] or 0
        if evento["tipo"] in ("Impro", "Sinistro Allianz", "AR Rentway"):
            meses_idx[mes_key]["riscos"] += 1

    timeline_meses.sort(key=lambda mes: mes["ordem"], reverse=True)
    for mes in timeline_meses:
        mes["valor"] = round(mes["valor"], 2)

    def date_from_value(value):
        parsed = parse_excel_datetime(value)
        if not parsed:
            return None
        try:
            return datetime.strptime(parsed[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    def month_iter(start_date, end_date):
        current = start_date.replace(day=1)
        last = end_date.replace(day=1)
        while current <= last:
            yield current
            if current.month == 12:
                current = current.replace(year=current.year + 1, month=1)
            else:
                current = current.replace(month=current.month + 1)

    meses_encontrados = {}
    for evento in eventos:
        data_evento = date_from_value(evento["data"])
        if data_evento:
            meses_encontrados[data_evento.strftime("%Y-%m")] = data_evento.replace(day=1)
    for periodo in periodos_estado:
        inicio = date_from_value(periodo["inicio"])
        fim = date_from_value(periodo["fim"]) or inicio
        if inicio and fim:
            if fim < inicio:
                inicio, fim = fim, inicio
            for mes_dt in month_iter(inicio, fim):
                meses_encontrados[mes_dt.strftime("%Y-%m")] = mes_dt

    meses_calendario = sorted(meses_encontrados.values(), reverse=True)[:12]
    calendario_mensal = []
    for mes_dt in meses_calendario:
        dias_mes = calendar.monthrange(mes_dt.year, mes_dt.month)[1]
        dias = []
        for dia in range(1, dias_mes + 1):
            data_dia = mes_dt.replace(day=dia)
            estado = {
                "classe": "livre",
                "label": "Livre",
                "titulo": f"{data_dia.isoformat()} | Livre",
                "url": "",
            }
            for periodo in periodos_estado:
                inicio = date_from_value(periodo["inicio"])
                fim = date_from_value(periodo["fim"]) or inicio
                if not inicio or not fim:
                    continue
                if fim < inicio:
                    inicio, fim = fim, inicio
                if inicio <= data_dia <= fim:
                    if periodo["tipo"] == "impro" or estado["classe"] == "livre":
                        estado = {
                            "classe": periodo["tipo"],
                            "label": "Impro" if periodo["tipo"] == "impro" else "Contrato",
                            "titulo": f"{data_dia.isoformat()} | {periodo['titulo']} | {periodo['detalhe']}",
                            "url": periodo["url"],
                        }

            op_docs = []
            fin_docs = []
            diag_docs = []
            for evento in eventos:
                data_evento = date_from_value(evento["data"])
                if data_evento != data_dia:
                    continue
                if evento["tipo"] in ("Contrato", "Fim contrato"):
                    continue
                doc_cal = {
                    "tipo": evento["tipo"],
                    "sigla": {
                        "Impro": "IMP",
                        "FO Rentway": "FO",
                        "Fatura fornecedor": "FAT",
                        "Sinistro Allianz": "SIN",
                        "AR Rentway": "AR",
                        "Processo interno": "PROC",
                        "Diagnostico": "DIAG",
                    }.get(evento["tipo"], "DOC"),
                    "titulo": f"{evento['tipo']} | {evento['titulo']} | {evento['descricao']} | {evento['detalhe']}",
                    "url": evento["url"],
                }
                if evento["tipo"] == "Fatura fornecedor":
                    fin_docs.append(doc_cal)
                elif evento["tipo"] in ("Processo interno", "Diagnostico"):
                    diag_docs.append(doc_cal)
                else:
                    op_docs.append(doc_cal)
            dias.append({
                "dia": dia,
                "estado": estado,
                "op_docs": op_docs,
                "fin_docs": fin_docs,
                "diag_docs": diag_docs,
            })

        calendario_mensal.append({
            "key": mes_dt.strftime("%Y-%m"),
            "label": f"{meses_pt[mes_dt.month]} {mes_dt.year}",
            "dias": dias,
        })

    resumo_timeline = {
        "eventos": len(eventos),
        "contratos": sum(1 for e in eventos if e["tipo"] in ("Contrato", "Fim contrato")),
        "custos": round(sum((e["valor"] or 0) for e in eventos), 2),
        "riscos": sum(1 for e in eventos if e["tipo"] in ("Impro", "Sinistro Allianz", "AR Rentway")),
    }

    return render_template(
        "auditoria/viatura_timeline.html",
        matricula=matricula,
        viatura=viatura,
        eventos=eventos,
        timeline_meses=timeline_meses,
        calendario_mensal=calendario_mensal,
        resumo_timeline=resumo_timeline,
        viaturas_sugeridas=viaturas_sugeridas,
    )


@app.route("/stock")
def index():
    with db() as c:
        artigos = c.execute("""
            SELECT
                *,
                CASE WHEN stock_atual <= stock_minimo THEN 1 ELSE 0 END AS alerta
            FROM artigos
            ORDER BY alerta DESC, categoria, nome
        """).fetchall()
    return render_template("index.html", artigos=artigos)


@app.route("/artigos/novo", methods=["GET", "POST"])
def novo_artigo():
    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        nome = request.form.get("nome", "").strip()
        if not codigo or not nome:
            flash("Codigo e nome sao obrigatorios.", "error")
            return redirect(url_for("novo_artigo"))

        try:
            with db() as c:
                c.execute("""
                    INSERT INTO artigos
                    (codigo, nome, categoria, unidade, localizacao, fornecedor, preco_unitario, stock_minimo, stock_atual, criado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
                """, (
                    codigo,
                    nome,
                    request.form.get("categoria"),
                    request.form.get("unidade"),
                    request.form.get("localizacao"),
                    request.form.get("fornecedor"),
                    clean_float(request.form.get("preco_unitario")) or 0,
                    clean_int(request.form.get("stock_minimo")) or 0,
                    now(),
                ))
            flash("Artigo criado.", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            flash("Ja existe um artigo com esse codigo.", "error")

    return render_template("novo_artigo.html")


@app.route("/movimentos/novo", methods=["GET", "POST"])
def novo_movimento():
    with db() as c:
        artigos = c.execute("""
            SELECT *
            FROM artigos
            ORDER BY categoria, nome
        """).fetchall()

        if request.method == "POST":
            artigo_id = clean_int(request.form.get("artigo_id"))
            tipo = request.form.get("tipo", "Entrada")
            quantidade = clean_int(request.form.get("quantidade")) or 0
            if not artigo_id or quantidade <= 0:
                flash("Seleciona artigo e quantidade valida.", "error")
                return redirect(url_for("novo_movimento"))

            artigo = c.execute("SELECT * FROM artigos WHERE id = ?", (artigo_id,)).fetchone()
            if not artigo:
                flash("Artigo nao encontrado.", "error")
                return redirect(url_for("novo_movimento"))

            delta = quantidade
            if tipo == "Saída":
                delta = -quantidade
            elif tipo == "Ajuste":
                delta = quantidade - (artigo["stock_atual"] or 0)

            c.execute("""
                INSERT INTO movimentos_stock
                (artigo_id, tipo, quantidade, motivo, matricula, responsavel, observacoes, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                artigo_id,
                tipo,
                quantidade,
                request.form.get("motivo"),
                request.form.get("matricula", "").strip().upper(),
                request.form.get("responsavel"),
                request.form.get("observacoes"),
                now(),
            ))
            c.execute("""
                UPDATE artigos
                SET stock_atual = COALESCE(stock_atual, 0) + ?
                WHERE id = ?
            """, (delta, artigo_id))
            flash("Movimento registado.", "success")
            return redirect(url_for("movimentos"))

    return render_template("novo_movimento.html", artigos=artigos)


@app.route("/movimentos")
def movimentos():
    with db() as c:
        movimentos_rows = c.execute("""
            SELECT m.*, a.codigo, a.nome
            FROM movimentos_stock m
            JOIN artigos a ON a.id = m.artigo_id
            ORDER BY m.id DESC
            LIMIT 500
        """).fetchall()
    return render_template("movimentos.html", movimentos=movimentos_rows)


@app.route("/protocolos")
def protocolos_home():
    with db() as c:
        stats = {
            "fases": c.execute("SELECT COUNT(*) c FROM protocolo_fases WHERE ativo = 1").fetchone()["c"],
            "campos": c.execute("SELECT COUNT(*) c FROM protocolo_campos WHERE ativo = 1").fetchone()["c"],
            "regras": c.execute("SELECT COUNT(*) c FROM protocolo_regras WHERE ativo = 1").fetchone()["c"],
            "processos": c.execute("SELECT COUNT(*) c FROM processos_manutencao").fetchone()["c"],
        }
        regras = c.execute("""
            SELECT r.*,
                   COUNT(rc.id) AS total_campos,
                   SUM(CASE WHEN rc.obrigatorio = 1 THEN 1 ELSE 0 END) AS obrigatorios
            FROM protocolo_regras r
            LEFT JOIN protocolo_regra_campos rc ON rc.regra_id = r.id
            GROUP BY r.id
            ORDER BY r.ativo DESC, r.nome_regra
        """).fetchall()
    return render_template("protocolos/home.html", stats=stats, regras=regras)


@app.route("/protocolos/campos")
def protocolos_campos():
    with db() as c:
        rows = c.execute("""
            SELECT pc.*, pf.nome_fase, pf.codigo_fase, pf.ordem AS fase_ordem
            FROM protocolo_campos pc
            LEFT JOIN protocolo_fases pf ON pf.id = pc.fase_id
            WHERE pc.ativo = 1
            ORDER BY pf.ordem, pc.ordem
        """).fetchall()
    return render_template(
        "protocolos/campos.html",
        fases=agrupar_campos_por_fase(rows),
        total=len(rows),
    )


@app.route("/protocolos/regras")
def protocolos_regras():
    with db() as c:
        regras = c.execute("""
            SELECT r.*,
                   COUNT(rc.id) AS total_campos,
                   SUM(CASE WHEN rc.obrigatorio = 1 THEN 1 ELSE 0 END) AS obrigatorios,
                   SUM(CASE WHEN rc.exige_anexo = 1 THEN 1 ELSE 0 END) AS anexos
            FROM protocolo_regras r
            LEFT JOIN protocolo_regra_campos rc ON rc.regra_id = r.id
            GROUP BY r.id
            ORDER BY r.ativo DESC, r.nome_regra
        """).fetchall()
    return render_template("protocolos/regras.html", regras=regras)


@app.route("/protocolos/regras/<int:regra_id>")
def protocolo_regra_detail(regra_id):
    with db() as c:
        regra = c.execute("SELECT * FROM protocolo_regras WHERE id = ?", (regra_id,)).fetchone()
        if not regra:
            flash("Regra de protocolo não encontrada.", "error")
            return redirect(url_for("protocolos_regras"))
        campos = protocolo_campos_da_regra(c, regra_id)
    return render_template(
        "protocolos/regra_detail.html",
        regra=regra,
        fases=agrupar_campos_por_fase(campos),
        total=len(campos),
    )


@app.route("/incidentes-frota")
def incidentes_frota():
    estado = request.args.get("estado", "")
    categoria = request.args.get("categoria", "")
    classificacao = request.args.get("classificacao", "")
    matricula = request.args.get("matricula", "").strip().upper()
    gravidade = request.args.get("gravidade", "")
    faturavel = request.args.get("faturavel", "")

    where = ["1=1"]
    params = []
    if estado:
        where.append("i.estado = ?")
        params.append(estado)
    if categoria:
        where.append("i.categoria_id = ?")
        params.append(categoria)
    if classificacao:
        where.append("i.classificacao_id = ?")
        params.append(classificacao)
    if matricula:
        where.append("i.matricula LIKE ?")
        params.append(f"%{matricula}%")
    if gravidade:
        where.append("i.gravidade = ?")
        params.append(gravidade)
    if faturavel:
        where.append("i.faturavel = ?")
        params.append(faturavel)

    with db() as c:
        rows = c.execute(f"""
            SELECT i.*, cat.nome AS categoria_nome, cls.nome AS classificacao_nome
            FROM incidentes_frota i
            JOIN incidentes_categorias cat ON cat.id = i.categoria_id
            JOIN incidentes_classificacoes cls ON cls.id = i.classificacao_id
            WHERE {' AND '.join(where)}
            ORDER BY i.criado_em DESC, i.id DESC
            LIMIT 500
        """, params).fetchall()
        categorias = c.execute("""
            SELECT * FROM incidentes_categorias WHERE ativo = 1 ORDER BY nome
        """).fetchall()
        classificacoes = c.execute("""
            SELECT cls.*, cat.nome AS categoria_nome
            FROM incidentes_classificacoes cls
            JOIN incidentes_categorias cat ON cat.id = cls.categoria_id
            WHERE cls.ativo = 1
            ORDER BY cat.nome, cls.nome
        """).fetchall()
        resumo = {
            "abertos": c.execute("""
                SELECT COUNT(*) c FROM incidentes_frota
                WHERE estado NOT IN ('Fechado', 'Cancelado')
            """).fetchone()["c"],
            "criticos": c.execute("""
                SELECT COUNT(*) c FROM incidentes_frota
                WHERE gravidade = 'Crítica' AND estado NOT IN ('Fechado', 'Cancelado')
            """).fetchone()["c"],
            "faturaveis": c.execute("""
                SELECT COUNT(*) c FROM incidentes_frota
                WHERE faturavel IN ('Sim', 'A avaliar') AND estado NOT IN ('Fechado', 'Cancelado')
            """).fetchone()["c"],
            "tarefas": c.execute("""
                SELECT COUNT(*) c FROM tarefas_frota
                WHERE origem_tipo = 'incidente_frota'
                  AND estado IN ('Pendente', 'Em curso', 'Bloqueada')
            """).fetchone()["c"],
        }

    return render_template(
        "oficina/incidentes_frota.html",
        incidentes=rows,
        categorias=categorias,
        classificacoes=classificacoes,
        resumo=resumo,
        filtros={
            "estado": estado,
            "categoria": categoria,
            "classificacao": classificacao,
            "matricula": matricula,
            "gravidade": gravidade,
            "faturavel": faturavel,
        },
        estados=INCIDENTE_ESTADOS,
        gravidades=INCIDENTE_GRAVIDADES,
        faturaveis=INCIDENTE_FATURAVEL,
    )


@app.route("/incidentes-frota/novo", methods=["GET", "POST"])
def novo_incidente_frota():
    with db() as c:
        categorias = c.execute("""
            SELECT * FROM incidentes_categorias WHERE ativo = 1 ORDER BY nome
        """).fetchall()
        classificacoes = c.execute("""
            SELECT cls.*, cat.nome AS categoria_nome
            FROM incidentes_classificacoes cls
            JOIN incidentes_categorias cat ON cat.id = cls.categoria_id
            WHERE cls.ativo = 1
            ORDER BY cat.nome, cls.nome
        """).fetchall()

        if request.method == "POST":
            matricula = request.form.get("matricula", "").strip().upper().replace(" ", "")
            viatura_id = request.form.get("viatura_id")
            if viatura_id:
                v = c.execute("SELECT * FROM viaturas WHERE id = ?", (viatura_id,)).fetchone()
            else:
                v = c.execute("SELECT * FROM viaturas WHERE matricula = ?", (matricula,)).fetchone()
            if not v:
                flash("Viatura não encontrada. Neste módulo não são criadas viaturas automaticamente.", "error")
                return redirect(url_for("novo_incidente_frota"))

            classificacao_id = request.form.get("classificacao_id")
            classificacao = c.execute("""
                SELECT cls.*, cat.nome AS categoria_nome
                FROM incidentes_classificacoes cls
                JOIN incidentes_categorias cat ON cat.id = cls.categoria_id
                WHERE cls.id = ? AND cls.ativo = 1
            """, (classificacao_id,)).fetchone()
            if not classificacao:
                flash("Classificação inválida.", "error")
                return redirect(url_for("novo_incidente_frota"))

            descricao = request.form.get("descricao", "").strip()
            if not descricao:
                flash("A descrição do incidente é obrigatória.", "error")
                return redirect(url_for("novo_incidente_frota"))

            cur = c.execute("""
                INSERT INTO incidentes_frota
                (viatura_id, matricula, categoria_id, classificacao_id, processo_id,
                 contrato_id, cliente, oficina_marca, descricao, origem, gravidade,
                 estado, faturavel, valor_estimado, responsavel_followup,
                 acao_necessaria, data_limite, criado_por, criado_em, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Aberto', ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                v["id"],
                v["matricula"],
                classificacao["categoria_id"],
                classificacao["id"],
                request.form.get("processo_id") or None,
                request.form.get("contrato_id"),
                request.form.get("cliente"),
                request.form.get("oficina_marca"),
                descricao,
                request.form.get("origem"),
                request.form.get("gravidade") or classificacao["prioridade_default"] or "Média",
                request.form.get("faturavel") or classificacao["faturavel_default"] or "A avaliar",
                clean_float(request.form.get("valor_estimado")),
                request.form.get("responsavel_followup"),
                request.form.get("acao_necessaria"),
                request.form.get("data_limite") or None,
                request.form.get("criado_por"),
                now(),
                now(),
            ))
            incidente_id = cur.lastrowid
            save_incidente_anexos(
                c,
                incidente_id,
                request.files.getlist("anexos"),
                request.form.get("criado_por"),
            )
            incidente = c.execute(
                "SELECT * FROM incidentes_frota WHERE id = ?",
                (incidente_id,)
            ).fetchone()
            tarefa_id = criar_tarefa_frota_para_incidente(c, incidente, classificacao)
            log_action("criação incidente", "incidentes_frota", incidente_id, f"{v['matricula']} · {classificacao['nome']}", conn=c)
            if tarefa_id:
                log_action("criação tarefa", "tarefas_frota", tarefa_id, f"Follow-up incidente {incidente_id}", conn=c)
            if tarefa_id:
                flash("Incidente criado e tarefa de follow-up gerada.", "success")
            else:
                flash("Incidente criado.", "success")
            return redirect(url_for("incidente_frota_detail", incidente_id=incidente_id))

    return render_template(
        "oficina/novo_incidente_frota.html",
        categorias=categorias,
        classificacoes=classificacoes,
        gravidades=INCIDENTE_GRAVIDADES,
        faturaveis=INCIDENTE_FATURAVEL,
    )


@app.route("/incidentes-frota/<int:incidente_id>")
def incidente_frota_detail(incidente_id):
    with db() as c:
        incidente = c.execute("""
            SELECT i.*, cat.nome AS categoria_nome, cls.nome AS classificacao_nome,
                   cls.instrucoes, v.vin, v.marca, v.modelo, v.motorizacao
            FROM incidentes_frota i
            JOIN incidentes_categorias cat ON cat.id = i.categoria_id
            JOIN incidentes_classificacoes cls ON cls.id = i.classificacao_id
            JOIN viaturas v ON v.id = i.viatura_id
            WHERE i.id = ?
        """, (incidente_id,)).fetchone()
        if not incidente:
            flash("Incidente não encontrado.", "error")
            return redirect(url_for("incidentes_frota"))
        anexos = c.execute("""
            SELECT * FROM incidentes_anexos
            WHERE incidente_id = ?
            ORDER BY id DESC
        """, (incidente_id,)).fetchall()
        tarefas = c.execute("""
            SELECT * FROM tarefas_frota
            WHERE origem_tipo = 'incidente_frota' AND origem_id = ?
            ORDER BY id DESC
        """, (str(incidente_id),)).fetchall()

    return render_template(
        "oficina/incidente_frota_detail.html",
        incidente=incidente,
        anexos=anexos,
        tarefas=tarefas,
        estados=INCIDENTE_ESTADOS,
        tarefa_estados=TAREFA_FROTA_ESTADOS,
    )


@app.route("/incidentes-frota/<int:incidente_id>/estado", methods=["POST"])
def atualizar_incidente_frota_estado(incidente_id):
    estado = request.form.get("estado") or "Aberto"
    fechado_em = now() if estado in {"Fechado", "Cancelado"} else None
    with db() as c:
        c.execute("""
            UPDATE incidentes_frota
            SET estado = ?,
                responsavel_followup = ?,
                acao_necessaria = ?,
                atualizado_em = ?,
                fechado_em = COALESCE(?, fechado_em)
            WHERE id = ?
        """, (
            estado,
            request.form.get("responsavel_followup"),
            request.form.get("acao_necessaria"),
            now(),
            fechado_em,
            incidente_id,
        ))
        if estado in {"Fechado", "Cancelado"}:
            log_action("fecho incidente", "incidentes_frota", incidente_id, estado, conn=c)
    flash("Incidente atualizado.", "success")
    return redirect(url_for("incidente_frota_detail", incidente_id=incidente_id))


@app.route("/incidentes-classificacoes")
def incidentes_classificacoes():
    with db() as c:
        rows = c.execute("""
            SELECT cls.*, cat.nome AS categoria_nome
            FROM incidentes_classificacoes cls
            JOIN incidentes_categorias cat ON cat.id = cls.categoria_id
            ORDER BY cat.nome, cls.nome
        """).fetchall()
    return render_template("oficina/incidentes_classificacoes.html", classificacoes=rows)


@app.route("/incidentes-classificacoes/nova", methods=["GET", "POST"])
def nova_classificacao_incidente():
    with db() as c:
        categorias = c.execute("""
            SELECT * FROM incidentes_categorias WHERE ativo = 1 ORDER BY nome
        """).fetchall()
        if request.method == "POST":
            try:
                c.execute("""
                    INSERT INTO incidentes_classificacoes
                    (categoria_id, nome, descricao, ativo, prioridade_default,
                     gera_tarefa, faturavel_default, sla_horas_default,
                     equipa_responsavel_default, exige_anexo, instrucoes, criado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    request.form.get("categoria_id"),
                    request.form.get("nome", "").strip(),
                    request.form.get("descricao"),
                    1 if request.form.get("ativo") else 0,
                    request.form.get("prioridade_default") or "Média",
                    1 if request.form.get("gera_tarefa") else 0,
                    request.form.get("faturavel_default") or "A avaliar",
                    request.form.get("sla_horas_default") or 24,
                    request.form.get("equipa_responsavel_default"),
                    1 if request.form.get("exige_anexo") else 0,
                    request.form.get("instrucoes"),
                    now(),
                ))
                flash("Classificação criada.", "success")
                return redirect(url_for("incidentes_classificacoes"))
            except sqlite3.IntegrityError:
                flash("Já existe uma classificação com esse nome nessa categoria.", "error")

    return render_template(
        "oficina/nova_classificacao_incidente.html",
        categorias=categorias,
        gravidades=INCIDENTE_GRAVIDADES,
        faturaveis=INCIDENTE_FATURAVEL,
    )


@app.route("/tarefas-frota")
def tarefas_frota():
    estado = request.args.get("estado", "")
    responsavel = request.args.get("responsavel", "")
    prioridade = request.args.get("prioridade", "")
    matricula = request.args.get("matricula", "").strip().upper()
    where = ["1=1"]
    params = []
    if estado:
        where.append("tf.estado = ?")
        params.append(estado)
    if responsavel:
        where.append("tf.responsavel LIKE ?")
        params.append(f"%{responsavel}%")
    if prioridade:
        where.append("tf.prioridade = ?")
        params.append(prioridade)
    if matricula:
        where.append("tf.matricula LIKE ?")
        params.append(f"%{matricula}%")

    with db() as c:
        rows = c.execute(f"""
            SELECT tf.*, i.estado AS incidente_estado
            FROM tarefas_frota tf
            LEFT JOIN incidentes_frota i
              ON tf.origem_tipo = 'incidente_frota' AND tf.origem_id = CAST(i.id AS TEXT)
            WHERE {' AND '.join(where)}
            ORDER BY
              CASE tf.estado WHEN 'Pendente' THEN 1 WHEN 'Em curso' THEN 2 WHEN 'Bloqueada' THEN 3 ELSE 4 END,
              tf.data_limite ASC,
              tf.id DESC
            LIMIT 500
        """, params).fetchall()

    return render_template(
        "oficina/tarefas_frota.html",
        tarefas=rows,
        estados=TAREFA_FROTA_ESTADOS,
        prioridades=INCIDENTE_GRAVIDADES,
        filtros={
            "estado": estado,
            "responsavel": responsavel,
            "prioridade": prioridade,
            "matricula": matricula,
        },
    )


@app.route("/tarefas-frota/<int:tarefa_id>/estado", methods=["POST"])
def atualizar_tarefa_frota_estado(tarefa_id):
    estado = request.form.get("estado") or "Pendente"
    concluida_em = now() if estado in {"Concluída", "Validada"} else None
    validada_por = request.form.get("validada_por")
    validada_em = now() if estado == "Validada" else None
    with db() as c:
        c.execute("""
            UPDATE tarefas_frota
            SET estado = ?,
                responsavel = ?,
                concluida_em = COALESCE(?, concluida_em),
                validada_por = COALESCE(?, validada_por),
                validada_em = COALESCE(?, validada_em)
            WHERE id = ?
        """, (
            estado,
            request.form.get("responsavel"),
            concluida_em,
            validada_por,
            validada_em,
            tarefa_id,
        ))
        if estado == "Validada":
            log_action("validação tarefa", "tarefas_frota", tarefa_id, request.form.get("validada_por"), conn=c)
    flash("Tarefa atualizada.", "success")
    return redirect(request.referrer or url_for("tarefas_frota"))


@app.route("/venda-usados")
def venda_usados():
    filtros = {
        "marca": request.args.get("marca", "").strip(),
        "modelo": request.args.get("modelo", "").strip(),
        "versao": request.args.get("versao", "").strip(),
        "ano": request.args.get("ano", "").strip(),
        "km_min": request.args.get("km_min", "").strip(),
        "km_max": request.args.get("km_max", "").strip(),
        "estado_frota": request.args.get("estado_frota", "").strip(),
        "estado_operacional": request.args.get("estado_operacional", "").strip(),
        "grupo_categoria": request.args.get("grupo_categoria", "").strip(),
        "proxima_revisao": request.args.get("proxima_revisao", "").strip(),
        "fornecedor_financeiro": request.args.get("fornecedor_financeiro", "").strip(),
    }
    where = ["COALESCE(v.ativo, 1) = 1", "COALESCE(v.estado_frota, 'Ativa') NOT IN ('Vendida', 'Abatida', 'Baixada')"]
    params = []
    if filtros["marca"]:
        where.append("v.marca LIKE ?")
        params.append(f"%{filtros['marca']}%")
    if filtros["modelo"]:
        where.append("v.modelo LIKE ?")
        params.append(f"%{filtros['modelo']}%")
    if filtros["versao"]:
        where.append("v.versao LIKE ?")
        params.append(f"%{filtros['versao']}%")
    if filtros["ano"]:
        where.append("v.ano = ?")
        params.append(filtros["ano"])
    if filtros["km_min"]:
        where.append("COALESCE(v.km_atual, 0) >= ?")
        params.append(clean_float(filtros["km_min"]) or 0)
    if filtros["km_max"]:
        where.append("COALESCE(v.km_atual, 0) <= ?")
        params.append(clean_float(filtros["km_max"]) or 0)
    if filtros["estado_frota"]:
        where.append("v.estado_frota = ?")
        params.append(filtros["estado_frota"])
    if filtros["estado_operacional"]:
        where.append("v.estado_operacional LIKE ?")
        params.append(f"%{filtros['estado_operacional']}%")
    if filtros["grupo_categoria"]:
        where.append("(v.grupo LIKE ? OR v.categoria LIKE ?)")
        params.extend([f"%{filtros['grupo_categoria']}%", f"%{filtros['grupo_categoria']}%"])
    if filtros["proxima_revisao"]:
        where.append("v.proxima_revisao_data <= ?")
        params.append(filtros["proxima_revisao"])
    if filtros["fornecedor_financeiro"]:
        where.append("v.fornecedor_financeiro LIKE ?")
        params.append(f"%{filtros['fornecedor_financeiro']}%")

    with db() as c:
        rows = c.execute(f"""
            SELECT v.*,
                   vv.id AS venda_id,
                   vv.estado_venda AS venda_estado
            FROM viaturas v
            LEFT JOIN viaturas_venda vv
              ON vv.viatura_id = v.id
             AND vv.selecionada_venda = 1
             AND vv.estado_venda NOT IN ('Vendida', 'Cancelada')
            WHERE {' AND '.join(where)}
            ORDER BY v.marca, v.modelo, v.matricula
            LIMIT 500
        """, params).fetchall()

    viaturas = []
    for row in rows:
        viaturas.append({
            "row": row,
            "venda": dados_venda_da_viatura(row),
        })

    return render_template(
        "venda/venda_usados.html",
        viaturas=viaturas,
        filtros=filtros,
    )


@app.route("/venda-usados/adicionar", methods=["POST"])
def venda_usados_adicionar():
    ids = request.form.getlist("viatura_id")
    if not ids:
        flash("Seleciona pelo menos uma viatura.", "error")
        return redirect(url_for("venda_usados"))

    criadas = 0
    ignoradas = 0
    with db() as c:
        for viatura_id in ids:
            v = c.execute("SELECT * FROM viaturas WHERE id = ?", (viatura_id,)).fetchone()
            if not v:
                ignoradas += 1
                continue
            venda = dados_venda_da_viatura(v)
            try:
                c.execute("""
                    INSERT INTO viaturas_venda
                    (viatura_id, matricula, selecionada_venda, estado_venda,
                     fornecedor_financeiro, proxima_revisao, valor_aquisicao,
                     data_aquisicao, meses_depreciacao, depreciacao_mensal,
                     depreciacao_acumulada, valor_atual_teorico, exportada,
                     criado_em, atualizado_em)
                    VALUES (?, ?, 1, 'Selecionada', ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                """, (
                    v["id"],
                    v["matricula"],
                    venda["fornecedor_financeiro"],
                    venda["proxima_revisao"],
                    venda["valor_aquisicao"],
                    venda["data_aquisicao"],
                    venda["meses_depreciacao"],
                    venda["depreciacao_mensal"],
                    venda["depreciacao_acumulada"],
                    venda["valor_atual_teorico"],
                    now(),
                    now(),
                ))
                criadas += 1
            except sqlite3.IntegrityError:
                ignoradas += 1

    flash(f"{criadas} viatura(s) adicionada(s). {ignoradas} já estavam selecionadas ou foram ignoradas.", "success")
    return redirect(url_for("venda_usados_selecionadas"))


@app.route("/venda-usados/selecionadas")
def venda_usados_selecionadas():
    with db() as c:
        rows = c.execute("""
            SELECT vv.*, v.marca, v.modelo, v.versao, v.ano, v.km_atual, v.vin, v.combustivel
            FROM viaturas_venda vv
            JOIN viaturas v ON v.id = vv.viatura_id
            WHERE vv.selecionada_venda = 1
            ORDER BY
              CASE vv.estado_venda
                WHEN 'Selecionada' THEN 1
                WHEN 'Em análise' THEN 2
                WHEN 'Aprovada' THEN 3
                WHEN 'Exportada' THEN 4
                ELSE 5
              END,
              vv.id DESC
        """).fetchall()
    return render_template(
        "venda/venda_usados_selecionadas.html",
        viaturas=rows,
        estados=VENDA_USADOS_ESTADOS,
    )


@app.route("/venda-usados/<int:venda_id>/editar", methods=["POST"])
def venda_usados_editar(venda_id):
    with db() as c:
        row = c.execute("""
            SELECT vv.*, v.valor_compra, v.valor_compra_com_iva, v.data_compra
            FROM viaturas_venda vv
            JOIN viaturas v ON v.id = vv.viatura_id
            WHERE vv.id = ?
        """, (venda_id,)).fetchone()
        if not row:
            flash("Registo de venda não encontrado.", "error")
            return redirect(url_for("venda_usados_selecionadas"))
        valor_aquisicao = request.form.get("valor_aquisicao") or row["valor_aquisicao"] or row["valor_compra_com_iva"] or row["valor_compra"]
        data_aquisicao = request.form.get("data_aquisicao") or row["data_aquisicao"] or row["data_compra"]
        dep = calcular_depreciacao_venda(valor_aquisicao, data_aquisicao)
        c.execute("""
            UPDATE viaturas_venda
            SET preco_sugerido = ?,
                preco_minimo = ?,
                preco_comercio = ?,
                observacoes = ?,
                estado_venda = ?,
                valor_aquisicao = ?,
                data_aquisicao = ?,
                meses_depreciacao = ?,
                depreciacao_mensal = ?,
                depreciacao_acumulada = ?,
                valor_atual_teorico = ?,
                atualizado_em = ?
            WHERE id = ?
        """, (
            clean_float(request.form.get("preco_sugerido")),
            clean_float(request.form.get("preco_minimo")),
            clean_float(request.form.get("preco_comercio")),
            request.form.get("observacoes"),
            request.form.get("estado_venda") or "Selecionada",
            dep["valor_aquisicao"],
            dep["data_aquisicao"],
            dep["meses_depreciacao"],
            dep["depreciacao_mensal"],
            dep["depreciacao_acumulada"],
            dep["valor_atual_teorico"],
            now(),
            venda_id,
        ))
    flash("Registo atualizado.", "success")
    return redirect(url_for("venda_usados_selecionadas"))


@app.route("/venda-usados/<int:venda_id>/remover", methods=["POST"])
def venda_usados_remover(venda_id):
    with db() as c:
        c.execute("""
            UPDATE viaturas_venda
            SET selecionada_venda = 0,
                estado_venda = 'Cancelada',
                atualizado_em = ?
            WHERE id = ?
        """, (now(), venda_id))
    flash("Viatura removida da seleção.", "success")
    return redirect(url_for("venda_usados_selecionadas"))


@app.route("/venda-usados/<int:venda_id>/exportada", methods=["POST"])
def venda_usados_marcar_exportada(venda_id):
    with db() as c:
        c.execute("""
            UPDATE viaturas_venda
            SET exportada = 1,
                exportada_em = ?,
                estado_venda = 'Exportada',
                atualizado_em = ?
            WHERE id = ?
        """, (now(), now(), venda_id))
    flash("Viatura marcada como exportada.", "success")
    return redirect(url_for("venda_usados_selecionadas"))


@app.route("/venda-usados/exportar")
def venda_usados_exportar():
    if session.get("role") not in ("Admin", "Gestor"):
        flash("Exportação reservada a Admin ou Gestor.", "error")
        return redirect(url_for("venda_usados_selecionadas"))
    if Workbook is None:
        flash("Falta instalar openpyxl para exportar Excel.", "error")
        return redirect(url_for("venda_usados_selecionadas"))

    with db() as c:
        rows = c.execute("""
            SELECT vv.*, v.vin, v.marca, v.modelo, v.versao, v.ano, v.km_atual,
                   v.combustivel
            FROM viaturas_venda vv
            JOIN viaturas v ON v.id = vv.viatura_id
            WHERE vv.selecionada_venda = 1
              AND vv.estado_venda IN ('Aprovada', 'Exportada')
            ORDER BY v.marca, v.modelo, vv.matricula
        """).fetchall()
        if not rows:
            flash("Não existem viaturas aprovadas/exportáveis.", "error")
            return redirect(url_for("venda_usados_selecionadas"))

        wb = Workbook()
        ws = wb.active
        ws.title = "Venda Usados"
        headers = [
            "Matrícula", "VIN", "Marca", "Modelo", "Versão", "Ano", "KM",
            "Combustível", "Fornecedor financeiro", "Próxima revisão",
            "Valor aquisição", "Valor atual teórico", "Preço sugerido",
            "Preço mínimo", "Preço comércio", "Observações",
        ]
        ws.append(headers)
        for row in rows:
            ws.append([
                row["matricula"],
                row["vin"],
                row["marca"],
                row["modelo"],
                row["versao"],
                row["ano"],
                row["km_atual"],
                row["combustivel"],
                row["fornecedor_financeiro"],
                row["proxima_revisao"],
                row["valor_aquisicao"],
                row["valor_atual_teorico"],
                row["preco_sugerido"],
                row["preco_minimo"],
                row["preco_comercio"],
                row["observacoes"],
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 38)

        export_ids = [row["id"] for row in rows]
        c.executemany("""
            UPDATE viaturas_venda
            SET exportada = 1,
                exportada_em = ?,
                estado_venda = 'Exportada',
                atualizado_em = ?
            WHERE id = ?
        """, [(now(), now(), vid) for vid in export_ids])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"venda_usados_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/processos")
def processos():
    with db() as c:
        rows = c.execute("""
            SELECT p.*, v.vin, v.marca, v.modelo
            FROM processos p
            JOIN viaturas v ON v.id = p.viatura_id
            WHERE COALESCE(p.arquivo, 0) = 0
              AND COALESCE(p.visivel_gestao, 1) = 1
            ORDER BY p.id DESC
        """).fetchall()

    return render_template("oficina/processos.html", processos=rows)


@app.route("/processos/novo", methods=["GET", "POST"])
def novo_processo():
    if request.method == "POST":
        impro = request.form["numero_impro"].strip()
        mat = request.form["matricula"].strip().upper()
        km = int(request.form.get("km_abertura") or 0) or None

        vid = ensure_viatura(mat, km)

        try:
            with db() as c:
                cur = c.execute("""
                    INSERT INTO processos
                    (numero_impro, viatura_id, matricula, km_abertura, responsavel, descricao_inicial, estado, data_abertura)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    impro,
                    vid,
                    mat,
                    km,
                    request.form.get("responsavel", ""),
                    request.form.get("descricao_inicial", ""),
                    "Receção / ServiceBox",
                    now(),
                ))
                log_action("criação processo", "processos", cur.lastrowid, f"Impro {impro} · {mat}", conn=c)

            flash("Processo aberto. Avança para Receção / ServiceBox.", "success")
            return redirect(url_for("processo_detail", processo_id=cur.lastrowid))

        except sqlite3.IntegrityError:
            flash("Já existe processo com esse Nº Impro.", "error")

    return render_template("oficina/novo_processo.html")


@app.route("/diagnosticos-pdf/importar", methods=["GET", "POST"])
def importar_diagnosticos_pdf_view():
    if request.method == "POST":
        responsavel = request.form.get("responsavel")
        pasta = request.form.get("pasta", "").strip()
        ficheiros = request.files.getlist("ficheiros")
        try:
            if pasta:
                folder = Path(pasta)
                if not folder.exists() or not folder.is_dir():
                    raise RuntimeError("A pasta indicada não existe ou não é uma pasta.")
                pdfs = sorted(folder.glob("*.pdf"))
                stats = import_diagnosticos_pdf_paths(pdfs, pasta, responsavel=responsavel, copy_files=True)
            else:
                saved = []
                for f in ficheiros:
                    if not f or not f.filename:
                        continue
                    safe = secure_filename(f.filename)
                    stored_name = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe}"
                    path = IMPORT_DIR / stored_name
                    f.save(path)
                    saved.append(path)
                stats = import_diagnosticos_pdf_paths(saved, "upload_pdf", responsavel=responsavel, copy_files=False)
        except Exception as exc:
            flash(f"Não foi possível importar diagnósticos PDF: {exc}", "error")
            return redirect(url_for("importar_diagnosticos_pdf_view"))

        flash(
            "Importação de diagnósticos PDF concluída: "
            f"{stats['criadas']} criados, {stats['atualizadas']} atualizados, "
            f"{stats['ignoradas']} ignorados, {stats['alertas']} a validar.",
            "success"
        )
        return redirect(url_for("importar_diagnosticos_pdf_view"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Diagnosticos PDF'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_diagnosticos_pdf.html", importacoes=importacoes)


@app.route("/processos/<int:processo_id>")
def processo_detail(processo_id):
    p = processo(processo_id)

    with db() as c:
        rec = c.execute(
            "SELECT * FROM rececoes WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

        diags = c.execute(
            "SELECT * FROM diagnosticos WHERE processo_id = ? ORDER BY id DESC",
            (processo_id,)
        ).fetchall()

        fo = c.execute(
            "SELECT * FROM folhas_obra WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

        tarefas = c.execute(
            "SELECT * FROM tarefas_fo WHERE fo_id = ? ORDER BY id",
            (fo["id"],)
        ).fetchall() if fo else []

        val = c.execute(
            "SELECT * FROM validacoes WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

        anexos = c.execute(
            "SELECT * FROM anexos WHERE processo_id = ? ORDER BY id DESC",
            (processo_id,)
        ).fetchall()

    return render_template(
        "oficina/processo_detail.html",
        p=p,
        rececao=rec,
        diagnosticos=diags,
        fo=fo,
        tarefas=tarefas,
        validacao=val,
        anexos=anexos,
        estados=ESTADOS
    )


@app.route("/processos/<int:processo_id>/checklist", methods=["GET", "POST"])
def processo_checklist(processo_id):
    p = processo(processo_id)
    if not p:
        flash("Processo não encontrado.", "error")
        return redirect(url_for("processos"))

    with db() as c:
        regra = protocolo_regra_aplicavel(c, p)
        regra_id = regra["id"] if regra else None
        ensure_motor_processo(c, p, regra_id)
        campos = protocolo_campos_da_regra(c, regra_id)

        if request.method == "POST":
            responsavel = request.form.get("preenchido_por") or p["responsavel"] or ""
            for campo in campos:
                field_name = f"campo_{campo['id']}"
                raw_value = request.form.get(field_name)
                presente = field_name in request.form
                valor_texto, valor_numero, valor_data, valor_booleano = valor_para_campo(campo, raw_value, presente)
                c.execute("""
                    INSERT INTO processo_campos_valores
                    (processo_id, campo_id, valor_texto, valor_numero, valor_data,
                     valor_booleano, observacoes, preenchido_por, preenchido_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(processo_id, campo_id) DO UPDATE SET
                        valor_texto = excluded.valor_texto,
                        valor_numero = excluded.valor_numero,
                        valor_data = excluded.valor_data,
                        valor_booleano = excluded.valor_booleano,
                        observacoes = excluded.observacoes,
                        preenchido_por = excluded.preenchido_por,
                        preenchido_em = excluded.preenchido_em
                """, (
                    processo_id,
                    campo["id"],
                    valor_texto,
                    valor_numero,
                    valor_data,
                    valor_booleano,
                    request.form.get(f"obs_{campo['id']}"),
                    responsavel,
                    now(),
                ))

                for f in request.files.getlist(f"anexo_{campo['id']}"):
                    if not f or not f.filename:
                        continue
                    safe = secure_filename(f.filename)
                    stored_name = f"{processo_id}_{campo['id']}_{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{safe}"
                    path = UPLOAD_DIR / stored_name
                    f.save(path)
                    c.execute("""
                        INSERT INTO processo_campos_anexos
                        (processo_id, campo_id, fase_id, tipo_anexo, filename,
                         original_name, carregado_por, carregado_em)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        processo_id,
                        campo["id"],
                        campo["fase_id"],
                        campo["tipo_campo"],
                        stored_name,
                        f.filename,
                        responsavel,
                        now(),
                    ))

            flash("Checklist dinâmica guardada.", "success")
            return redirect(url_for("processo_checklist", processo_id=processo_id))

        valores = {
            row["campo_id"]: row
            for row in c.execute(
                "SELECT * FROM processo_campos_valores WHERE processo_id = ?",
                (processo_id,)
            ).fetchall()
        }
        anexos_por_campo = {
            row["campo_id"]: row["total"]
            for row in c.execute("""
                SELECT campo_id, COUNT(*) AS total
                FROM processo_campos_anexos
                WHERE processo_id = ?
                GROUP BY campo_id
            """, (processo_id,)).fetchall()
        }

    total_obrigatorios = sum(1 for campo in campos if campo["obrigatorio_regra"])
    preenchidos = 0
    em_falta = []
    for campo in campos:
        valor = valores.get(campo["id"])
        tem_valor = bool(valor_visivel(campo, valor))
        tem_anexo = anexos_por_campo.get(campo["id"], 0) > 0
        if campo["obrigatorio_regra"] and (tem_valor or tem_anexo):
            preenchidos += 1
        if campo["obrigatorio_regra"] and not tem_valor and not tem_anexo:
            em_falta.append(campo)

    return render_template(
        "protocolos/checklist.html",
        p=p,
        regra=regra,
        fases=agrupar_campos_por_fase(campos),
        valores=valores,
        anexos_por_campo=anexos_por_campo,
        total_obrigatorios=total_obrigatorios,
        preenchidos=preenchidos,
        em_falta=em_falta,
        valor_visivel=valor_visivel,
    )


@app.route("/processos/<int:processo_id>/estado", methods=["POST"])
def atualizar_estado(processo_id):
    with db() as c:
        c.execute(
            "UPDATE processos SET estado = ? WHERE id = ?",
            (request.form["estado"], processo_id)
        )

    return redirect(url_for("processo_detail", processo_id=processo_id))


@app.route("/processos/<int:processo_id>/rececao", methods=["GET", "POST"])
def rececao(processo_id):
    p = processo(processo_id)

    if request.method == "POST":
        with db() as c:
            c.execute("""
                INSERT INTO rececoes
                (processo_id, servicebox_verificado, campanhas_verificadas, campanhas_pendentes,
                 plano_manutencao_verificado, historico_oem_verificado, observacoes, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(processo_id) DO UPDATE SET
                    servicebox_verificado = excluded.servicebox_verificado,
                    campanhas_verificadas = excluded.campanhas_verificadas,
                    campanhas_pendentes = excluded.campanhas_pendentes,
                    plano_manutencao_verificado = excluded.plano_manutencao_verificado,
                    historico_oem_verificado = excluded.historico_oem_verificado,
                    observacoes = excluded.observacoes,
                    atualizado_em = excluded.atualizado_em
            """, (
                processo_id,
                1 if request.form.get("servicebox_verificado") else 0,
                1 if request.form.get("campanhas_verificadas") else 0,
                1 if request.form.get("campanhas_pendentes") else 0,
                request.form.get("plano_manutencao_verificado"),
                1 if request.form.get("historico_oem_verificado") else 0,
                request.form.get("observacoes"),
                now()
            ))

            c.execute(
                "UPDATE processos SET estado = ? WHERE id = ?",
                ("Diagnóstico Mecânico", processo_id)
            )

        uploads(processo_id, "rececao", request.files.getlist("anexos"))

        return redirect(url_for("processo_detail", processo_id=processo_id))

    with db() as c:
        r = c.execute(
            "SELECT * FROM rececoes WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

    return render_template("oficina/rececao.html", p=p, r=r)


@app.route("/processos/<int:processo_id>/diagnostico/<tipo>", methods=["GET", "POST"])
def diagnostico(processo_id, tipo):
    tipo_real = {
        "pre": "Pré",
        "mecanico": "Mecânico",
        "pos": "Pós"
    }.get(tipo)

    p = processo(processo_id)

    if request.method == "POST":
        with db() as c:
            c.execute("""
                INSERT INTO diagnosticos
                (processo_id, viatura_id, tipo, data_diagnostico, km_diagnostico,
                 n_manutencoes_ecu, intervencoes_reais, telecarregamento,
                 ultima_manutencao_registada, manutencao_nao_registada,
                 km_manutencao_anterior, km_ate_manutencao, diluicao_oleo,
                 carbono_oleo, problema_identificado, causa_provavel,
                 intervencao_recomendada, oficina_recomendada, pode_circular,
                 prioridade, inconsistencias, observacoes, criado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                processo_id,
                p["viatura_id"],
                tipo_real,
                request.form.get("data_diagnostico") or now(),
                request.form.get("km_diagnostico") or None,
                request.form.get("n_manutencoes_ecu"),
                request.form.get("intervencoes_reais"),
                request.form.get("telecarregamento"),
                request.form.get("ultima_manutencao_registada"),
                request.form.get("manutencao_nao_registada"),
                request.form.get("km_manutencao_anterior") or None,
                request.form.get("km_ate_manutencao") or None,
                request.form.get("diluicao_oleo") or None,
                request.form.get("carbono_oleo") or None,
                request.form.get("problema_identificado"),
                request.form.get("causa_provavel"),
                request.form.get("intervencao_recomendada"),
                request.form.get("oficina_recomendada"),
                request.form.get("pode_circular"),
                request.form.get("prioridade"),
                request.form.get("inconsistencias"),
                request.form.get("observacoes"),
                now()
            ))

            if tipo_real == "Mecânico":
                c.execute(
                    "UPDATE processos SET estado = ? WHERE id = ?",
                    ("Pendente Validação", processo_id)
                )

            if tipo_real == "Pós":
                c.execute(
                    "UPDATE processos SET estado = ? WHERE id = ?",
                    ("Concluído pelo Mecânico", processo_id)
                )

        uploads(
            processo_id,
            "diagnostico_" + tipo_real.lower(),
            request.files.getlist("anexos")
        )

        return redirect(url_for("processo_detail", processo_id=processo_id))

    return render_template("oficina/diagnostico.html", p=p, tipo=tipo_real)


@app.route("/processos/<int:processo_id>/fo/criar", methods=["POST"])
def criar_fo(processo_id):
    with db() as c:
        c.execute("""
            INSERT OR IGNORE INTO folhas_obra
            (processo_id, numero_fo, estado, responsavel, observacoes, criada_em)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            processo_id,
            request.form.get("numero_fo") or f"FO-{processo_id:04d}",
            "Aberta",
            request.form.get("responsavel"),
            request.form.get("observacoes"),
            now()
        ))

        c.execute(
            "UPDATE processos SET estado = ? WHERE id = ?",
            ("FO Criada", processo_id)
        )

    return redirect(url_for("processo_detail", processo_id=processo_id))


@app.route("/fo/<int:fo_id>/tarefas/nova", methods=["POST"])
def nova_tarefa_fo(fo_id):
    with db() as c:
        fo = c.execute(
            "SELECT * FROM folhas_obra WHERE id = ?",
            (fo_id,)
        ).fetchone()

        c.execute("""
            INSERT INTO tarefas_fo
            (fo_id, descricao, tipo_servico, responsavel, autorizada, estado, observacoes, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            fo_id,
            request.form.get("descricao"),
            request.form.get("tipo_servico"),
            request.form.get("responsavel"),
            1 if request.form.get("autorizada") else 0,
            request.form.get("estado") or "Pendente",
            request.form.get("observacoes"),
            now()
        ))

    return redirect(url_for("processo_detail", processo_id=fo["processo_id"]))


@app.route("/tarefas/<int:tarefa_id>/estado", methods=["POST"])
def atualizar_tarefa(tarefa_id):
    with db() as c:
        t = c.execute("""
            SELECT t.*, f.processo_id
            FROM tarefas_fo t
            JOIN folhas_obra f ON f.id = t.fo_id
            WHERE t.id = ?
        """, (tarefa_id,)).fetchone()

        c.execute(
            "UPDATE tarefas_fo SET estado = ? WHERE id = ?",
            (request.form.get("estado"), tarefa_id)
        )

    return redirect(url_for("processo_detail", processo_id=t["processo_id"]))


@app.route("/processos/<int:processo_id>/iniciar_execucao", methods=["POST"])
def iniciar_execucao(processo_id):
    with db() as c:
        fo = c.execute(
            "SELECT * FROM folhas_obra WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

        if not fo:
            flash("Não é possível iniciar execução: não existe FO criada.", "error")
            return redirect(url_for("processo_detail", processo_id=processo_id))

        n = c.execute(
            "SELECT COUNT(*) c FROM tarefas_fo WHERE fo_id = ? AND autorizada = 1",
            (fo["id"],)
        ).fetchone()["c"]

        if n == 0:
            flash("Não é possível iniciar execução: a FO não tem tarefas autorizadas.", "error")
            return redirect(url_for("processo_detail", processo_id=processo_id))

        c.execute(
            "UPDATE processos SET estado = ? WHERE id = ?",
            ("Em Execução", processo_id)
        )

    return redirect(url_for("processo_detail", processo_id=processo_id))


@app.route("/processos/<int:processo_id>/validacao", methods=["GET", "POST"])
def validacao(processo_id):
    p = processo(processo_id)

    if request.method == "POST":
        vals = [
            1 if request.form.get(k) else 0
            for k in [
                "fo_fechada",
                "diagnostico_pos_anexado",
                "ficha_viatura_atualizada",
                "parametros_atualizados",
                "documentacao_completa"
            ]
        ]

        with db() as c:
            c.execute("""
                INSERT INTO validacoes
                (processo_id, fo_fechada, diagnostico_pos_anexado, ficha_viatura_atualizada,
                 parametros_atualizados, documentacao_completa, observacoes, validado_por, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(processo_id) DO UPDATE SET
                    fo_fechada = excluded.fo_fechada,
                    diagnostico_pos_anexado = excluded.diagnostico_pos_anexado,
                    ficha_viatura_atualizada = excluded.ficha_viatura_atualizada,
                    parametros_atualizados = excluded.parametros_atualizados,
                    documentacao_completa = excluded.documentacao_completa,
                    observacoes = excluded.observacoes,
                    validado_por = excluded.validado_por,
                    atualizado_em = excluded.atualizado_em
            """, (
                processo_id,
                *vals,
                request.form.get("observacoes"),
                request.form.get("validado_por"),
                now()
            ))

            if all(vals):
                c.execute(
                    "UPDATE processos SET estado = ?, data_fecho = ? WHERE id = ?",
                    ("Fechado", now(), processo_id)
                )
                log_action("fecho processo", "processos", processo_id, p["numero_impro"], conn=c)
            else:
                c.execute(
                    "UPDATE processos SET estado = ? WHERE id = ?",
                    ("Validação Administrativa", processo_id)
                )

        return redirect(url_for("processo_detail", processo_id=processo_id))

    with db() as c:
        val = c.execute(
            "SELECT * FROM validacoes WHERE processo_id = ?",
            (processo_id,)
        ).fetchone()

    return render_template("oficina/validacao.html", p=p, val=val)


@app.route("/processos/<int:processo_id>/comparacao")
def comparacao(processo_id):
    p = processo(processo_id)

    with db() as c:
        diags = c.execute("""
            SELECT *
            FROM diagnosticos
            WHERE viatura_id = ?
            ORDER BY data_diagnostico DESC, id DESC
        """, (p["viatura_id"],)).fetchall()

    campos = [
        ("km_diagnostico", "Quilometragem"),
        ("n_manutencoes_ecu", "Nº manutenções ECU"),
        ("intervencoes_reais", "Intervenções reais"),
        ("telecarregamento", "Telecarregamento"),
        ("ultima_manutencao_registada", "Última manutenção registada"),
        ("manutencao_nao_registada", "Manutenção não registada"),
        ("km_manutencao_anterior", "Km manutenção anterior"),
        ("km_ate_manutencao", "Km até manutenção"),
        ("diluicao_oleo", "Diluição óleo"),
        ("carbono_oleo", "Carbono óleo"),
    ]

    return render_template(
        "oficina/comparacao.html",
        p=p,
        diags=diags[:4],
        campos=campos
    )


@app.route("/viaturas")
def viaturas():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "estado": request.args.get("estado", "Ativa").strip(),
        "operacional": request.args.get("operacional", "").strip(),
        "marca": request.args.get("marca", "").strip(),
        "modelo": request.args.get("modelo", "").strip(),
        "grupo": request.args.get("grupo", "").strip(),
        "categoria": request.args.get("categoria", "").strip(),
        "estacao": request.args.get("estacao", "").strip(),
        "km_min": request.args.get("km_min", "").strip(),
        "km_max": request.args.get("km_max", "").strip(),
    }

    where = []
    params = []

    if filtros["q"]:
        where.append("(matricula LIKE ? OR vin LIKE ?)")
        params.extend([f"%{filtros['q']}%", f"%{filtros['q']}%"])
    if filtros["estado"] and filtros["estado"] != "Todas":
        where.append("COALESCE(estado_frota, 'Ativa') = ?")
        params.append(filtros["estado"])
    if filtros["operacional"]:
        where.append("COALESCE(estado_operacional, '') = ?")
        params.append(filtros["operacional"])
    if filtros["marca"]:
        where.append("marca = ?")
        params.append(filtros["marca"])
    if filtros["modelo"]:
        where.append("modelo = ?")
        params.append(filtros["modelo"])
    if filtros["grupo"]:
        where.append("grupo = ?")
        params.append(filtros["grupo"])
    if filtros["categoria"]:
        where.append("categoria = ?")
        params.append(filtros["categoria"])
    if filtros["estacao"]:
        where.append("rental_station = ?")
        params.append(filtros["estacao"])
    if filtros["km_min"]:
        where.append("km_atual >= ?")
        params.append(int(filtros["km_min"]))
    if filtros["km_max"]:
        where.append("km_atual <= ?")
        params.append(int(filtros["km_max"]))

    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        rows = c.execute(
            f"SELECT * FROM viaturas {sql_where} ORDER BY matricula LIMIT 500",
            params
        ).fetchall()

        resumo = c.execute("""
            SELECT
                COUNT(*) total,
                SUM(CASE WHEN COALESCE(estado_frota, 'Ativa') = 'Ativa' THEN 1 ELSE 0 END) ativas,
                SUM(CASE WHEN COALESCE(estado_frota, 'Ativa') = 'Em venda' THEN 1 ELSE 0 END) em_venda,
                SUM(CASE WHEN COALESCE(estado_frota, 'Ativa') = 'Vendida' THEN 1 ELSE 0 END) vendidas,
                SUM(CASE WHEN COALESCE(estado_operacional, '') = 'Em impro' THEN 1 ELSE 0 END) em_impro
            FROM viaturas
        """).fetchone()

        marcas = c.execute("SELECT DISTINCT marca FROM viaturas WHERE marca IS NOT NULL AND marca <> '' ORDER BY marca").fetchall()
        modelos = c.execute("SELECT DISTINCT modelo FROM viaturas WHERE modelo IS NOT NULL AND modelo <> '' ORDER BY modelo").fetchall()
        grupos = c.execute("SELECT DISTINCT grupo FROM viaturas WHERE grupo IS NOT NULL AND grupo <> '' ORDER BY grupo").fetchall()
        categorias = c.execute("SELECT DISTINCT categoria FROM viaturas WHERE categoria IS NOT NULL AND categoria <> '' ORDER BY categoria").fetchall()
        estacoes = c.execute("SELECT DISTINCT rental_station FROM viaturas WHERE rental_station IS NOT NULL AND rental_station <> '' ORDER BY rental_station").fetchall()
        operacionais = c.execute("SELECT DISTINCT estado_operacional FROM viaturas WHERE estado_operacional IS NOT NULL AND estado_operacional <> '' ORDER BY estado_operacional").fetchall()

    return render_template(
        "oficina/viaturas.html",
        viaturas=rows,
        filtros=filtros,
        resumo=resumo,
        marcas=marcas,
        modelos=modelos,
        grupos=grupos,
        categorias=categorias,
        estacoes=estacoes,
        operacionais=operacionais
    )


@app.route("/viaturas/importar", methods=["GET", "POST"])
def importar_frota_view():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona um ficheiro Excel do Rentway.", "error")
            return redirect(url_for("importar_frota_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            stats = import_frota_excel(
                path,
                f.filename,
                request.form.get("responsavel")
            )
        except Exception as exc:
            flash(f"Não foi possível importar o ficheiro: {exc}", "error")
            return redirect(url_for("importar_frota_view"))

        flash(
            "Importação concluída: "
            f"{stats['criadas']} criadas, {stats['atualizadas']} atualizadas, "
            f"{stats['vendidas']} vendidas/históricas, {stats['ignoradas']} ignoradas.",
            "success"
        )
        log_action("importação frota", "importacoes", f.filename, json.dumps(stats, ensure_ascii=False))
        return redirect(url_for("viaturas"))

    with db() as c:
        importacoes = c.execute(
            "SELECT * FROM importacoes ORDER BY id DESC LIMIT 20"
        ).fetchall()

    return render_template("oficina/importar_frota.html", importacoes=importacoes)


@app.route("/impros-rentway")
def impros_rentway():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper(),
        "estacao": request.args.get("estacao", "").strip(),
    }
    where = []
    params = []
    if filtros["q"]:
        where.append("(numero_impro LIKE ? OR matricula LIKE ? OR condutor LIKE ? OR oficina LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 4)
    if filtros["matricula"]:
        where.append("matricula = ?")
        params.append(filtros["matricula"])
    if filtros["estacao"]:
        where.append("estacao_out = ?")
        params.append(filtros["estacao"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        rows = c.execute(f"""
            SELECT *
            FROM rentway_impros
            {sql_where}
            ORDER BY COALESCE(data_fecho, data_in, data_abertura) DESC, id DESC
            LIMIT 500
        """, params).fetchall()
        resumo = c.execute("""
            SELECT
                COUNT(*) total,
                COUNT(DISTINCT matricula) viaturas,
                SUM(CASE WHEN oficina IS NOT NULL AND oficina <> '' THEN 1 ELSE 0 END) com_oficina
            FROM rentway_impros
        """).fetchone()
        estacoes = c.execute("""
            SELECT DISTINCT estacao_out
            FROM rentway_impros
            WHERE estacao_out IS NOT NULL AND estacao_out <> ''
            ORDER BY estacao_out
        """).fetchall()

    return render_template(
        "oficina/impros_rentway.html",
        impros=rows,
        filtros=filtros,
        resumo=resumo,
        estacoes=estacoes
    )


@app.route("/impros-rentway/importar", methods=["GET", "POST"])
def importar_impros_view():
    if request.method == "POST":
        files = [f for f in request.files.getlist("ficheiros") if f and f.filename]
        if not files:
            f = request.files.get("ficheiro")
            files = [f] if f and f.filename else []

        if not files:
            flash("Seleciona um ou mais ficheiros de impros do Rentway.", "error")
            return redirect(url_for("importar_impros_view"))

        total = {"criadas": 0, "atualizadas": 0, "ignoradas": 0, "ficheiros": 0}
        errors = []

        for index, f in enumerate(files, start=1):
            safe = secure_filename(f.filename)
            filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{index}_{safe}"
            path = IMPORT_DIR / filename
            f.save(path)

            try:
                stats = import_impros_excel(
                    path,
                    f.filename,
                    request.form.get("responsavel")
                )
            except Exception as exc:
                errors.append(f"{f.filename}: {exc}")
                continue

            total["ficheiros"] += 1
            total["criadas"] += stats["criadas"]
            total["atualizadas"] += stats["atualizadas"]
            total["ignoradas"] += stats["ignoradas"]

        if errors:
            flash("Alguns ficheiros não foram importados: " + " | ".join(errors), "error")
            if total["ficheiros"] == 0:
                return redirect(url_for("importar_impros_view"))

        flash(
            "Importação de impros concluída: "
            f"{total['ficheiros']} ficheiros, {total['criadas']} criadas, "
            f"{total['atualizadas']} atualizadas, {total['ignoradas']} linhas ignoradas.",
            "success"
        )
        return redirect(url_for("impros_rentway"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Histórico Impros'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_impros.html", importacoes=importacoes)


@app.route("/folhas-obra-rentway")
def folhas_obra_rentway():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper(),
        "fornecedor": request.args.get("fornecedor", "").strip(),
        "natureza": request.args.get("natureza", "").strip(),
    }
    where = []
    params = []
    if filtros["q"]:
        where.append("(folha_obra_origem LIKE ? OR matricula LIKE ? OR fornecedor LIKE ? OR descricao LIKE ? OR numero_fatura LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 5)
    if filtros["matricula"]:
        where.append("matricula = ?")
        params.append(filtros["matricula"])
    if filtros["fornecedor"]:
        where.append("fornecedor = ?")
        params.append(filtros["fornecedor"])
    if filtros["natureza"]:
        where.append("natureza = ?")
        params.append(filtros["natureza"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        rows = c.execute(f"""
            SELECT *
            FROM rentway_folhas_obra
            {sql_where}
            ORDER BY data_documento DESC, id DESC
            LIMIT 500
        """, params).fetchall()
        resumo = c.execute("""
            SELECT
                COUNT(*) total,
                COUNT(DISTINCT folha_obra_origem) total_fo,
                COUNT(DISTINCT matricula) viaturas,
                ROUND(COALESCE(SUM(valor_com_iva), 0), 2) valor_total,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) validar
            FROM rentway_folhas_obra
        """).fetchone()
        fornecedores = c.execute("""
            SELECT DISTINCT fornecedor
            FROM rentway_folhas_obra
            WHERE fornecedor IS NOT NULL AND fornecedor <> ''
            ORDER BY fornecedor
        """).fetchall()
        naturezas = c.execute("""
            SELECT DISTINCT natureza
            FROM rentway_folhas_obra
            WHERE natureza IS NOT NULL AND natureza <> ''
            ORDER BY natureza
        """).fetchall()

    return render_template(
        "oficina/folhas_obra_rentway.html",
        folhas=rows,
        filtros=filtros,
        resumo=resumo,
        fornecedores=fornecedores,
        naturezas=naturezas
    )


@app.route("/folhas-obra-rentway/importar", methods=["GET", "POST"])
def importar_folhas_obra_view():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona o ficheiro de histórico de folhas de obra.", "error")
            return redirect(url_for("importar_folhas_obra_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            stats = import_folhas_obra_excel(
                path,
                f.filename,
                request.form.get("responsavel")
            )
        except Exception as exc:
            flash(f"Não foi possível importar folhas de obra: {exc}", "error")
            return redirect(url_for("importar_folhas_obra_view"))

        flash(
            "Importação de folhas de obra concluída: "
            f"{stats['criadas']} criadas, {stats['atualizadas']} atualizadas, "
            f"{stats['ignoradas']} ignoradas, {stats['alertas']} alertas.",
            "success"
        )
        return redirect(url_for("folhas_obra_rentway"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Histórico FO'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_folhas_obra.html", importacoes=importacoes)


@app.route("/sinistros-allianz")
def sinistros_allianz():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper().replace(" ", ""),
        "causa": request.args.get("causa", "").strip(),
    }
    where = []
    params = []
    if filtros["q"]:
        where.append("(matricula LIKE ? OR referencia_sinistro LIKE ? OR causa LIKE ? OR adesao LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 4)
    if filtros["matricula"]:
        where.append("matricula = ?")
        params.append(filtros["matricula"])
    if filtros["causa"]:
        where.append("causa = ?")
        params.append(filtros["causa"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        rows = c.execute(f"""
            SELECT *
            FROM sinistros_allianz
            {sql_where}
            ORDER BY data_sinistro DESC, id DESC
            LIMIT 500
        """, params).fetchall()
        resumo = c.execute("""
            SELECT
                COUNT(*) total_sinistros,
                COUNT(DISTINCT matricula) viaturas,
                ROUND(COALESCE(SUM(custo_total), 0), 2) custo_total,
                SUM(CASE WHEN data_encerramento IS NULL OR data_encerramento = '' THEN 1 ELSE 0 END) por_encerrar,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) validar
            FROM sinistros_allianz
        """).fetchone()
        causas = c.execute("""
            SELECT DISTINCT causa
            FROM sinistros_allianz
            WHERE causa IS NOT NULL AND causa <> ''
            ORDER BY causa
        """).fetchall()

    return render_template(
        "oficina/sinistros_allianz.html",
        sinistros=rows,
        filtros=filtros,
        resumo=resumo,
        causas=causas
    )


@app.route("/sinistros-allianz/importar", methods=["GET", "POST"])
def importar_sinistros_allianz_view():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona o ficheiro de sinistros Allianz.", "error")
            return redirect(url_for("importar_sinistros_allianz_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            stats = import_sinistros_allianz_excel(
                path,
                f.filename,
                request.form.get("responsavel")
            )
        except Exception as exc:
            flash(f"Não foi possível importar sinistros Allianz: {exc}", "error")
            return redirect(url_for("importar_sinistros_allianz_view"))

        flash(
            "Importação de sinistros Allianz concluída: "
            f"{stats['criadas']} sinistros criados, {stats['atualizadas']} linhas agregadas/atualizadas, "
            f"{stats['ignoradas']} ignoradas.",
            "success"
        )
        return redirect(url_for("sinistros_allianz"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Sinistros Allianz'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_sinistros_allianz.html", importacoes=importacoes)


@app.route("/ars-rentway")
def ars_rentway():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper().replace(" ", ""),
        "document_type": request.args.get("document_type", "").strip(),
        "status": request.args.get("status", "").strip(),
        "estacao": request.args.get("estacao", "").strip(),
    }
    where = []
    params = []
    if filtros["q"]:
        where.append("(accident_report_id LIKE ? OR matricula LIKE ? OR document_no LIKE ? OR insurance_policy LIKE ? OR status LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 5)
    if filtros["matricula"]:
        where.append("matricula = ?")
        params.append(filtros["matricula"])
    if filtros["document_type"]:
        where.append("document_type = ?")
        params.append(filtros["document_type"])
    if filtros["status"]:
        where.append("status = ?")
        params.append(filtros["status"])
    if filtros["estacao"]:
        where.append("rental_station_out = ?")
        params.append(filtros["estacao"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        rows = c.execute(f"""
            SELECT *
            FROM rentway_accident_reports
            {sql_where}
            ORDER BY accident_date DESC, id DESC
            LIMIT 500
        """, params).fetchall()
        resumo = c.execute("""
            SELECT
                COUNT(*) total,
                COUNT(DISTINCT matricula) viaturas,
                SUM(CASE WHEN document_type = 'Rental Agreement' THEN 1 ELSE 0 END) contratos,
                SUM(CASE WHEN document_type = 'Impro' THEN 1 ELSE 0 END) impros,
                SUM(CASE WHEN status LIKE '%ABERTO%' OR status LIKE '%PEND%' OR status LIKE '%CURSO%' THEN 1 ELSE 0 END) em_aberto,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) validar
            FROM rentway_accident_reports
        """).fetchone()
        tipos = c.execute("""
            SELECT DISTINCT document_type
            FROM rentway_accident_reports
            WHERE document_type IS NOT NULL AND document_type <> ''
            ORDER BY document_type
        """).fetchall()
        statuses = c.execute("""
            SELECT DISTINCT status
            FROM rentway_accident_reports
            WHERE status IS NOT NULL AND status <> ''
            ORDER BY status
        """).fetchall()
        estacoes = c.execute("""
            SELECT DISTINCT rental_station_out
            FROM rentway_accident_reports
            WHERE rental_station_out IS NOT NULL AND rental_station_out <> ''
            ORDER BY rental_station_out
        """).fetchall()

    return render_template(
        "oficina/ars_rentway.html",
        ars=rows,
        filtros=filtros,
        resumo=resumo,
        tipos=tipos,
        statuses=statuses,
        estacoes=estacoes
    )


@app.route("/ars-rentway/importar", methods=["GET", "POST"])
def importar_ars_rentway_view():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona o ficheiro de ARs Rentway.", "error")
            return redirect(url_for("importar_ars_rentway_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            stats = import_rentway_accident_reports_excel(
                path,
                f.filename,
                request.form.get("responsavel")
            )
        except Exception as exc:
            flash(f"Não foi possível importar ARs Rentway: {exc}", "error")
            return redirect(url_for("importar_ars_rentway_view"))

        flash(
            "Importação de ARs Rentway concluída: "
            f"{stats['criadas']} criados, {stats['atualizadas']} atualizados, "
            f"{stats['ignoradas']} ignorados, {stats['alertas']} a validar.",
            "success"
        )
        return redirect(url_for("ars_rentway"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Histórico AR Rentway'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_ars_rentway.html", importacoes=importacoes)


@app.route("/faturas-fornecedores")
def faturas_fornecedores():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper().replace(" ", ""),
        "fornecedor": request.args.get("fornecedor", "").strip(),
        "ct": request.args.get("ct", "").strip(),
        "vista": request.args.get("vista", "documento").strip(),
        "ordem": request.args.get("ordem", "data").strip(),
    }
    if filtros["vista"] not in ("documento", "viatura", "fornecedor", "servico"):
        filtros["vista"] = "documento"
    if filtros["ordem"] not in ("data", "valor", "linhas", "nome"):
        filtros["ordem"] = "data"
    where = []
    params = []
    if filtros["q"]:
        where.append("(documento LIKE ? OR matricula LIKE ? OR fornecedor LIKE ? OR descricao LIKE ? OR referencia LIKE ? OR or_reparacao LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 6)
    if filtros["matricula"]:
        where.append("matricula = ?")
        params.append(filtros["matricula"])
    if filtros["fornecedor"]:
        where.append("fornecedor = ?")
        params.append(filtros["fornecedor"])
    if filtros["ct"]:
        where.append("ct = ?")
        params.append(filtros["ct"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""
    doc_key_expr = "COALESCE(fornecedor, '') || '|' || COALESCE(NULLIF(documento, ''), fonte_pdf, '')"
    doc_label_expr = "COALESCE(NULLIF(documento, ''), fonte_pdf, 'Sem documento')"

    if filtros["vista"] == "viatura":
        group_key_expr = "COALESCE(NULLIF(matricula, ''), 'Sem matricula')"
        group_select = f"""
            {group_key_expr} AS group_key,
            {group_key_expr} AS titulo,
            COUNT(DISTINCT {doc_key_expr}) || ' documentos' AS subtitulo,
            MAX(data_doc) AS meta_data,
            GROUP_CONCAT(DISTINCT fornecedor) AS meta_um,
            GROUP_CONCAT(DISTINCT or_reparacao) AS meta_dois
        """
    elif filtros["vista"] == "fornecedor":
        group_key_expr = "COALESCE(NULLIF(fornecedor, ''), 'Sem fornecedor')"
        group_select = f"""
            {group_key_expr} AS group_key,
            {group_key_expr} AS titulo,
            COUNT(DISTINCT {doc_key_expr}) || ' documentos' AS subtitulo,
            MAX(data_doc) AS meta_data,
            COUNT(DISTINCT matricula) || ' viaturas' AS meta_um,
            GROUP_CONCAT(DISTINCT ct) AS meta_dois
        """
    else:
        group_key_expr = doc_key_expr
        group_select = f"""
            {group_key_expr} AS group_key,
            {doc_label_expr} AS titulo,
            fornecedor AS subtitulo,
            MAX(data_doc) AS meta_data,
            GROUP_CONCAT(DISTINCT matricula) AS meta_um,
            GROUP_CONCAT(DISTINCT or_reparacao) AS meta_dois
        """

    order_sql = {
        "data": "meta_data DESC, titulo DESC",
        "valor": "valor_total DESC, titulo ASC",
        "linhas": "linhas DESC, titulo ASC",
        "nome": "titulo ASC",
    }[filtros["ordem"]]

    with db() as c:
        if filtros["vista"] == "servico":
            base_rows = c.execute(f"""
                SELECT *, {doc_key_expr} AS doc_key
                FROM faturas_fornecedores_linhas
                {sql_where}
                ORDER BY data_doc DESC, id ASC
            """, params).fetchall()
            grupos_tmp = {}
            for row in base_rows:
                servicos = detect_supplier_services([row]) or ["Sem sinal"]
                for servico in servicos:
                    grupo = grupos_tmp.setdefault(servico, {"linhas": [], "valor_total": 0.0, "meta_data": None})
                    grupo["linhas"].append(row)
                    grupo["valor_total"] += row["total_liq"] or 0
                    if row["data_doc"] and (not grupo["meta_data"] or row["data_doc"] > grupo["meta_data"]):
                        grupo["meta_data"] = row["data_doc"]

            def sort_key(item):
                nome, grupo = item
                if filtros["ordem"] == "valor":
                    return (-grupo["valor_total"], nome)
                if filtros["ordem"] == "linhas":
                    return (-len(grupo["linhas"]), nome)
                if filtros["ordem"] == "nome":
                    return (nome,)
                return (grupo["meta_data"] or "", nome)

            reverse = filtros["ordem"] == "data"
            documentos = []
            for servico, grupo in sorted(grupos_tmp.items(), key=sort_key, reverse=reverse)[:250]:
                docs = {
                    (linha["fornecedor"] or "") + "|" + (linha["documento"] or linha["fonte_pdf"] or "")
                    for linha in grupo["linhas"]
                }
                matriculas = sorted({linha["matricula"] for linha in grupo["linhas"] if linha["matricula"]})
                resumo_grupo = {
                    "titulo": servico,
                    "subtitulo": f"{len(docs)} documentos",
                    "meta_data": grupo["meta_data"],
                    "meta_um": ", ".join(matriculas[:6]),
                    "meta_dois": "",
                    "linhas": len(grupo["linhas"]),
                    "valor_total": round(grupo["valor_total"], 2),
                }
                documentos.append({
                    "resumo": resumo_grupo,
                    "linhas": grupo["linhas"][:500],
                    "servicos": [servico] if servico != "Sem sinal" else [],
                })
        else:
            doc_rows = c.execute(f"""
            SELECT
                {group_select},
                COUNT(*) AS linhas,
                ROUND(COALESCE(SUM(total_liq), 0), 2) AS valor_total,
                ROUND(COALESCE(SUM(CASE WHEN UPPER(ct) = 'MO' THEN total_liq ELSE 0 END), 0), 2) AS valor_mo,
                ROUND(COALESCE(SUM(CASE WHEN UPPER(ct) = 'MAT' THEN total_liq ELSE 0 END), 0), 2) AS valor_mat,
                MAX(fonte_pdf) AS fonte_pdf,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) AS validar
            FROM faturas_fornecedores_linhas
            {sql_where}
            GROUP BY group_key
            ORDER BY {order_sql}
            LIMIT 250
        """, params).fetchall()
            doc_keys = [row["group_key"] for row in doc_rows]
            linhas_por_doc = {key: [] for key in doc_keys}
            if doc_keys:
                placeholders = ", ".join("?" for _ in doc_keys)
                detail_sql = f"""
                    SELECT *, {group_key_expr} AS group_key
                    FROM faturas_fornecedores_linhas
                    {sql_where}
                    AND {group_key_expr} IN ({placeholders})
                    ORDER BY data_doc DESC, documento DESC, id ASC
                """ if sql_where else f"""
                    SELECT *, {group_key_expr} AS group_key
                    FROM faturas_fornecedores_linhas
                    WHERE {group_key_expr} IN ({placeholders})
                    ORDER BY data_doc DESC, documento DESC, id ASC
                """
                detail_rows = c.execute(detail_sql, params + doc_keys).fetchall()
                for row in detail_rows:
                    linhas_por_doc.setdefault(row["group_key"], []).append(row)
            documentos = [
                {
                    "resumo": row,
                    "linhas": linhas_por_doc.get(row["group_key"], []),
                    "servicos": detect_supplier_services(linhas_por_doc.get(row["group_key"], [])),
                }
                for row in doc_rows
            ]
        resumo = c.execute("""
            SELECT
                COUNT(*) linhas,
                COUNT(DISTINCT COALESCE(fornecedor, '') || '|' || COALESCE(NULLIF(documento, ''), fonte_pdf, '')) documentos,
                COUNT(DISTINCT matricula) viaturas,
                ROUND(COALESCE(SUM(total_liq), 0), 2) valor_total,
                ROUND(COALESCE(SUM(CASE WHEN UPPER(ct) = 'MO' THEN total_liq ELSE 0 END), 0), 2) valor_mo,
                ROUND(COALESCE(SUM(CASE WHEN UPPER(ct) = 'MAT' THEN total_liq ELSE 0 END), 0), 2) valor_mat,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) validar
            FROM faturas_fornecedores_linhas
        """).fetchone()
        fornecedores = c.execute("""
            SELECT DISTINCT fornecedor
            FROM faturas_fornecedores_linhas
            WHERE fornecedor IS NOT NULL AND fornecedor <> ''
            ORDER BY fornecedor
        """).fetchall()
        cts = c.execute("""
            SELECT DISTINCT ct
            FROM faturas_fornecedores_linhas
            WHERE ct IS NOT NULL AND ct <> ''
            ORDER BY ct
        """).fetchall()

    return render_template(
        "oficina/faturas_fornecedores.html",
        documentos=documentos,
        filtros=filtros,
        resumo=resumo,
        fornecedores=fornecedores,
        cts=cts
    )


@app.route("/faturas-fornecedores/importar", methods=["GET", "POST"])
def importar_faturas_fornecedores_view():
    if request.method == "POST":
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona o ficheiro de faturas de fornecedores.", "error")
            return redirect(url_for("importar_faturas_fornecedores_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            stats = import_faturas_fornecedores_excel(
                path,
                f.filename,
                request.form.get("responsavel")
            )
        except Exception as exc:
            flash(f"Não foi possível importar faturas de fornecedores: {exc}", "error")
            return redirect(url_for("importar_faturas_fornecedores_view"))

        flash(
            "Importação de faturas concluída: "
            f"{stats['criadas']} linhas criadas, {stats['atualizadas']} atualizadas, "
            f"{stats['ignoradas']} ignoradas, {stats['alertas']} a validar.",
            "success"
        )
        return redirect(url_for("faturas_fornecedores"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo = 'Faturas Fornecedores'
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_faturas_fornecedores.html", importacoes=importacoes)


@app.route("/contratos-rentway")
def contratos_rentway():
    filtros = {
        "q": request.args.get("q", "").strip(),
        "matricula": request.args.get("matricula", "").strip().upper().replace(" ", ""),
        "cliente": request.args.get("cliente", "").strip(),
        "origem": request.args.get("origem", "").strip(),
    }
    where = []
    params = []
    if filtros["q"]:
        where.append("(cf.contrato_nr LIKE ? OR cf.matricula_final LIKE ? OR cv.matricula LIKE ? OR cv.cliente_nome LIKE ? OR cf.origem LIKE ?)")
        params.extend([f"%{filtros['q']}%"] * 5)
    if filtros["matricula"]:
        where.append("(cf.matricula_final = ? OR cv.matricula = ?)")
        params.extend([filtros["matricula"], filtros["matricula"]])
    if filtros["cliente"]:
        where.append("cv.cliente_nome LIKE ?")
        params.append(f"%{filtros['cliente']}%")
    if filtros["origem"]:
        where.append("cf.origem = ?")
        params.append(filtros["origem"])
    sql_where = "WHERE " + " AND ".join(where) if where else ""

    with db() as c:
        contratos = c.execute(f"""
            SELECT
                cf.*,
                COALESCE(COUNT(DISTINCT cv.id), 0) AS movimentos,
                COALESCE(COUNT(DISTINCT cv.matricula), 0) AS viaturas_reais,
                GROUP_CONCAT(DISTINCT cv.matricula) AS matriculas_reais,
                MAX(cv.cliente_nome) AS cliente_nome
            FROM contratos_financeiro cf
            LEFT JOIN contratos_viaturas cv ON cv.contrato_nr = cf.contrato_nr
            {sql_where}
            GROUP BY cf.id
            ORDER BY cf.data_out DESC, cf.id DESC
            LIMIT 500
        """, params).fetchall()
        contrato_nrs = [row["contrato_nr"] for row in contratos]
        movimentos_por_contrato = {nr: [] for nr in contrato_nrs}
        if contrato_nrs:
            placeholders = ", ".join("?" for _ in contrato_nrs)
            movimentos = c.execute(f"""
                SELECT *
                FROM contratos_viaturas
                WHERE contrato_nr IN ({placeholders})
                ORDER BY contrato_nr, data_out, id
            """, contrato_nrs).fetchall()
            for row in movimentos:
                movimentos_por_contrato.setdefault(row["contrato_nr"], []).append(row)

        resumo = c.execute("""
            SELECT
                COUNT(*) contratos,
                COUNT(DISTINCT matricula_final) viaturas_financeiro,
                ROUND(COALESCE(SUM(valor_total), 0), 2) valor_total,
                ROUND(COALESCE(SUM(dias), 0), 2) dias_total,
                SUM(CASE WHEN valor_total < 0 THEN 1 ELSE 0 END) negativos,
                SUM(CASE WHEN valor_total = 0 THEN 1 ELSE 0 END) total_zero
            FROM contratos_financeiro
        """).fetchone()
        resumo_viaturas = c.execute("""
            SELECT
                COUNT(*) movimentos,
                COUNT(DISTINCT contrato_nr) contratos_operacionais,
                COUNT(DISTINCT matricula) viaturas_operacionais,
                SUM(CASE WHEN estado_importacao = 'VALIDAR' THEN 1 ELSE 0 END) validar
            FROM contratos_viaturas
        """).fetchone()
        origens = c.execute("""
            SELECT DISTINCT origem
            FROM contratos_financeiro
            WHERE origem IS NOT NULL AND origem <> ''
            ORDER BY origem
        """).fetchall()

    return render_template(
        "oficina/contratos_rentway.html",
        contratos=contratos,
        movimentos_por_contrato=movimentos_por_contrato,
        filtros=filtros,
        resumo=resumo,
        resumo_viaturas=resumo_viaturas,
        origens=origens
    )


@app.route("/contratos-rentway/importar", methods=["GET", "POST"])
def importar_contratos_rentway_view():
    if request.method == "POST":
        tipo = request.form.get("tipo")
        f = request.files.get("ficheiro")
        if not f or not f.filename:
            flash("Seleciona o ficheiro de contratos.", "error")
            return redirect(url_for("importar_contratos_rentway_view"))

        safe = secure_filename(f.filename)
        filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{safe}"
        path = IMPORT_DIR / filename
        f.save(path)

        try:
            if tipo == "financeiro":
                stats = import_contratos_financeiro_excel(path, f.filename, request.form.get("responsavel"))
                label = "contratos financeiros"
            elif tipo == "viaturas":
                stats = import_contratos_viaturas_excel(path, f.filename, request.form.get("responsavel"))
                label = "movimentos de viaturas"
            else:
                raise RuntimeError("Tipo de importação inválido.")
        except Exception as exc:
            flash(f"Não foi possível importar contratos: {exc}", "error")
            return redirect(url_for("importar_contratos_rentway_view"))

        flash(
            f"Importação de {label} concluída: "
            f"{stats['criadas']} criados, {stats['atualizadas']} atualizados, "
            f"{stats['ignoradas']} ignorados, {stats['alertas']} a validar.",
            "success"
        )
        return redirect(url_for("contratos_rentway"))

    with db() as c:
        importacoes = c.execute("""
            SELECT *
            FROM importacoes
            WHERE tipo IN ('Contratos Financeiro', 'Contratos Viaturas')
            ORDER BY id DESC
            LIMIT 20
        """).fetchall()

    return render_template("oficina/importar_contratos_rentway.html", importacoes=importacoes)


@app.route("/viaturas/<int:viatura_id>", methods=["GET", "POST"])
def viatura_detail(viatura_id):
    with db() as c:
        if request.method == "POST":
            c.execute("""
                UPDATE viaturas SET
                    vin = ?,
                    marca = ?,
                    modelo = ?,
                    versao = ?,
                    motorizacao = ?,
                    combustivel = ?,
                    caixa = ?,
                    ano = ?,
                    data_compra = ?,
                    km_atual = ?,
                    proxima_revisao_km = ?,
                    proxima_revisao_data = ?,
                    ultima_revisao = ?,
                    campanhas_pendentes = ?,
                    risco_tecnico = ?,
                    observacoes = ?,
                    ativo = ?,
                    estado_frota = ?,
                    estado_operacional = ?,
                    rental_station = ?,
                    grupo = ?,
                    categoria = ?,
                    data_venda = ?,
                    valor_venda = ?,
                    preco_venda_calculado = ?,
                    preco_venda_retalho_calculado = ?,
                    depreciacao_codigo_fiscal_calculado = ?,
                    depreciacao_ultima_calculo_calculado = ?,
                    codigo_conta_calculado = ?
                WHERE id = ?
            """, (
                request.form.get("vin"),
                request.form.get("marca"),
                request.form.get("modelo"),
                request.form.get("versao"),
                request.form.get("motorizacao"),
                request.form.get("combustivel"),
                request.form.get("caixa"),
                request.form.get("ano") or None,
                request.form.get("data_compra"),
                request.form.get("km_atual") or None,
                request.form.get("proxima_revisao_km") or None,
                request.form.get("proxima_revisao_data"),
                request.form.get("ultima_revisao"),
                1 if request.form.get("campanhas_pendentes") else 0,
                request.form.get("risco_tecnico"),
                request.form.get("observacoes"),
                1 if request.form.get("ativo") else 0,
                request.form.get("estado_frota"),
                request.form.get("estado_operacional"),
                request.form.get("rental_station"),
                request.form.get("grupo"),
                request.form.get("categoria"),
                request.form.get("data_venda"),
                request.form.get("valor_venda") or None,
                request.form.get("preco_venda_calculado") or None,
                request.form.get("preco_venda_retalho_calculado") or None,
                request.form.get("depreciacao_codigo_fiscal_calculado"),
                request.form.get("depreciacao_ultima_calculo_calculado"),
                request.form.get("codigo_conta_calculado"),
                viatura_id
            ))

            return redirect(url_for("viatura_detail", viatura_id=viatura_id))

        v = c.execute(
            "SELECT * FROM viaturas WHERE id = ?",
            (viatura_id,)
        ).fetchone()

        processos_rows = c.execute(
            "SELECT * FROM processos WHERE viatura_id = ? ORDER BY id DESC",
            (viatura_id,)
        ).fetchall()

        impros_rentway_rows = c.execute("""
            SELECT *
            FROM rentway_impros
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY COALESCE(data_fecho, data_in, data_abertura) DESC, id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        folhas_obra_rentway_rows = c.execute("""
            SELECT *
            FROM rentway_folhas_obra
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY data_documento DESC, id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        sinistros_allianz_rows = c.execute("""
            SELECT *
            FROM sinistros_allianz
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY data_sinistro DESC, id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        ars_rentway_rows = c.execute("""
            SELECT *
            FROM rentway_accident_reports
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY accident_date DESC, id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        faturas_fornecedores_rows = c.execute("""
            SELECT *
            FROM faturas_fornecedores_linhas
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY data_doc DESC, id DESC
            LIMIT 500
        """, (viatura_id, v["matricula"])).fetchall()

        faturas_docs_tmp = {}
        for linha in faturas_fornecedores_rows:
            doc_key = (
                (linha["fornecedor"] or ""),
                (linha["documento"] or linha["fonte_pdf"] or "Sem documento"),
            )
            doc = faturas_docs_tmp.setdefault(doc_key, {
                "resumo": {
                    "titulo": linha["documento"] or linha["fonte_pdf"] or "Sem documento",
                    "subtitulo": linha["fornecedor"] or "",
                    "meta_data": linha["data_doc"] or "",
                    "meta_um": linha["matricula"] or "",
                    "meta_dois": "",
                    "linhas": 0,
                    "valor_total": 0,
                    "validar": 0,
                },
                "linhas": [],
            })
            doc["linhas"].append(linha)
            doc["resumo"]["linhas"] += 1
            doc["resumo"]["valor_total"] += linha["total_liq"] or 0
            if linha["estado_importacao"] == "VALIDAR":
                doc["resumo"]["validar"] += 1
            if linha["data_doc"] and linha["data_doc"] > (doc["resumo"]["meta_data"] or ""):
                doc["resumo"]["meta_data"] = linha["data_doc"]
            if linha["or_reparacao"]:
                existing_or = doc["resumo"]["meta_dois"]
                ors = {part.strip() for part in existing_or.split(",") if part.strip()}
                ors.add(linha["or_reparacao"])
                doc["resumo"]["meta_dois"] = ", ".join(sorted(ors))

        faturas_fornecedores_docs = []
        for doc in faturas_docs_tmp.values():
            doc["resumo"]["valor_total"] = round(doc["resumo"]["valor_total"], 2)
            doc["servicos"] = detect_supplier_services(doc["linhas"])
            faturas_fornecedores_docs.append(doc)
        faturas_fornecedores_docs.sort(
            key=lambda doc: (doc["resumo"]["meta_data"] or "", doc["resumo"]["titulo"] or ""),
            reverse=True,
        )

        contratos_viaturas_rows = c.execute("""
            SELECT cv.*, cf.valor_total, cf.dias, cf.origem, cf.grupo_reservado, cf.grupo_entregue
            FROM contratos_viaturas cv
            LEFT JOIN contratos_financeiro cf ON cf.contrato_nr = cv.contrato_nr
            WHERE cv.viatura_id = ? OR cv.matricula = ?
            ORDER BY cv.data_out DESC, cv.id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        diags = c.execute("""
            SELECT d.*, p.numero_impro
            FROM diagnosticos d
            JOIN processos p ON p.id = d.processo_id
            WHERE d.viatura_id = ?
            ORDER BY d.data_diagnostico DESC, d.id DESC
        """, (viatura_id,)).fetchall()

        incidentes_frota_rows = c.execute("""
            SELECT i.*, cat.nome AS categoria_nome, cls.nome AS classificacao_nome
            FROM incidentes_frota i
            JOIN incidentes_categorias cat ON cat.id = i.categoria_id
            JOIN incidentes_classificacoes cls ON cls.id = i.classificacao_id
            WHERE i.viatura_id = ? OR i.matricula = ?
            ORDER BY i.criado_em DESC, i.id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

        tarefas_frota_rows = c.execute("""
            SELECT *
            FROM tarefas_frota
            WHERE viatura_id = ? OR matricula = ?
            ORDER BY
              CASE estado WHEN 'Pendente' THEN 1 WHEN 'Em curso' THEN 2 WHEN 'Bloqueada' THEN 3 ELSE 4 END,
              data_limite ASC,
              id DESC
            LIMIT 50
        """, (viatura_id, v["matricula"])).fetchall()

    return render_template(
        "oficina/viatura_detail.html",
        v=v,
        processos=processos_rows,
        impros_rentway=impros_rentway_rows,
        folhas_obra_rentway=folhas_obra_rentway_rows,
        sinistros_allianz=sinistros_allianz_rows,
        ars_rentway=ars_rentway_rows,
        faturas_fornecedores=faturas_fornecedores_rows,
        faturas_fornecedores_docs=faturas_fornecedores_docs,
        contratos_viaturas=contratos_viaturas_rows,
        diagnosticos=diags,
        incidentes_frota=incidentes_frota_rows,
        tarefas_frota=tarefas_frota_rows
    )


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_file(UPLOAD_DIR / filename, as_attachment=True)


try:
    init_db()
    print("CarFast startup: init_db concluido.", flush=True)
except Exception:
    print("CarFast startup: erro no init_db.", flush=True)
    traceback.print_exc()
    raise

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
